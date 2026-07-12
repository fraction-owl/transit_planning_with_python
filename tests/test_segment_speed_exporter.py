from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from scripts.gtfs_exports.segment_speed_exporter import (
    MISSING_VAL,
    band_rows,
    build_index,
    convert_distance,
    export_excel,
    minutes_to_hhmm,
    mph,
    parse_time_to_minutes,
    safe_sheet,
    segment_metrics,
)

FIXTURES = Path(__file__).parent / "fixtures"

# ---------------------------------------------------------------------------
# parse_time_to_minutes
# ---------------------------------------------------------------------------


def test_parse_time_to_minutes_basic() -> None:
    assert parse_time_to_minutes("07:05:00") == 425


def test_parse_time_to_minutes_no_seconds() -> None:
    assert parse_time_to_minutes("07:05") == 425


def test_parse_time_to_minutes_none_returns_none() -> None:
    assert parse_time_to_minutes(None) is None


def test_parse_time_to_minutes_malformed_returns_none() -> None:
    assert parse_time_to_minutes("not-a-time") is None


def test_parse_time_to_minutes_past_midnight() -> None:
    assert parse_time_to_minutes("25:00:00") == 1500


# ---------------------------------------------------------------------------
# minutes_to_hhmm
# ---------------------------------------------------------------------------


def test_minutes_to_hhmm_basic() -> None:
    assert minutes_to_hhmm(425) == "07:05"


def test_minutes_to_hhmm_none_returns_default_missing() -> None:
    assert minutes_to_hhmm(None) == ""


def test_minutes_to_hhmm_none_returns_sentinel() -> None:
    assert minutes_to_hhmm(None, MISSING_VAL) == MISSING_VAL


# ---------------------------------------------------------------------------
# convert_distance
# ---------------------------------------------------------------------------


def test_convert_distance_meters_to_miles() -> None:
    result = convert_distance(1609.344, "meters")
    assert result == pytest.approx(1.0, rel=1e-4)


def test_convert_distance_feet_to_miles() -> None:
    result = convert_distance(5280.0, "feet")
    assert result == pytest.approx(1.0, rel=1e-4)


def test_convert_distance_km_to_miles() -> None:
    result = convert_distance(1.609344, "km")
    assert result == pytest.approx(1.0, rel=1e-4)


def test_convert_distance_none_returns_none() -> None:
    assert convert_distance(None, "meters") is None


def test_convert_distance_empty_string_returns_none() -> None:
    assert convert_distance("", "meters") is None


def test_convert_distance_non_numeric_returns_none() -> None:
    assert convert_distance("abc", "meters") is None


def test_convert_distance_unknown_output_unit_raises() -> None:
    with pytest.raises(ValueError, match="output_unit"):
        convert_distance(1.0, "meters", "leagues")  # type: ignore[arg-type]


# ---------------------------------------------------------------------------
# mph
# ---------------------------------------------------------------------------


def test_mph_basic() -> None:
    assert mph(1.0, 60) == pytest.approx(1.0)


def test_mph_half_hour() -> None:
    assert mph(1.0, 30) == pytest.approx(2.0)


def test_mph_zero_runtime_returns_sentinel() -> None:
    assert mph(1.0, 0) == MISSING_VAL


def test_mph_none_dist_returns_sentinel() -> None:
    assert mph(None, 30) == MISSING_VAL


def test_mph_none_runtime_returns_sentinel() -> None:
    assert mph(1.0, None) == MISSING_VAL


# ---------------------------------------------------------------------------
# safe_sheet
# ---------------------------------------------------------------------------


def test_safe_sheet_normal() -> None:
    assert safe_sheet("Dir_0") == "Dir_0"


def test_safe_sheet_strips_invalid_chars() -> None:
    assert safe_sheet("Dir[0]") == "Dir_0_"


def test_safe_sheet_truncates_to_31() -> None:
    long = "A" * 40
    assert len(safe_sheet(long)) == 31


def test_safe_sheet_empty_becomes_sheet() -> None:
    assert safe_sheet("") == "Sheet"


# ---------------------------------------------------------------------------
# segment_metrics
# ---------------------------------------------------------------------------


def _two_stop_grp() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "stop_id": ["S1", "S2"],
            "stop_sequence": [1, 2],
            "departure_time": ["07:00:00", "07:10:00"],
            "arrival_time": ["07:00:00", "07:10:00"],
            "shape_dist_traveled": ["0", "1609.344"],
        }
    )


def test_segment_metrics_returns_tuple() -> None:
    result = segment_metrics(_two_stop_grp())
    assert isinstance(result, tuple)
    assert len(result) == 3


def test_segment_metrics_speed_positive() -> None:
    seg_speeds, dist_total, runtime = segment_metrics(_two_stop_grp())
    # 1 mile in 10 min = 6 mph
    assert seg_speeds[1] == pytest.approx(6.0, rel=0.01)


def test_segment_metrics_runtime_correct() -> None:
    _, _, runtime = segment_metrics(_two_stop_grp())
    assert runtime == 10


# ---------------------------------------------------------------------------
# build_index — empty after filter
# ---------------------------------------------------------------------------


def _minimal_gtfs() -> dict:
    trips = pd.DataFrame(
        {
            "trip_id": ["T1"],
            "route_id": ["R1"],
            "service_id": ["1"],
            "direction_id": ["0"],
        }
    )
    routes = pd.DataFrame({"route_id": ["R1"], "route_short_name": ["101"]})
    stop_times = pd.DataFrame(
        {
            "trip_id": ["T1", "T1"],
            "stop_id": ["S1", "S2"],
            "stop_sequence": [1, 2],
            "departure_time": ["07:00:00", "07:10:00"],
            "arrival_time": ["07:00:00", "07:10:00"],
            "shape_dist_traveled": ["0", "1609.344"],
        }
    )
    stops = pd.DataFrame(
        {
            "stop_id": ["S1", "S2"],
            "stop_name": ["First", "Second"],
            "stop_code": ["1", "2"],
        }
    )
    return {"trips": trips, "routes": routes, "stop_times": stop_times, "stops": stops}


def test_build_index_returns_dataframe() -> None:
    import scripts.gtfs_exports.segment_speed_exporter as mod

    orig_filter = mod.FILTER_IN_ROUTE_SHORT_NAMES
    orig_service = mod.FILTER_IN_SERVICE_IDS
    mod.FILTER_IN_ROUTE_SHORT_NAMES = []
    mod.FILTER_IN_SERVICE_IDS = []
    try:
        idx, pat_lut, speed_lut, header_lut = build_index(_minimal_gtfs())
        assert isinstance(idx, pd.DataFrame)
    finally:
        mod.FILTER_IN_ROUTE_SHORT_NAMES = orig_filter
        mod.FILTER_IN_SERVICE_IDS = orig_service


def test_build_index_empty_after_filter_returns_empty() -> None:
    import scripts.gtfs_exports.segment_speed_exporter as mod

    orig = mod.FILTER_IN_ROUTE_SHORT_NAMES
    mod.FILTER_IN_ROUTE_SHORT_NAMES = ["NOMATCH"]
    try:
        idx, pat_lut, speed_lut, header_lut = build_index(_minimal_gtfs())
        assert idx.empty
    finally:
        mod.FILTER_IN_ROUTE_SHORT_NAMES = orig


# ---------------------------------------------------------------------------
# band_rows
# ---------------------------------------------------------------------------


def test_band_rows_groups_correctly() -> None:
    idx = pd.DataFrame(
        {
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
            "pattern_hash": [111, 111],
            "speed_hash": [222, 222],
            "start": [420, 480],
        }
    )
    bands = band_rows(idx)
    assert len(bands) == 1
    assert bands.iloc[0]["TripCount"] == 2


def test_band_rows_frtime_is_min() -> None:
    idx = pd.DataFrame(
        {
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
            "pattern_hash": [111, 111],
            "speed_hash": [222, 222],
            "start": [420, 480],
        }
    )
    bands = band_rows(idx)
    # band_rows converts minutes to HH:MM strings via minutes_to_hhmm
    assert bands.iloc[0]["FrTime"] == "07:00"
    assert bands.iloc[0]["ToTime"] == "08:00"


# ---------------------------------------------------------------------------
# export_excel — real openpyxl output
# ---------------------------------------------------------------------------


def test_export_excel_writes_real_workbook(tmp_path: Path, monkeypatch: pytest.MonkeyPatch) -> None:
    import scripts.gtfs_exports.segment_speed_exporter as mod

    monkeypatch.setattr(mod, "OUTPUT_FOLDER", tmp_path)
    monkeypatch.setattr(mod, "FILTER_IN_ROUTE_SHORT_NAMES", [])
    monkeypatch.setattr(mod, "FILTER_IN_SERVICE_IDS", [])

    gtfs = _minimal_gtfs()
    idx, pat_lut, speed_lut, header_lut = build_index(gtfs)
    bands = band_rows(idx)
    export_excel(bands, pat_lut, speed_lut, header_lut, gtfs["routes"])

    out = tmp_path / "route_101_cal1_speed_table.xlsx"
    assert out.exists()

    wb = openpyxl.load_workbook(out)
    assert wb.sheetnames == ["Dir_0"]
    rows = list(wb["Dir_0"].iter_rows(values_only=True))
    assert rows[0][:4] == ("Pattern", "FrTime", "ToTime", "Mean_mph")
    # Single trip: 1 mile in 10 minutes = 6 mph, band spans 07:00 only.
    assert rows[1][1] == "07:00"
    assert rows[1][2] == "07:00"
    assert rows[1][3] == pytest.approx(6.0, rel=0.01)
