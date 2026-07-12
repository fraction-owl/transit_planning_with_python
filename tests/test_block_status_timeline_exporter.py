from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from scripts.gtfs_exports.block_status_timeline_exporter import (
    fill_stop_ids_for_dwell_layover_loading,
    find_cluster,
    get_status_for_minute,
    mark_first_and_last_stops,
    minutes_to_hhmm,
    parse_time_to_minutes,
    process_block,
    validate_folders,
)

# ---------------------------------------------------------------------------
# parse_time_to_minutes
# ---------------------------------------------------------------------------


def test_parse_time_to_minutes_hhmmss() -> None:
    assert parse_time_to_minutes("07:05:00") == 425


def test_parse_time_to_minutes_hhmm_no_seconds() -> None:
    assert parse_time_to_minutes("07:05") == 425


def test_parse_time_to_minutes_past_midnight() -> None:
    assert parse_time_to_minutes("26:30:00") == 1590


def test_parse_time_to_minutes_seconds_rounded() -> None:
    assert parse_time_to_minutes("00:01:30") == 1


# ---------------------------------------------------------------------------
# minutes_to_hhmm
# ---------------------------------------------------------------------------


def test_minutes_to_hhmm_basic() -> None:
    assert minutes_to_hhmm(425) == "07:05"


def test_minutes_to_hhmm_zero() -> None:
    assert minutes_to_hhmm(0) == "00:00"


def test_minutes_to_hhmm_past_midnight() -> None:
    assert minutes_to_hhmm(1590) == "26:30"


# ---------------------------------------------------------------------------
# find_cluster
# ---------------------------------------------------------------------------

_CLUSTERS = [
    {"name": "Hub", "stops": ["100", "101"]},
    {"name": "Terminal", "stops": ["200"]},
]


def test_find_cluster_found() -> None:
    assert find_cluster("100", _CLUSTERS) == "Hub"


def test_find_cluster_second_cluster() -> None:
    assert find_cluster("200", _CLUSTERS) == "Terminal"


def test_find_cluster_not_found() -> None:
    assert find_cluster("999", _CLUSTERS) is None


# ---------------------------------------------------------------------------
# validate_folders
# ---------------------------------------------------------------------------


def test_validate_folders_missing_input_raises(tmp_path: Path) -> None:
    with pytest.raises(NotADirectoryError):
        validate_folders(str(tmp_path / "no_such_dir"), str(tmp_path / "out"))


def test_validate_folders_creates_output(tmp_path: Path) -> None:
    src = tmp_path / "src"
    src.mkdir()
    out = tmp_path / "out" / "nested"
    validate_folders(str(src), str(out))
    assert out.is_dir()


# ---------------------------------------------------------------------------
# mark_first_and_last_stops
# ---------------------------------------------------------------------------


def _make_stop_times() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "trip_id": ["T1", "T1", "T1"],
            "stop_sequence": [1, 2, 3],
            "stop_id": ["S1", "S2", "S3"],
        }
    )


def test_mark_first_and_last_stops_returns_dataframe() -> None:
    df = mark_first_and_last_stops(_make_stop_times())
    assert isinstance(df, pd.DataFrame)


def test_mark_first_and_last_stops_first_flag() -> None:
    df = mark_first_and_last_stops(_make_stop_times())
    assert df.loc[df["stop_sequence"] == 1, "is_first_stop"].all()


def test_mark_first_and_last_stops_last_flag() -> None:
    df = mark_first_and_last_stops(_make_stop_times())
    assert df.loc[df["stop_sequence"] == 3, "is_last_stop"].all()


def test_mark_first_and_last_stops_middle_neither() -> None:
    df = mark_first_and_last_stops(_make_stop_times())
    middle = df[df["stop_sequence"] == 2]
    assert not middle["is_first_stop"].any()
    assert not middle["is_last_stop"].any()


# ---------------------------------------------------------------------------
# get_status_for_minute
# ---------------------------------------------------------------------------

# stop_info tuple: (arr, dep, stop_id, stop_name, trip_id, is_first, is_last, stop_seq, t_val)

_SEQ_SINGLE = [
    (420, 425, "S1", "Main St", "T1", True, False, 1, 0),
    (430, 435, "S2", "Oak Ave", "T1", False, True, 2, 0),
]


def test_get_status_for_minute_empty_sequence() -> None:
    status, *_ = get_status_for_minute(420, [], [])
    assert status == "EMPTY"


def test_get_status_for_minute_depart() -> None:
    # DEPART fires when minute == departure_time and stop is the first stop.
    # First stop has arr=420, dep=425, is_first=True.
    status, *_ = get_status_for_minute(425, _SEQ_SINGLE, [])
    assert status == "DEPART"


def test_get_status_for_minute_arrive() -> None:
    status, *_ = get_status_for_minute(430, _SEQ_SINGLE, [])
    assert status == "ARRIVE"


def test_get_status_for_minute_traveling() -> None:
    status, *_ = get_status_for_minute(427, _SEQ_SINGLE, [])
    assert status == "TRAVELING BETWEEN STOPS"


def test_get_status_for_minute_dwell() -> None:
    seq = [(420, 430, "S1", "Main St", "T1", True, False, 1, 0)]
    status, *_ = get_status_for_minute(425, seq, [])
    assert status == "DWELL"


# ---------------------------------------------------------------------------
# fill_stop_ids_for_dwell_layover_loading
# ---------------------------------------------------------------------------


def _make_status_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "Status": ["ARRIVE", "DWELL", "LAYOVER"],
            "Stop ID": ["S1", "", ""],
            "Stop Name": ["Main St", "", ""],
            "Stop Sequence": [1, "", ""],
            "Arrival Time": ["07:00", "", ""],
            "Departure Time": ["07:05", "", ""],
            "Trip ID": ["T1", "", ""],
        }
    )


def test_fill_dwell_stop_id_propagated() -> None:
    df = fill_stop_ids_for_dwell_layover_loading(_make_status_df())
    assert df.loc[1, "Stop ID"] == "S1"


def test_fill_layover_stop_id_propagated() -> None:
    df = fill_stop_ids_for_dwell_layover_loading(_make_status_df())
    assert df.loc[2, "Stop ID"] == "S1"


def test_fill_returns_dataframe() -> None:
    df = fill_stop_ids_for_dwell_layover_loading(_make_status_df())
    assert isinstance(df, pd.DataFrame)


# ---------------------------------------------------------------------------
# process_block — integration smoke test
# ---------------------------------------------------------------------------


def _make_block_subset() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "trip_id": ["T1", "T1"],
            "stop_sequence": [1, 2],
            "stop_id": ["S1", "S2"],
            "stop_name": ["Stop A", "Stop B"],
            "arrival_min": [420, 425],
            "departure_min": [420, 425],
            "is_first_stop": [True, False],
            "is_last_stop": [False, True],
            "route_id": ["R1", "R1"],
            "direction_id": ["0", "0"],
            "timepoint": [0, 0],
        }
    )


def test_process_block_returns_dataframe() -> None:
    df = process_block(_make_block_subset(), "BLK1", range(415, 430), [])
    assert isinstance(df, pd.DataFrame)


def test_process_block_has_expected_columns() -> None:
    df = process_block(_make_block_subset(), "BLK1", range(415, 430), [])
    for col in ("Timestamp", "Block", "Status"):
        assert col in df.columns, f"Missing column: {col}"


def test_process_block_row_count_matches_timeline() -> None:
    timeline = range(415, 430)
    df = process_block(_make_block_subset(), "BLK1", timeline, [])
    assert len(df) == len(timeline)


# ---------------------------------------------------------------------------
# run_step1_gtfs_to_blocks — full pipeline against gtfs_basic, real xlsx output
# ---------------------------------------------------------------------------


def test_run_step1_writes_real_block_workbooks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import scripts.gtfs_exports.block_status_timeline_exporter as mod

    fixture = Path(__file__).parent / "fixtures" / "gtfs_basic"
    monkeypatch.setattr(mod, "GTFS_FOLDER_PATH", str(fixture))
    monkeypatch.setattr(mod, "BLOCK_OUTPUT_FOLDER", str(tmp_path))
    monkeypatch.setattr(mod, "CALENDAR_SERVICE_IDS", ["WKDY"])

    mod.run_step1_gtfs_to_blocks()

    names = sorted(p.name for p in tmp_path.glob("block_*.xlsx"))
    # Six blocks; B1-B3 interline routes R1 and R2, B4-B6 are R3 only.
    assert names == [
        "block_B1_R1_R2.xlsx",
        "block_B2_R1_R2.xlsx",
        "block_B3_R1_R2.xlsx",
        "block_B4_R3.xlsx",
        "block_B5_R3.xlsx",
        "block_B6_R3.xlsx",
    ]

    wb = openpyxl.load_workbook(tmp_path / "block_B1_R1_R2.xlsx")
    ws = wb.active
    header = [c.value for c in ws[1]]
    for col in ("Timestamp", "Block", "Status"):
        assert col in header, f"Missing column: {col}"
    # One row per minute of the 26-hour timeline.
    assert ws.max_row - 1 == mod.DEFAULT_HOURS * 60
    status_col = header.index("Status") + 1
    statuses = {ws.cell(row=r, column=status_col).value for r in range(2, ws.max_row + 1)}
    assert len(statuses) > 1  # active statuses beyond the inactive filler
