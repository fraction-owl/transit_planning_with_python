from __future__ import annotations

import logging
import shutil
import zipfile
from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

import scripts.field_tools.gtfs_service_change_dates as target

_CAL_HEADER = (
    "service_id,monday,tuesday,wednesday,thursday,friday,saturday,sunday,start_date,end_date"
)


def _write_feed(
    parent: Path,
    name: str,
    calendar_rows=(),
    calendar_dates_rows=(),
    agency_name: str = "Metro Transit",
    timezone: str = "America/New_York",
    feed_info=None,
    extra_files=None,
    zipped: bool = False,
) -> Path:
    feed_dir = parent / name
    feed_dir.mkdir(parents=True)
    if calendar_rows:
        lines = [_CAL_HEADER]
        for sid, days, start, end in calendar_rows:
            lines.append(f"{sid},{','.join(days)},{start},{end}")
        (feed_dir / "calendar.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if calendar_dates_rows:
        lines = ["service_id,date,exception_type"]
        for sid, date, etype in calendar_dates_rows:
            lines.append(f"{sid},{date},{etype}")
        (feed_dir / "calendar_dates.txt").write_text("\n".join(lines) + "\n", encoding="utf-8")
    if agency_name:
        (feed_dir / "agency.txt").write_text(
            "agency_id,agency_name,agency_url,agency_timezone\n"
            f"A1,{agency_name},https://example.com,{timezone}\n",
            encoding="utf-8",
        )
    if feed_info:
        header = ",".join(feed_info)
        values = ",".join(str(v) for v in feed_info.values())
        (feed_dir / "feed_info.txt").write_text(f"{header}\n{values}\n", encoding="utf-8")
    for file_name, content in (extra_files or {}).items():
        (feed_dir / file_name).write_text(content, encoding="utf-8")
    if zipped:
        zip_path = parent / f"{name}.zip"
        with zipfile.ZipFile(zip_path, "w") as archive:
            for member in sorted(feed_dir.iterdir()):
                archive.write(member, arcname=f"{name}/{member.name}")
        shutil.rmtree(feed_dir)
        return zip_path
    return feed_dir


def test_within_feed_change_detected(tmp_path: Path) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "single",
        calendar_rows=[
            ("WKD1", "1111100", "20250106", "20250613"),
            ("WKD2", "1111100", "20250616", "20251226"),
            ("SAT", "0000010", "20250111", "20251227"),
        ],
    )
    out = tmp_path / "out"
    changes, feeds_df = target.run(feeds_dir=feeds, output_dir=out)

    assert list(changes["change_date"]) == ["2025-06-16"]
    row = changes.iloc[0]
    assert row["day_of_week"] == "Monday"
    assert row["change_type"] == "Service change"
    assert row["services_added"] == "WKD2"
    assert row["services_removed"] == "WKD1"

    assert len(feeds_df) == 1
    assert feeds_df.iloc[0]["first_active_date"] == "2025-01-06"
    assert feeds_df.iloc[0]["last_active_date"] == "2025-12-27"
    assert feeds_df.iloc[0]["service_ids"] == 3

    assert (out / "service_change_quick_reference.xlsx").exists()
    runlog = out / "gtfs_service_change_dates_runlog.txt"
    assert runlog.exists()
    assert "# === BEGIN CONFIG ===" in runlog.read_text(encoding="utf-8")


def test_holiday_week_is_transient(tmp_path: Path, caplog) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "holiday",
        calendar_rows=[("WKD", "1111100", "20250106", "20251226")],
        calendar_dates_rows=[("WKD", "20250704", "2"), ("HOL", "20250704", "1")],
    )
    with caplog.at_level(logging.INFO):
        changes, _ = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")
    assert changes.empty
    assert "transient" in caplog.text.lower()


def test_cross_feed_succession_between_zips(tmp_path: Path) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "2025_spring",
        calendar_rows=[("WKD_A", "1111100", "20250106", "20250613")],
        zipped=True,
    )
    _write_feed(
        feeds,
        "2025_fall",
        calendar_rows=[("WKD_B", "1111100", "20250616", "20251226")],
        zipped=True,
    )
    changes, feeds_df = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")

    assert list(changes["change_date"]) == ["2025-06-16"]
    row = changes.iloc[0]
    assert row["change_type"] == "New service period"
    assert row["services_added"] == "WKD_B"
    assert row["services_removed"] == "WKD_A"
    assert "2025_spring" in row["source_feeds"]
    assert "2025_fall" in row["source_feeds"]
    assert len(feeds_df) == 2


def test_overlapping_feeds_disagree(tmp_path: Path, caplog) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "archive_flat",
        calendar_rows=[("WKD_ALL", "1111100", "20250106", "20251226")],
    )
    _write_feed(
        feeds,
        "archive_split",
        calendar_rows=[
            ("WKD1", "1111100", "20250106", "20250613"),
            ("WKD2", "1111100", "20250616", "20251226"),
        ],
    )
    with caplog.at_level(logging.WARNING):
        changes, _ = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")

    assert list(changes["change_date"]) == ["2025-06-16"]
    assert changes.iloc[0]["disputed_by"] == "archive_flat"
    assert "disagree" in caplog.text.lower()


def test_mixed_agencies_warn_about_same_system(tmp_path: Path, caplog) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "metro",
        calendar_rows=[("WKD", "1111100", "20250106", "20250613")],
        agency_name="Metro Transit",
    )
    _write_feed(
        feeds,
        "sunshine",
        calendar_rows=[("WKD", "1111100", "20250616", "20251226")],
        agency_name="Sunshine Shuttles",
        timezone="America/Chicago",
    )
    with caplog.at_level(logging.WARNING):
        _, feeds_df = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")
    assert "same system" in caplog.text.lower()
    assert "agency_timezones" in caplog.text
    assert feeds_df["issues"].str.contains("same system").any()


def test_cutoffs_limit_changes(tmp_path: Path) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "quarterly",
        calendar_rows=[
            ("W1", "1111100", "20250106", "20250328"),
            ("W2", "1111100", "20250331", "20250627"),
            ("W3", "1111100", "20250630", "20250926"),
            ("W4", "1111100", "20250929", "20251226"),
        ],
    )
    out = tmp_path / "out"

    all_changes, _ = target.run(feeds_dir=feeds, output_dir=out)
    assert list(all_changes["change_date"]) == ["2025-03-31", "2025-06-30", "2025-09-29"]

    recent, _ = target.run(feeds_dir=feeds, output_dir=out, max_changes=1)
    assert list(recent["change_date"]) == ["2025-09-29"]

    # Half a year back from the newest active date (2025-12-26) is 2025-06-26.
    within, _ = target.run(feeds_dir=feeds, output_dir=out, max_years=0.5)
    assert list(within["change_date"]) == ["2025-06-30", "2025-09-29"]


def test_folder_that_is_itself_a_feed(tmp_path: Path) -> None:
    feed = _write_feed(
        tmp_path, "rootfeed", calendar_rows=[("WKD", "1111100", "20250106", "20251226")]
    )
    changes, feeds_df = target.run(feeds_dir=feed, output_dir=tmp_path / "out")
    assert changes.empty
    assert len(feeds_df) == 1
    assert feeds_df.iloc[0]["feed"] == "rootfeed"


def test_unusable_feed_is_listed_with_issue(tmp_path: Path, caplog) -> None:
    feeds = tmp_path / "archive"
    _write_feed(feeds, "good", calendar_rows=[("WKD", "1111100", "20250106", "20251226")])
    _write_feed(
        feeds,
        "no_dates",
        extra_files={"trips.txt": "route_id,service_id,trip_id\nR1,A,T1\n"},
    )
    with caplog.at_level(logging.WARNING):
        _, feeds_df = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")
    assert len(feeds_df) == 2
    bad = feeds_df[feeds_df["feed"] == "no_dates"].iloc[0]
    assert "calendar" in bad["issues"]
    assert bad["first_active_date"] == ""
    assert "no_dates" in caplog.text


def test_feed_info_disagreement_warns(tmp_path: Path, caplog) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "short_calendar",
        calendar_rows=[("WKD", "1111100", "20250106", "20250613")],
        feed_info={
            "feed_publisher_name": "Metro Transit",
            "feed_publisher_url": "https://example.com",
            "feed_lang": "en",
            "feed_start_date": "20250101",
            "feed_end_date": "20251231",
            "feed_version": "2025.1",
        },
    )
    with caplog.at_level(logging.WARNING):
        _, feeds_df = target.run(feeds_dir=feeds, output_dir=tmp_path / "out")
    row = feeds_df.iloc[0]
    assert row["feed_version"] == "2025.1"
    assert row["declared_end"] == "2025-12-31"
    assert "feed_info declares service through" in row["issues"]
    assert "feed_info declares service through" in caplog.text


def test_csv_outputs_are_machine_readable(tmp_path: Path) -> None:
    feeds = tmp_path / "archive"
    calendar_rows = [("WKD0", "1111100", "20250106", "20250613")] + [
        (f"W_{letter}", "1111100", "20250616", "20251226") for letter in "ABCDEFG"
    ]
    _write_feed(feeds, "many_services", calendar_rows=calendar_rows)
    out = tmp_path / "out"
    changes, feeds_df = target.run(feeds_dir=feeds, output_dir=out)

    # The CSV carries the full, untruncated service_id list.
    csv_changes = pd.read_csv(out / "service_changes.csv")
    assert list(csv_changes["change_date"]) == ["2025-06-16"]
    assert csv_changes.iloc[0]["services_added"] == "W_A; W_B; W_C; W_D; W_E; W_F; W_G"
    assert csv_changes.iloc[0]["services_removed"] == "WKD0"
    assert list(csv_changes.columns) == list(changes.columns)

    # The printable XLSX shortens the same list for display.
    workbook = load_workbook(out / "service_change_quick_reference.xlsx")
    assert "+1 more" in str(workbook["Service Changes"]["D4"].value)

    csv_feeds = pd.read_csv(out / "service_change_feeds.csv")
    assert list(csv_feeds["feed"]) == list(feeds_df["feed"])
    assert csv_feeds.iloc[0]["service_ids"] == 8


def test_empty_folder_raises(tmp_path: Path) -> None:
    empty = tmp_path / "empty"
    empty.mkdir()
    with pytest.raises(ValueError, match="No GTFS feeds"):
        target.run(feeds_dir=empty, output_dir=tmp_path / "out")


def test_main_placeholder_paths_return_2() -> None:
    assert target.main([]) == 2


def test_main_missing_folder_returns_1(tmp_path: Path) -> None:
    rc = target.main(["--feeds-dir", str(tmp_path / "nope"), "--output-dir", str(tmp_path / "out")])
    assert rc == 1


def test_main_runs_end_to_end(tmp_path: Path) -> None:
    feeds = tmp_path / "archive"
    _write_feed(
        feeds,
        "single",
        calendar_rows=[
            ("WKD1", "1111100", "20250106", "20250613"),
            ("WKD2", "1111100", "20250616", "20251226"),
        ],
    )
    out = tmp_path / "out"
    rc = target.main(["--feeds-dir", str(feeds), "--output-dir", str(out), "--max-changes", "5"])
    assert rc == 0

    workbook = load_workbook(out / "service_change_quick_reference.xlsx")
    assert workbook.sheetnames == ["Service Changes", "Feeds"]
    sheet = workbook["Service Changes"]
    assert str(sheet["A1"].value).startswith("Service Change Quick Reference")
    assert sheet["A4"].value == "2025-06-16"
