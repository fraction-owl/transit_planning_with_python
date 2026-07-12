from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from scripts.gtfs_exports.time_band_exporter import (
    MISSING_TIME,
    build_index,
    export_excel,
    load_gtfs,
    make_bands,
    minutes_to_hhmm,
    parse_time_to_minutes,
    safe_sheet,
    segment_runtimes,
)

FIXTURES = Path(__file__).parent / "fixtures"

# ---------------------------------------------------------------------------
# parse_time_to_minutes
# ---------------------------------------------------------------------------


def test_parse_time_to_minutes_basic() -> None:
    assert parse_time_to_minutes("07:05:00") == 425


def test_parse_time_to_minutes_no_seconds() -> None:
    assert parse_time_to_minutes("07:05") == 425


def test_parse_time_to_minutes_past_midnight() -> None:
    assert parse_time_to_minutes("25:00:00") == 1500


def test_parse_time_to_minutes_none_returns_none() -> None:
    assert parse_time_to_minutes(None) is None


def test_parse_time_to_minutes_malformed_returns_none() -> None:
    assert parse_time_to_minutes("bad") is None


# ---------------------------------------------------------------------------
# minutes_to_hhmm
# ---------------------------------------------------------------------------


def test_minutes_to_hhmm_basic() -> None:
    assert minutes_to_hhmm(425) == "07:05"


def test_minutes_to_hhmm_midnight() -> None:
    assert minutes_to_hhmm(0) == "00:00"


def test_minutes_to_hhmm_none_returns_sentinel() -> None:
    assert minutes_to_hhmm(None, MISSING_TIME) == MISSING_TIME


# ---------------------------------------------------------------------------
# safe_sheet
# ---------------------------------------------------------------------------


def test_safe_sheet_normal_name() -> None:
    assert safe_sheet("Dir_0") == "Dir_0"


def test_safe_sheet_removes_invalid_chars() -> None:
    result = safe_sheet("Route[101]")
    assert "[" not in result
    assert "]" not in result


def test_safe_sheet_truncates_long_name() -> None:
    assert len(safe_sheet("X" * 50)) == 31


def test_safe_sheet_empty_becomes_sheet() -> None:
    assert safe_sheet("") == "Sheet"


# ---------------------------------------------------------------------------
# segment_runtimes
# ---------------------------------------------------------------------------


def _two_stop_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "stop_id": ["S1", "S2"],
            "stop_sequence": [1, 2],
            "departure_time": ["07:00:00", "07:10:00"],
            "arrival_time": ["07:00:00", "07:10:00"],
        }
    )


def test_segment_runtimes_first_element_is_sentinel() -> None:
    segs = segment_runtimes(_two_stop_df())
    assert segs[0] == MISSING_TIME


def test_segment_runtimes_correct_runtime() -> None:
    segs = segment_runtimes(_two_stop_df())
    assert segs[1] == 10


def test_segment_runtimes_length() -> None:
    segs = segment_runtimes(_two_stop_df())
    assert len(segs) == 2


def test_segment_runtimes_missing_time_yields_empty_string() -> None:
    df = pd.DataFrame(
        {
            "stop_id": ["S1", "S2"],
            "stop_sequence": [1, 2],
            "departure_time": [None, "07:10:00"],
            "arrival_time": [None, "07:10:00"],
        }
    )
    segs = segment_runtimes(df)
    assert segs[1] == ""


# ---------------------------------------------------------------------------
# load_gtfs
# ---------------------------------------------------------------------------


def test_load_gtfs_missing_file_raises(tmp_path: Path) -> None:
    with pytest.raises(FileNotFoundError):
        load_gtfs(tmp_path)


def test_load_gtfs_loads_basic_fixture() -> None:
    data = load_gtfs(FIXTURES / "gtfs_basic")
    assert "trips" in data
    assert "stop_times" in data
    assert "routes" in data
    assert "stops" in data


def test_load_gtfs_stop_sequence_is_numeric() -> None:
    data = load_gtfs(FIXTURES / "gtfs_basic")
    assert pd.api.types.is_numeric_dtype(data["stop_times"]["stop_sequence"])


# ---------------------------------------------------------------------------
# build_index
# ---------------------------------------------------------------------------


def _minimal_gtfs() -> dict:
    trips = pd.DataFrame(
        {
            "trip_id": ["T1", "T1"],
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
        }
    ).drop_duplicates(subset="trip_id")
    trips = pd.DataFrame(
        {
            "trip_id": ["T1"],
            "route_id": ["R1"],
            "service_id": ["1"],
            "direction_id": ["0"],
        }
    )
    routes = pd.DataFrame({"route_id": ["R1"], "route_short_name": ["101"]})
    stop_times = pd.DataFrame(
        {
            "trip_id": ["T1", "T1"],
            "stop_id": ["S1", "S2"],
            "stop_sequence": [1, 2],
            "departure_time": ["07:00:00", "07:10:00"],
            "arrival_time": ["07:00:00", "07:10:00"],
        }
    )
    stops = pd.DataFrame(
        {
            "stop_id": ["S1", "S2"],
            "stop_name": ["First", "Second"],
            "stop_code": ["001", "002"],
        }
    )
    return {"trips": trips, "routes": routes, "stop_times": stop_times, "stops": stops}


def test_build_index_returns_dataframe() -> None:
    import scripts.gtfs_exports.time_band_exporter as mod

    orig_in = mod.FILTER_IN_ROUTE_SHORT_NAMES
    orig_service = mod.FILTER_IN_SERVICE_IDS
    mod.FILTER_IN_ROUTE_SHORT_NAMES = []
    mod.FILTER_IN_SERVICE_IDS = []
    try:
        idx, _, _, _ = build_index(_minimal_gtfs())
        assert isinstance(idx, pd.DataFrame)
    finally:
        mod.FILTER_IN_ROUTE_SHORT_NAMES = orig_in
        mod.FILTER_IN_SERVICE_IDS = orig_service


def test_build_index_empty_when_filtered_out() -> None:
    import scripts.gtfs_exports.time_band_exporter as mod

    orig = mod.FILTER_IN_ROUTE_SHORT_NAMES
    mod.FILTER_IN_ROUTE_SHORT_NAMES = ["NOMATCH"]
    try:
        idx, _, _, _ = build_index(_minimal_gtfs())
        assert idx.empty
    finally:
        mod.FILTER_IN_ROUTE_SHORT_NAMES = orig


# ---------------------------------------------------------------------------
# make_bands
# ---------------------------------------------------------------------------


def test_make_bands_one_group() -> None:
    idx = pd.DataFrame(
        {
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
            "pattern_hash": [100, 100],
            "seg_hash": [200, 200],
            "start": [420, 480],
        }
    )
    bands = make_bands(idx)
    assert len(bands) == 1


def test_make_bands_frtime_totime() -> None:
    idx = pd.DataFrame(
        {
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
            "pattern_hash": [100, 100],
            "seg_hash": [200, 200],
            "start": [420, 480],
        }
    )
    bands = make_bands(idx)
    assert bands.iloc[0]["FrTime"] == 420
    assert bands.iloc[0]["ToTime"] == 480
    assert bands.iloc[0]["Total"] == 2


def test_make_bands_two_patterns_two_rows() -> None:
    idx = pd.DataFrame(
        {
            "route_id": ["R1", "R1"],
            "service_id": ["1", "1"],
            "direction_id": ["0", "0"],
            "pattern_hash": [100, 101],
            "seg_hash": [200, 201],
            "start": [420, 480],
        }
    )
    bands = make_bands(idx)
    assert len(bands) == 2


# ---------------------------------------------------------------------------
# export_excel — full pipeline against the gtfs_basic fixture, real openpyxl
# ---------------------------------------------------------------------------


def test_export_excel_full_pipeline_writes_real_workbooks(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    import scripts.gtfs_exports.time_band_exporter as mod

    monkeypatch.setattr(mod, "OUTPUT_FOLDER", tmp_path)

    data = load_gtfs(FIXTURES / "gtfs_basic")
    idx, stop_dict, seg_dict, header_names = build_index(data)
    bands = make_bands(idx)
    export_excel(bands, stop_dict, seg_dict, header_names, data["routes"])

    names = sorted(p.name for p in tmp_path.glob("*.xlsx"))
    assert names == [
        "route_R1_calWKDY_timeband_table.xlsx",
        "route_R2_calWKDY_timeband_table.xlsx",
        "route_R3_calWKDY_timeband_table.xlsx",
    ]

    wb = openpyxl.load_workbook(tmp_path / "route_R1_calWKDY_timeband_table.xlsx")
    assert wb.sheetnames == ["Dir_0"]
    rows = list(wb["Dir_0"].iter_rows(values_only=True))
    assert rows[0][:4] == ("Pattern", "FrTime", "ToTime", "Total")
    # Header continues with the pattern's stop names and codes.
    assert rows[0][4] == "Main St & Mt Vernon Ln (1001)"
    # R1's three trips share one pattern and identical runtimes → one band
    # from the 07:00 trip to the 17:00 trip with 5-minute segments.
    assert len(rows) == 2
    assert rows[1][1:4] == ("07:00", "17:00", 3)
    assert rows[1][5] == 5
