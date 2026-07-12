from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd

from scripts.gtfs_exports.timepoint_schedule_exporter import (
    MAX_COLUMN_WIDTH,
    MISSING_TIME,
    adjust_time,
    apply_in_out_filters,
    build_service_id_schedule_map,
    export_to_excel_multiple_sheets,
    format_output_folder_name,
    get_all_route_short_names,
    map_service_id_to_schedule,
    prepare_timepoints,
    remove_empty_schedule_columns,
    time_to_minutes,
)

# ---------------------------------------------------------------------------
# time_to_minutes
# ---------------------------------------------------------------------------


def test_time_to_minutes_24h() -> None:
    assert time_to_minutes("08:30") == 510


def test_time_to_minutes_midnight() -> None:
    assert time_to_minutes("00:00") == 0


def test_time_to_minutes_past_midnight() -> None:
    assert time_to_minutes("25:00") == 1500


def test_time_to_minutes_sentinel_returns_none() -> None:
    assert time_to_minutes(MISSING_TIME) is None


def test_time_to_minutes_non_string_returns_none() -> None:
    assert time_to_minutes(None) is None  # type: ignore[arg-type]


def test_time_to_minutes_am_pm() -> None:
    assert time_to_minutes("8:30 AM") == 510


def test_time_to_minutes_pm() -> None:
    assert time_to_minutes("1:00 PM") == 780


# ---------------------------------------------------------------------------
# adjust_time
# ---------------------------------------------------------------------------


def test_adjust_time_24h_format() -> None:
    assert adjust_time("8:05:00", "24") == "08:05"


def test_adjust_time_12h_am() -> None:
    assert adjust_time("08:05", "12") == "8:05 AM"


def test_adjust_time_12h_pm() -> None:
    assert adjust_time("13:00", "12") == "1:00 PM"


def test_adjust_time_none_returns_none() -> None:
    assert adjust_time(None, "24") is None  # type: ignore[arg-type]


def test_adjust_time_sentinel_returns_sentinel() -> None:
    assert adjust_time(MISSING_TIME, "24") == MISSING_TIME


# ---------------------------------------------------------------------------
# prepare_timepoints
# ---------------------------------------------------------------------------


def _stop_times_with_timepoint() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "trip_id": ["T1", "T1", "T1"],
            "stop_sequence": ["1", "2", "3"],
            "stop_id": ["S1", "S2", "S3"],
            "timepoint": ["1", "0", "1"],
        }
    )


def test_prepare_timepoints_filters_to_one() -> None:
    df = prepare_timepoints(_stop_times_with_timepoint())
    assert len(df) == 2


def test_prepare_timepoints_all_rows_returned_when_no_column() -> None:
    df = pd.DataFrame(
        {
            "trip_id": ["T1", "T1"],
            "stop_sequence": ["1", "2"],
            "stop_id": ["S1", "S2"],
        }
    )
    result = prepare_timepoints(df)
    assert len(result) == 2


def test_prepare_timepoints_stop_sequence_is_numeric() -> None:
    df = prepare_timepoints(_stop_times_with_timepoint())
    assert pd.api.types.is_numeric_dtype(df["stop_sequence"])


# ---------------------------------------------------------------------------
# remove_empty_schedule_columns
# ---------------------------------------------------------------------------


def test_remove_empty_schedule_columns_drops_blank_cols() -> None:
    df = pd.DataFrame(
        {
            "Stop A Schedule": [MISSING_TIME, MISSING_TIME],
            "Stop B Schedule": ["07:05", MISSING_TIME],
            "Route Name": ["101", "101"],
        }
    )
    result = remove_empty_schedule_columns(df)
    assert "Stop A Schedule" not in result.columns
    assert "Stop B Schedule" in result.columns


def test_remove_empty_schedule_columns_preserves_non_schedule_cols() -> None:
    df = pd.DataFrame(
        {
            "Route Name": ["101"],
            "Stop A Schedule": [MISSING_TIME],
        }
    )
    result = remove_empty_schedule_columns(df)
    assert "Route Name" in result.columns


# ---------------------------------------------------------------------------
# map_service_id_to_schedule
# ---------------------------------------------------------------------------


def _cal_row(days: dict) -> pd.Series:
    base = {
        d: "0"
        for d in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    }
    base.update(days)
    return pd.Series(base)


def test_map_weekday() -> None:
    row = _cal_row(
        {"monday": "1", "tuesday": "1", "wednesday": "1", "thursday": "1", "friday": "1"}
    )
    assert map_service_id_to_schedule(row) == "Weekday"


def test_map_saturday() -> None:
    row = _cal_row({"saturday": "1"})
    assert map_service_id_to_schedule(row) == "Saturday"


def test_map_sunday() -> None:
    row = _cal_row({"sunday": "1"})
    assert map_service_id_to_schedule(row) == "Sunday"


def test_map_weekend() -> None:
    row = _cal_row({"saturday": "1", "sunday": "1"})
    assert map_service_id_to_schedule(row) == "Weekend"


def test_map_daily() -> None:
    row = _cal_row(
        {
            d: "1"
            for d in ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
        }
    )
    assert map_service_id_to_schedule(row) == "Daily"


def test_map_no_days_returns_holiday() -> None:
    row = _cal_row({})
    assert map_service_id_to_schedule(row) == "Holiday"


# ---------------------------------------------------------------------------
# format_output_folder_name
# ---------------------------------------------------------------------------


def test_format_output_folder_name_weekday() -> None:
    name = format_output_folder_name("1", "Weekday")
    assert name == "weekday_sid_1"


def test_format_output_folder_name_saturday() -> None:
    name = format_output_folder_name("2", "Saturday")
    assert name == "saturday_sid_2"


# ---------------------------------------------------------------------------
# get_all_route_short_names
# ---------------------------------------------------------------------------


def test_get_all_route_short_names_sorted() -> None:
    routes = pd.DataFrame({"route_short_name": ["202", "101", "303"]})
    assert get_all_route_short_names(routes) == ["101", "202", "303"]


def test_get_all_route_short_names_drops_na() -> None:
    routes = pd.DataFrame({"route_short_name": ["101", None, "202"]})
    result = get_all_route_short_names(routes)
    assert None not in result
    assert len(result) == 2


# ---------------------------------------------------------------------------
# apply_in_out_filters
# ---------------------------------------------------------------------------


def test_apply_in_out_filters_filter_in() -> None:
    import scripts.gtfs_exports.timepoint_schedule_exporter as mod

    orig_in = mod.FILTER_IN_ROUTES
    orig_out = mod.FILTER_OUT_ROUTES
    mod.FILTER_IN_ROUTES = ["101"]
    mod.FILTER_OUT_ROUTES = []
    try:
        result = apply_in_out_filters(["101", "202", "303"])
        assert result == ["101"]
    finally:
        mod.FILTER_IN_ROUTES = orig_in
        mod.FILTER_OUT_ROUTES = orig_out


def test_apply_in_out_filters_filter_out() -> None:
    import scripts.gtfs_exports.timepoint_schedule_exporter as mod

    orig_in = mod.FILTER_IN_ROUTES
    orig_out = mod.FILTER_OUT_ROUTES
    mod.FILTER_IN_ROUTES = []
    mod.FILTER_OUT_ROUTES = ["202"]
    try:
        result = apply_in_out_filters(["101", "202", "303"])
        assert "202" not in result
        assert len(result) == 2
    finally:
        mod.FILTER_IN_ROUTES = orig_in
        mod.FILTER_OUT_ROUTES = orig_out


# ---------------------------------------------------------------------------
# build_service_id_schedule_map
# ---------------------------------------------------------------------------


def test_build_service_id_schedule_map_basic() -> None:
    import scripts.gtfs_exports.timepoint_schedule_exporter as mod

    orig = mod.SERVICE_LABEL_OVERRIDES
    mod.SERVICE_LABEL_OVERRIDES = {}
    cal = pd.DataFrame(
        {
            "service_id": ["1"],
            "monday": ["1"],
            "tuesday": ["1"],
            "wednesday": ["1"],
            "thursday": ["1"],
            "friday": ["1"],
            "saturday": ["0"],
            "sunday": ["0"],
        }
    )
    try:
        mapping = build_service_id_schedule_map(cal)
        assert mapping["1"] == "Weekday"
    finally:
        mod.SERVICE_LABEL_OVERRIDES = orig


def test_build_service_id_schedule_map_override_takes_precedence() -> None:
    import scripts.gtfs_exports.timepoint_schedule_exporter as mod

    orig = mod.SERVICE_LABEL_OVERRIDES
    mod.SERVICE_LABEL_OVERRIDES = {"1": "CustomLabel"}
    cal = pd.DataFrame(
        {
            "service_id": ["1"],
            "monday": ["1"],
            "tuesday": ["1"],
            "wednesday": ["1"],
            "thursday": ["1"],
            "friday": ["1"],
            "saturday": ["0"],
            "sunday": ["0"],
        }
    )
    try:
        mapping = build_service_id_schedule_map(cal)
        assert mapping["1"] == "CustomLabel"
    finally:
        mod.SERVICE_LABEL_OVERRIDES = orig


# ---------------------------------------------------------------------------
# export_to_excel_multiple_sheets — real openpyxl output
# ---------------------------------------------------------------------------


def test_export_to_excel_multiple_sheets_writes_real_workbook(tmp_path: Path) -> None:
    out = tmp_path / "schedule.xlsx"
    df = pd.DataFrame({"Route": ["R1", "R1"], "Main St & Doe St": ["07:10", "12:10"]})

    export_to_excel_multiple_sheets({"Weekday_0": df, "Saturday_0": pd.DataFrame()}, str(out))

    wb = openpyxl.load_workbook(out)
    # The empty DataFrame's sheet is skipped, the populated one written.
    assert wb.sheetnames == ["Weekday_0"]
    ws = wb["Weekday_0"]
    rows = list(ws.iter_rows(values_only=True))
    assert rows[0] == ("Route", "Main St & Doe St")
    assert rows[1] == ("R1", "07:10")
    assert rows[2] == ("R1", "12:10")
    # Styling applied by the real export path.
    assert ws["A1"].alignment.wrap_text is True
    assert ws["A2"].alignment.horizontal == "left"
    assert ws.column_dimensions["A"].width <= MAX_COLUMN_WIDTH


def test_export_to_excel_multiple_sheets_empty_dict_writes_nothing(tmp_path: Path) -> None:
    out = tmp_path / "schedule.xlsx"
    export_to_excel_multiple_sheets({}, str(out))
    assert not out.exists()
