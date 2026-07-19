"""Allocate stop-level ridership to districts and report each district's share.

This script joins stop-level ridership (a Ridecheck-style Excel export) to GTFS
stop locations, determines which district contains each stop via an ArcPy
point-in-polygon spatial join, and writes an Excel workbook of total and
percent ridership by district. A reconciliation sheet accounts for every
unallocated unit — ridership at stops outside all districts, STOP_IDs with no
GTFS match, and (under "full" boundary allocation) ridership double-counted
across districts — so discrepancies are auditable rather than silently hidden.

A stop exactly coincident with a shared district boundary joins every touching
district; BOUNDARY_ALLOCATION decides whether its ridership is divided evenly
among them ("split", the default — district totals reconcile exactly) or
counted fully in each ("full" — a boundary-as-inside convention that can push
summed percentages above 100). Boardings, alightings, and their sum all carry
through with separate percentage columns; boardings is the conventional
"ridership" figure, while boardings + alightings is stop *activity* and
double-counts riders. Percent columns default to shares of the allocated
total; PCT_BASE can widen the denominator to include out-of-district or
unmatched ridership. A GeoPandas twin, ``district_ridership_share_gpd.py``,
runs without an ArcGIS license.

Inputs
------
- GTFS feed (a folder, a stops.txt path, or a .zip archive): stop coordinates
  and the join key (``stop_id`` or ``stop_code``, per GTFS_JOIN_KEY).
- Stop-level ridership Excel; column defaults match the Ridecheck
  RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS) export, the same file
  consumed by ``data_request_by_stop_processor.py``.
- District polygon layer (e.g. a shapefile) with a district identifier field.

Outputs
-------
- ``district_ridership_share.xlsx`` — sheets ``district_ridership`` (totals
  and percent shares per district), ``ridership_allocation`` (per
  stop-district allocation detail), and ``reconciliation`` (allocated vs.
  outside-district vs. unmatched totals).
- ``district_ridership_share_runlog.txt`` — run-log sidecar capturing the
  verbatim CONFIGURATION block, the effective settings, and SHA-256
  fingerprints of the inputs.
- ``district_ridership_share_arcpy.log`` in ``LOG_DIR`` mirroring console
  output.

Typical usage
-------------
Update the paths in the CONFIGURATION section (or pass the matching CLI
flags) and run from a shell, ArcGIS Pro's Python window, or a Jupyter
notebook (requires ArcGIS Pro's ``arcpy``).
"""

from __future__ import annotations

import argparse
import hashlib
import logging
import os
import sys
import uuid
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Sequence

import arcpy
import pandas as pd

# Path to this script's own source. Undefined in a Jupyter kernel (no __file__),
# so fall back to None; the run log handles a missing source gracefully.
try:
    SELF_PATH: Optional[Path] = Path(__file__).resolve()
except NameError:
    SELF_PATH = None

# =============================================================================
# CONFIGURATION
# =============================================================================
# === BEGIN CONFIG ===

# INPUTS ----------------------------------------------------------------------
# GTFS feed: a folder, a stops.txt path, or a .zip archive of the feed.
GTFS_INPUT = r"Path\To\Your\GTFS_data"

# Stop-level ridership Excel. Column names are configurable below; defaults
# match the Ridecheck RIDERSHIP_BY_ROUTE_AND_STOP_(ALL_TIME_PERIODS) export.
EXCEL_FILE = r"Path\To\Your\RIDERSHIP_BY_ROUTE_AND_STOP.XLSX"

# Ridership Excel column names (Ridecheck ALL_TIME_PERIODS export).
RIDERSHIP_STOP_ID_FIELD: str = "STOP_ID"
RIDERSHIP_BOARDINGS_FIELD: str = "BOARD_ALL"
RIDERSHIP_ALIGHTINGS_FIELD: str = "ALIGHT_ALL"
RIDERSHIP_ROUTE_FIELD: str = "ROUTE_NAME"

# Which GTFS stops.txt field the ridership STOP_ID column matches. Ridecheck's
# STOP_ID is usually the public-facing stop code, which joins to GTFS
# "stop_code"; use "stop_id" if your export carries GTFS stop_ids instead.
GTFS_JOIN_KEY: str = "stop_code"

# ROUTES = keep-only list  |  ROUTES_EXCLUDE = toss-out list
ROUTES: list[str] = []  # keep these (empty keeps all)
ROUTES_EXCLUDE: list[str] = []  # drop these (empty drops none)

# Optional TIME_PERIOD filter. The Ridecheck ALL_TIME_PERIODS export has one
# row per route x direction x TIME_PERIOD, and the periods partition the day
# (no rollup row), so an empty list sums every row — the daily total, the
# same treatment as data_request_by_stop_processor.py's "All Time Periods"
# sheet. A non-empty list keeps only those periods (compared
# uppercased/stripped). NOTE: this rests on the periods being disjoint; a
# feed variant that adds a rollup row would double-count under the
# empty-list default.
TIME_PERIOD_FIELD: str = "TIME_PERIOD"
TIME_PERIODS: list[str] = []  # e.g. ["AM PEAK", "PM PEAK"]; empty keeps all

# BOUNDARY ALLOCATION ---------------------------------------------------------
# How a stop exactly coincident with a shared district boundary is allocated
# (the containment join yields one row per touching district for such stops):
# "split" — its ridership is divided evenly among the touching districts, so
#           district totals reconcile exactly to the systemwide total.
# "full"  — it counts fully in every touching district (a boundary-as-inside
#           convention). District totals can then sum above the systemwide
#           total and percentages above 100; the reconciliation sheet reports
#           the double-counted amount.
BOUNDARY_ALLOCATION: str = "split"

# PERCENT BASE ----------------------------------------------------------------
# Denominator for the pct_* share columns, named by reconciliation category:
# "allocated"   — ridership allocated to districts, counting each stop once.
#                 District shares sum to 100 under "split" (default).
# "geocoded"    — also counts ridership at geocoded stops outside every
#                 district, so shares sum below 100 by exactly the outside
#                 share.
# "grand_total" — the full ridership file: additionally counts STOP_IDs with
#                 no GTFS match (join failures, not geography — use with
#                 care).
PCT_BASE: str = "allocated"

# District polygons and the field carrying the district identifier/name.
DISTRICTS_FC = r"Path\To\Your\Districts.shp"
DISTRICT_FIELD = "DISTRICT"

# SPATIAL ---------------------------------------------------------------------
# Stops are projected into the district layer's own spatial reference before
# the containment join, so no target EPSG is configured and the districts
# layer itself is never reprojected. If the WGS84-to-district-datum shift
# matters at your stop locations, name an explicit geographic transformation
# (e.g. "WGS_1984_(ITRF00)_To_NAD_1983"); None applies no datum
# transformation.
GEO_TRANSFORMATION: Optional[str] = None

# WORKSPACE -------------------------------------------------------------------
# Working directory for the intermediate file geodatabase. A local (non-
# network) path keeps the geoprocessing steps fast.
WORK_DIR = os.path.abspath(r"temp\district_ridership_share_work")
WORK_GDB_NAME = "work.gdb"

# OUTPUTS ---------------------------------------------------------------------
OUTPUT_DIR = r"Path\To\Your\Output_Folder"
OUTPUT_EXCEL_NAME = "district_ridership_share.xlsx"
LOG_DIR = r"Path\To\Your\Logs"

LOG_LEVEL: int = logging.INFO  # DEBUG / INFO / WARNING / ERROR

# ROUNDING --------------------------------------------------------------------
# Per-stop values are rounded to ROUND_DECIMALS after the per-stop collapse,
# BEFORE allocation/summing, so the district totals on the summary sheet
# equal sums of the rounded per-stop rows on the allocation sheet — the
# workbook stays internally consistent at the displayed precision.
APPLY_ROUNDING: bool = True
ROUND_DECIMALS: int = 1
PCT_DECIMALS: int = 1

# Font applied to all cells in the output workbook.
EXCEL_FONT_NAME: str = "Arial"

# When True, a failed run-log write aborts the script so the analyst is never
# left with outputs that lack a matching configuration record.
REQUIRE_RUN_LOG: bool = True

# In a Jupyter kernel __file__ is undefined, so the run log cannot read this
# script's own source to capture the config block verbatim. Optionally point
# this at the .py on disk to restore verbatim capture. If left empty, the run
# log falls back to a snapshot of the live config values instead.
SOURCE_FILE_OVERRIDE: str = r""

# === END CONFIG ===

# Metrics carried through allocation. Keys of the per-stop ridership dicts and
# column stems in the output ("boardings", "pct_boardings", ...).
METRICS: tuple[str, ...] = ("boardings", "alightings", "total")


# =============================================================================
# LOGGING
# =============================================================================


def configure_logging(log_dir: str) -> None:
    """Configure root logging to write to console plus a file in *log_dir*."""
    os.makedirs(log_dir, exist_ok=True)
    log_path = os.path.join(log_dir, "district_ridership_share_arcpy.log")
    logging.basicConfig(
        level=LOG_LEVEL,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
        handlers=[
            logging.FileHandler(log_path, encoding="utf-8"),
            logging.StreamHandler(),
        ],
        force=True,
    )
    logging.info("Logging to: %s", log_path)


def log_arcpy_messages(level: int = logging.INFO) -> None:
    """Log any messages generated by the last ArcPy operation."""
    msg = arcpy.GetMessages()
    if msg:
        logging.log(level, "ArcPy messages:\n%s", msg)


# =============================================================================
# UTILITIES
# =============================================================================


def safe_name(prefix: str, workspace: str) -> str:
    """Generate a unique and valid name for an ArcGIS feature class or table.

    Args:
        prefix: The base prefix for the name (e.g., 'stops_wgs84').
        workspace: The workspace (GDB) where the name will be used.

    Returns:
        A validated, unique name string.
    """
    suffix = uuid.uuid4().hex[:8]
    return arcpy.ValidateTableName(f"{prefix}_{suffix}", workspace)


def ensure_work_gdb(work_dir: str, gdb_name: str) -> str:
    """Create a file geodatabase (GDB) if it doesn't already exist.

    Args:
        work_dir: The directory where the GDB should reside.
        gdb_name: The name of the GDB file (e.g., 'work.gdb').

    Returns:
        The full path to the geodatabase.
    """
    os.makedirs(work_dir, exist_ok=True)
    gdb = os.path.join(work_dir, gdb_name)
    if not arcpy.Exists(gdb):
        logging.info("Creating work GDB: %s", gdb)
        arcpy.management.CreateFileGDB(work_dir, gdb_name)
        log_arcpy_messages()
    return gdb


def sha256_of_file(path: str) -> Optional[str]:
    """Return the SHA-256 hex digest of *path*, or None if unreadable."""
    try:
        h = hashlib.sha256()
        with open(path, "rb") as f:
            for chunk in iter(lambda: f.read(1 << 20), b""):
                h.update(chunk)
        return h.hexdigest()
    except OSError:
        return None


# =============================================================================
# GTFS STOPS LOADING
# =============================================================================


def load_gtfs_stops(gtfs_input: str) -> pd.DataFrame:
    """Load GTFS stops.txt from a feed folder, a stops.txt path, or a .zip feed.

    Zip members may sit at the archive root or nested one level inside a single
    wrapper folder; both layouts are handled. All columns are read as strings
    to avoid pandas' type inference mangling identifiers (leading zeros).

    Args:
        gtfs_input: Path to a GTFS feed folder, a stops.txt file, or a .zip
            archive of the feed.

    Returns:
        The raw stops table as a DataFrame of strings.

    Raises:
        OSError: Path missing, or stops.txt not found in the feed.
        ValueError: Not a folder/stops.txt/.zip, an invalid zip, an ambiguous
            zip (stops.txt in multiple locations), or an unparseable/empty file.
    """
    if not os.path.exists(gtfs_input):
        raise OSError(f"The path '{gtfs_input}' does not exist.")

    lowered = gtfs_input.lower()
    try:
        if os.path.isdir(gtfs_input):
            path = os.path.join(gtfs_input, "stops.txt")
            if not os.path.isfile(path):
                raise OSError(f"stops.txt not found in folder '{gtfs_input}'.")
            df = pd.read_csv(path, dtype=str, low_memory=False)
        elif lowered.endswith(".txt"):
            df = pd.read_csv(gtfs_input, dtype=str, low_memory=False)
        elif lowered.endswith(".zip"):
            try:
                archive = zipfile.ZipFile(gtfs_input)
            except zipfile.BadZipFile as exc:
                raise ValueError(f"'{gtfs_input}' is not a valid zip archive.") from exc
            with archive:
                candidates = [n for n in archive.namelist() if os.path.basename(n) == "stops.txt"]
                if not candidates:
                    raise OSError(f"stops.txt not found inside '{gtfs_input}'.")
                if len(candidates) > 1:
                    raise ValueError(
                        f"Ambiguous stops.txt in '{gtfs_input}' (multiple locations): "
                        f"{', '.join(candidates)}"
                    )
                with archive.open(candidates[0]) as handle:
                    df = pd.read_csv(handle, dtype=str, low_memory=False)
        else:
            raise ValueError(f"'{gtfs_input}' is not a GTFS folder, a stops.txt, or a .zip feed.")
    except pd.errors.EmptyDataError as exc:
        raise ValueError(f"stops.txt in '{gtfs_input}' is empty.") from exc
    except pd.errors.ParserError as exc:
        raise ValueError(f"Parser error reading stops.txt in '{gtfs_input}': {exc}") from exc

    logging.info("Loaded stops.txt (%d records).", len(df))
    return df


def filter_stops(stops: pd.DataFrame, join_key: str = GTFS_JOIN_KEY) -> pd.DataFrame:
    """Filter GTFS stops to boarding locations with a usable join key and coords.

    The join key is the stops.txt field the ridership STOP_ID column matches
    ("stop_id" or "stop_code"). Keeps location_type 0/blank rows, drops rows
    missing the key or coordinates, and warns on duplicate key values.
    Duplicates are retained: each point participates in the containment join
    and the districts found are unioned per key, so a key whose points land
    in different districts is treated like a boundary-coincident stop and
    allocated per BOUNDARY_ALLOCATION. The warning exists so that outcome is
    a visible choice, not a silent one.

    Args:
        stops: Raw stops.txt DataFrame (string dtypes).
        join_key: stops.txt field to key stops on ("stop_id" or "stop_code").

    Returns:
        DataFrame with stop_key, stop_lon, stop_lat (floats for coords).

    Raises:
        ValueError: *join_key* invalid, required columns missing, or no
            usable stops remain.
    """
    if join_key not in {"stop_id", "stop_code"}:
        raise ValueError(f"GTFS_JOIN_KEY must be 'stop_id' or 'stop_code', got '{join_key}'.")
    required = [join_key, "stop_lat", "stop_lon"]
    missing = [c for c in required if c not in stops.columns]
    if missing:
        raise ValueError(
            f"stops.txt missing column(s): {', '.join(missing)}. "
            f"Available: {', '.join(stops.columns)}"
        )

    df = stops.copy()
    if "location_type" in df.columns:
        df = df[df["location_type"].fillna("0").replace("", "0").astype(str) == "0"]

    df["stop_key"] = df[join_key].fillna("").astype(str).str.strip()
    n_blank = int((df["stop_key"] == "").sum())
    if n_blank:
        logging.warning(
            "Dropping %d stop(s) with blank %s (cannot match ridership STOP_ID).",
            n_blank,
            join_key,
        )
    df = df[df["stop_key"] != ""]

    df = df.dropna(subset=["stop_lat", "stop_lon"])
    df["stop_lat"] = df["stop_lat"].astype(float)
    df["stop_lon"] = df["stop_lon"].astype(float)

    n_dupe = int(df["stop_key"].duplicated().sum())
    if n_dupe:
        logging.warning(
            "%d duplicate %s value(s) in GTFS. Districts are unioned per key; "
            "allocation across them follows BOUNDARY_ALLOCATION.",
            n_dupe,
            join_key,
        )

    if df.empty:
        raise ValueError(f"No usable boarding stops with {join_key} remain after filtering.")

    logging.info("Filtered to %d boarding stops keyed on %s.", len(df), join_key)
    return df[["stop_key", "stop_lon", "stop_lat"]]


# =============================================================================
# SPATIAL WORK
# =============================================================================


def stops_to_points(stops_df: pd.DataFrame, out_gdb: str) -> str:
    """Build a WGS84 point feature class from the stops DataFrame.

    Uses CreateFeatureclass + an insert cursor rather than XYTableToPoint from
    a CSV, so the join key is written as TEXT verbatim — CSV type inference can
    coerce numeric-looking IDs and drop leading zeros, which would silently
    break the ridership join.

    Args:
        stops_df: DataFrame with stop_key, stop_lon, stop_lat.
        out_gdb: Geodatabase for the output feature class.

    Returns:
        Full path to the created point feature class.
    """
    name = safe_name("stops_wgs84", out_gdb)
    out_fc = os.path.join(out_gdb, name)
    logging.info("Creating stops point feature class (%d points).", len(stops_df))

    arcpy.management.CreateFeatureclass(
        out_path=out_gdb,
        out_name=name,
        geometry_type="POINT",
        spatial_reference=arcpy.SpatialReference(4326),
    )
    log_arcpy_messages()
    arcpy.management.AddField(out_fc, "stop_key", "TEXT", field_length=64)
    log_arcpy_messages()

    with arcpy.da.InsertCursor(out_fc, ["SHAPE@XY", "stop_key"]) as cur:
        for code, lon, lat in stops_df[["stop_key", "stop_lon", "stop_lat"]].itertuples(
            index=False
        ):
            cur.insertRow(((float(lon), float(lat)), str(code)))

    return out_fc


def validate_district_field(districts_fc: str, district_field: str) -> None:
    """Check that *district_field* exists on the district layer before joining.

    Args:
        districts_fc: District polygon layer.
        district_field: Field expected to carry the district identifier/name.

    Raises:
        ValueError: The field is absent (the message lists available fields).
    """
    try:
        available = {f.name for f in arcpy.ListFields(districts_fc)}
    except OSError as exc:
        raise ValueError(
            f"Could not read fields from district layer '{districts_fc}': {exc}"
        ) from exc
    if district_field not in available:
        raise ValueError(
            f"DISTRICT_FIELD '{district_field}' not found in '{districts_fc}'. "
            f"Available fields: {', '.join(sorted(available))}"
        )


def project_stops_to_district_sr(
    stops_fc: str,
    districts_fc: str,
    out_gdb: str,
    transformation: Optional[str] = GEO_TRANSFORMATION,
) -> str:
    """Project the stops into the district layer's own spatial reference.

    Containment is evaluated in the district layer's native spatial reference
    (read via arcpy.Describe), so the polygons are never reprojected — only
    the points move. When *transformation* is named it is applied; otherwise
    no datum transformation is used.

    Args:
        stops_fc: WGS84 stops point feature class.
        districts_fc: District polygon layer whose SR is the target.
        out_gdb: The output geodatabase.
        transformation: Optional geographic (datum) transformation name.

    Returns:
        The full path to the projected stops feature class.

    Raises:
        ValueError: The district layer's spatial reference cannot be read or
            is unknown (containment in an unknown SR is meaningless).
    """
    try:
        desc = arcpy.Describe(districts_fc)
        target_sr = desc.spatialReference
    except OSError as exc:
        raise ValueError(f"Could not describe district layer '{districts_fc}': {exc}") from exc
    if target_sr is None or getattr(target_sr, "name", "Unknown") in ("", "Unknown"):
        raise ValueError(
            f"District layer '{districts_fc}' has no usable spatial reference "
            "(missing/unknown .prj). Define its projection before running."
        )

    out_fc = os.path.join(out_gdb, safe_name("stops_proj", out_gdb))
    logging.info(
        "Projecting stops into district SR '%s'%s",
        target_sr.name,
        f" (transformation: {transformation})" if transformation else "",
    )
    if transformation:
        arcpy.management.Project(stops_fc, out_fc, target_sr, transformation)
    else:
        arcpy.management.Project(stops_fc, out_fc, target_sr)
    log_arcpy_messages()
    return out_fc


def containment_join_stops_to_districts(
    stops_fc: str,
    districts_fc: str,
    out_gdb: str,
) -> str:
    """Spatially join stops to districts by point-in-polygon containment.

    Uses INTERSECT with JOIN_ONE_TO_MANY / KEEP_COMMON so a stop coincident
    with a shared district boundary produces one output row per district —
    the precondition for boundary allocation. No search radius: this is
    containment, unlike the proximity join used by
    ``gtfs_service_by_district_arcpy.py``.

    Args:
        stops_fc: Projected stops point feature class (carries stop_key).
        districts_fc: District polygon layer.
        out_gdb: The output geodatabase.

    Returns:
        The full path to the spatially joined output feature class.
    """
    out_fc = os.path.join(out_gdb, safe_name("stops_districts_contains", out_gdb))
    logging.info("SpatialJoin INTERSECT (containment, one-to-many)")
    arcpy.analysis.SpatialJoin(
        target_features=stops_fc,
        join_features=districts_fc,
        out_feature_class=out_fc,
        join_operation="JOIN_ONE_TO_MANY",
        join_type="KEEP_COMMON",
        match_option="INTERSECT",
    )
    log_arcpy_messages()
    return out_fc


def extract_stop_districts(fc: str, district_field: str) -> dict[str, set[str]]:
    """Read the spatial join result into a stop_key -> set-of-districts mapping.

    Args:
        fc: The spatially joined feature class.
        district_field: The name of the district ID field.

    Returns:
        A dictionary mapping stop_key strings to a set of district ID strings.

    Raises:
        ValueError: The join produced zero stop-district pairs (silent-zero
            guard: an unpopulated DISTRICT_FIELD or a projection mismatch
            would otherwise flow through as an empty report).
    """
    stop_to_districts: dict[str, set[str]] = {}
    with arcpy.da.SearchCursor(fc, ["stop_key", district_field]) as cur:
        for stop_key, dist in cur:
            if stop_key and dist is not None and str(dist).strip():
                stop_to_districts.setdefault(str(stop_key), set()).add(str(dist))
    if not stop_to_districts:
        raise ValueError(
            "Containment join produced zero stop-district pairs. Check that "
            f"DISTRICT_FIELD ('{district_field}') is populated and that the "
            "districts layer actually covers the stop locations."
        )
    n_multi = sum(1 for ds in stop_to_districts.values() if len(ds) > 1)
    logging.info(
        "Containment join: %d stop key(s) in a district; %d span more than one.",
        len(stop_to_districts),
        n_multi,
    )
    return stop_to_districts


# =============================================================================
# RIDERSHIP LOADING + ALLOCATION
# =============================================================================


def _clean_id(value: Any) -> str:
    """Render an identifier cell as a clean string.

    Excel stores numeric-looking IDs as floats, so a STOP_ID of 1234 reads
    back as 1234.0 and str() would yield '1234.0' — which never matches a
    GTFS join key. Integral floats are rendered without the trailing '.0';
    everything else is stripped str().
    """
    if pd.isna(value):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def load_stop_ridership(
    excel_file: str,
    routes: Optional[Sequence[str]] = None,
    routes_exclude: Optional[Sequence[str]] = None,
    time_periods: Optional[Sequence[str]] = None,
) -> dict[str, dict[str, float]]:
    """Load per-stop ridership from the Ridecheck Excel, keyed by stop ID.

    Column names come from the RIDERSHIP_*_FIELD config constants (defaults
    match the Ridecheck ALL_TIME_PERIODS export). Rows are filtered by the
    keep-list *routes*, then the drop-list *routes_exclude*, then by
    *time_periods*, then collapsed to one row per stop ID (summing across
    routes, directions, and kept periods).

    Time periods: the ALL_TIME_PERIODS export's periods partition the day
    (no rollup row), so an empty *time_periods* sums every row — the daily
    total, matching ``data_request_by_stop_processor.py``'s "All Time
    Periods" treatment. A non-empty *time_periods* keeps only those periods;
    values are compared uppercased/stripped, and naming a period absent from
    the file fails loud.

    Args:
        excel_file: Path to the ridership workbook.
        routes: Route names to keep (inclusive). Empty/None keeps all.
        routes_exclude: Route names to drop (exclusive). Empty/None drops none.
        time_periods: TIME_PERIOD values to keep. Empty/None keeps all.

    Returns:
        Mapping stop ID -> {"boardings", "alightings", "total"} floats.

    Raises:
        OSError: excel_file does not exist.
        ValueError: A required column is missing, *time_periods* names periods
            absent from the file, or no rows remain.
    """
    if not os.path.isfile(excel_file):
        raise OSError(f"Ridership Excel not found: {excel_file}")

    df = pd.read_excel(excel_file)
    required = [RIDERSHIP_STOP_ID_FIELD, RIDERSHIP_BOARDINGS_FIELD, RIDERSHIP_ALIGHTINGS_FIELD]
    missing = [c for c in required if c not in df.columns]
    if missing:
        raise ValueError(
            f"Ridership Excel '{excel_file}' missing column(s): {', '.join(missing)}. "
            f"Available: {', '.join(str(c) for c in df.columns)}"
        )

    if routes or routes_exclude:
        if RIDERSHIP_ROUTE_FIELD not in df.columns:
            raise ValueError(
                f"ROUTES/ROUTES_EXCLUDE given but '{RIDERSHIP_ROUTE_FIELD}' column not present."
            )
        route_col = df[RIDERSHIP_ROUTE_FIELD].map(_clean_id)
        if routes:
            initial = len(df)
            wanted = {str(r).strip() for r in routes}
            df = df[route_col.isin(wanted)]
            route_col = route_col[df.index]
            logging.info(
                "ROUTES keep-filter %s: records reduced from %d to %d.",
                sorted(wanted),
                initial,
                len(df),
            )
        if routes_exclude:
            initial = len(df)
            dropped = {str(r).strip() for r in routes_exclude}
            df = df[~route_col.isin(dropped)]
            logging.info(
                "ROUTES_EXCLUDE drop-filter %s: records reduced from %d to %d.",
                sorted(dropped),
                initial,
                len(df),
            )

    # Time-period filter: empty keeps all periods (disjoint by assumption).
    if TIME_PERIOD_FIELD and TIME_PERIOD_FIELD in df.columns:
        period_col = df[TIME_PERIOD_FIELD].map(_clean_id).str.upper()
        periods_present = sorted(period_col.unique())
        if time_periods:
            wanted_periods = {str(p).strip().upper() for p in time_periods}
            absent = sorted(wanted_periods - set(periods_present))
            if absent:
                raise ValueError(
                    f"TIME_PERIODS value(s) not found in file: {absent}. "
                    f"Available: {periods_present}"
                )
            initial = len(df)
            df = df[period_col.isin(wanted_periods)]
            logging.info(
                "Time-period filter %s: records reduced from %d to %d.",
                sorted(wanted_periods),
                initial,
                len(df),
            )
        else:
            logging.info(
                "TIME_PERIODS empty: summing all %d period(s) %s (daily total; "
                "assumes disjoint periods).",
                len(periods_present),
                periods_present,
            )
    elif TIME_PERIOD_FIELD:
        logging.info(
            "No '%s' column in ridership file; skipping time-period filter.",
            TIME_PERIOD_FIELD,
        )

    df = df.copy()
    df["_stop"] = df[RIDERSHIP_STOP_ID_FIELD].map(_clean_id)
    for col in (RIDERSHIP_BOARDINGS_FIELD, RIDERSHIP_ALIGHTINGS_FIELD):
        n_bad = int(pd.to_numeric(df[col], errors="coerce").isna().sum() - df[col].isna().sum())
        if n_bad > 0:
            logging.warning("Coercing %d non-numeric value(s) in %s to 0.", n_bad, col)
        df[col] = pd.to_numeric(df[col], errors="coerce").fillna(0.0)
    df["_total"] = df[RIDERSHIP_BOARDINGS_FIELD] + df[RIDERSHIP_ALIGHTINGS_FIELD]

    df = df[df["_stop"] != ""]
    if df.empty:
        raise ValueError(f"No usable ridership rows in '{excel_file}'.")

    per_stop = df.groupby("_stop", as_index=False).agg(
        {RIDERSHIP_BOARDINGS_FIELD: "sum", RIDERSHIP_ALIGHTINGS_FIELD: "sum", "_total": "sum"}
    )
    if APPLY_ROUNDING:
        # Round per-stop values BEFORE allocation so the district totals on
        # the summary sheet equal sums of the rounded per-stop rows on the
        # allocation sheet.
        for col in (RIDERSHIP_BOARDINGS_FIELD, RIDERSHIP_ALIGHTINGS_FIELD, "_total"):
            per_stop[col] = per_stop[col].round(ROUND_DECIMALS)
    logging.info(
        "Loaded ridership for %d stop(s); boardings total = %.1f.",
        len(per_stop),
        float(per_stop[RIDERSHIP_BOARDINGS_FIELD].sum()),
    )
    return {
        str(row["_stop"]): {
            "boardings": float(row[RIDERSHIP_BOARDINGS_FIELD]),
            "alightings": float(row[RIDERSHIP_ALIGHTINGS_FIELD]),
            "total": float(row["_total"]),
        }
        for _, row in per_stop.iterrows()
    }


def allocate_ridership_to_districts(
    stop_to_districts: dict[str, set[str]],
    stop_ridership: dict[str, dict[str, float]],
    geocoded_stops: set[str],
    metrics: Sequence[str] = METRICS,
    boundary_allocation: str = BOUNDARY_ALLOCATION,
    pct_base: str = PCT_BASE,
    join_key: str = GTFS_JOIN_KEY,
) -> tuple[pd.DataFrame, pd.DataFrame, dict[str, dict[str, float]]]:
    """Allocate each stop's ridership to the district(s) it falls in.

    Under point-in-polygon containment a stop belongs to one district; a stop
    exactly coincident with a shared boundary is joined to each touching
    district. *boundary_allocation* controls what happens then: "split"
    divides its ridership evenly among the touching districts (district
    totals reconcile exactly), while "full" counts the full value in every
    touching district (district totals then double-count such stops and
    percentages can sum above 100). Percentages default to shares of the
    allocated total per metric, counting each stop ONCE regardless of mode;
    *pct_base* can widen the denominator to include out-of-district or
    unmatched ridership. The diagnostics expose ridership that landed in no
    district, never matched GTFS, or was double-counted across boundaries,
    so every discrepancy is auditable.

    Args:
        stop_to_districts: Mapping stop_key -> set of district IDs.
        stop_ridership: Mapping stop_key -> {metric: value}.
        geocoded_stops: stop keys present in the GTFS feed with coordinates,
            used to separate "outside all districts" from "not in GTFS".
        metrics: Metric keys to allocate and report.
        boundary_allocation: "split" or "full" (see above).
        pct_base: Denominator for the pct_* columns — "allocated" (default),
            "geocoded" (adds ridership outside every district), or
            "grand_total" (also adds ridership unmatched to GTFS).
        join_key: The GTFS join key in use, echoed in the mismatch error.

    Returns:
        Tuple of (district_df, alloc_df, diagnostics) where diagnostics maps
        'matched' / 'no_district' / 'unmatched_to_gtfs' / 'boundary_extra'
        -> {metric: total}.

    Raises:
        ValueError: *boundary_allocation* or *pct_base* invalid, or nothing
            allocated (likely a join-key/STOP_ID mismatch — check
            GTFS_JOIN_KEY).
    """
    if boundary_allocation not in ("full", "split"):
        raise ValueError(
            f"BOUNDARY_ALLOCATION must be 'full' or 'split', got {boundary_allocation!r}."
        )
    if pct_base not in ("allocated", "geocoded", "grand_total"):
        raise ValueError(
            f"PCT_BASE must be 'allocated', 'geocoded', or 'grand_total', got {pct_base!r}."
        )
    district_totals: dict[str, dict[str, float]] = {}
    alloc_rows: list[dict[str, Any]] = []
    matched = {m: 0.0 for m in metrics}
    no_district = {m: 0.0 for m in metrics}
    unmatched_gtfs = {m: 0.0 for m in metrics}
    boundary_extra = {m: 0.0 for m in metrics}  # double-count under "full"

    for stop_key, vals in stop_ridership.items():
        if stop_key not in geocoded_stops:
            for m in metrics:
                unmatched_gtfs[m] += vals.get(m, 0.0)
            continue
        districts = stop_to_districts.get(stop_key)
        if not districts:
            for m in metrics:
                no_district[m] += vals.get(m, 0.0)
            continue
        n = len(districts)
        for m in metrics:
            matched[m] += vals.get(m, 0.0)
            if boundary_allocation == "full" and n > 1:
                boundary_extra[m] += vals.get(m, 0.0) * (n - 1)
        for d in sorted(districts):
            dd = district_totals.setdefault(d, {m: 0.0 for m in metrics})
            row: dict[str, Any] = {"stop_key": stop_key, "district": d, "n_districts": n}
            for m in metrics:
                share = vals.get(m, 0.0) if boundary_allocation == "full" else vals.get(m, 0.0) / n
                dd[m] += share
                row[f"{m}_alloc"] = share
            alloc_rows.append(row)

    if all(matched[m] == 0.0 for m in metrics):
        raise ValueError(
            "No ridership allocated to any district. Likely a join-key mismatch "
            f"(GTFS_JOIN_KEY='{join_key}'). Ridership sample: "
            f"{sorted(stop_ridership)[:5]}; GTFS sample: {sorted(geocoded_stops)[:5]}"
        )

    pct_denoms = {
        m: (
            matched[m]
            + (no_district[m] if pct_base in ("geocoded", "grand_total") else 0.0)
            + (unmatched_gtfs[m] if pct_base == "grand_total" else 0.0)
        )
        for m in metrics
    }
    district_rows: list[dict[str, Any]] = []
    for d in sorted(district_totals):
        r: dict[str, Any] = {"district": d}
        for m in metrics:
            r[m] = district_totals[d][m]
            r[f"pct_{m}"] = district_totals[d][m] / pct_denoms[m] * 100.0 if pct_denoms[m] else 0.0
        district_rows.append(r)

    district_df = pd.DataFrame(district_rows)
    if APPLY_ROUNDING:
        for m in metrics:
            district_df[m] = district_df[m].round(ROUND_DECIMALS)
            district_df[f"pct_{m}"] = district_df[f"pct_{m}"].round(PCT_DECIMALS)
    alloc_cols = ["stop_key", "district", "n_districts"] + [f"{m}_alloc" for m in metrics]
    alloc_df = pd.DataFrame(alloc_rows, columns=alloc_cols).sort_values(
        ["district", "stop_key"], kind="stable", ignore_index=True
    )
    diagnostics = {
        "matched": matched,
        "no_district": no_district,
        "unmatched_to_gtfs": unmatched_gtfs,
        "boundary_extra": boundary_extra,
    }
    return district_df, alloc_df, diagnostics


def build_reconciliation_frame(
    diagnostics: dict[str, dict[str, float]],
    metrics: Sequence[str] = METRICS,
) -> pd.DataFrame:
    """Build the reconciliation sheet: one row per category, one column per metric.

    Categories:
        allocated_to_districts: ridership successfully allocated, counting
            each stop once (the pct base).
        boundary_double_count: extra ridership counted under "full" boundary
            allocation because boundary stops count fully in more than one
            district (0.0 under "split"). Sum of district totals equals
            allocated_to_districts + this row.
        outside_all_districts: geocoded stops falling in no district polygon.
        unmatched_to_gtfs: ridership STOP_IDs with no GTFS join-key match.
        grand_total: allocated + outside + unmatched — the full ridership
            file, excluding the double-count row.

    Args:
        diagnostics: Output of allocate_ridership_to_districts.
        metrics: Metric keys, in column order.

    Returns:
        DataFrame with 'category' plus one column per metric.
    """
    rows = [
        {"category": "allocated_to_districts", **diagnostics["matched"]},
        {"category": "boundary_double_count", **diagnostics["boundary_extra"]},
        {"category": "outside_all_districts", **diagnostics["no_district"]},
        {"category": "unmatched_to_gtfs", **diagnostics["unmatched_to_gtfs"]},
    ]
    grand = {
        m: sum(diagnostics[k][m] for k in ("matched", "no_district", "unmatched_to_gtfs"))
        for m in metrics
    }
    rows.append({"category": "grand_total", **grand})
    frame = pd.DataFrame(rows, columns=["category", *metrics])
    if APPLY_ROUNDING:
        for m in metrics:
            frame[m] = frame[m].round(ROUND_DECIMALS)
    return frame


# =============================================================================
# EXCEL OUTPUT
# =============================================================================


def apply_font_to_workbook(xlsx_path: str, font_name: str) -> None:
    """Apply *font_name* to every cell in the workbook, preserving bold headers."""
    from openpyxl import load_workbook
    from openpyxl.styles import Font

    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                bold = bool(cell.font.bold) if cell.font else False
                cell.font = Font(name=font_name, bold=bold)
    wb.save(xlsx_path)


def write_output_workbook(
    out_path: str,
    district_df: pd.DataFrame,
    alloc_df: pd.DataFrame,
    recon_df: pd.DataFrame,
) -> None:
    """Write the multi-sheet output workbook and apply the house font.

    Sheets:
        district_ridership: totals and percent shares per district (summary).
        ridership_allocation: per stop-district allocation detail, including
            the split rows for any boundary-coincident stops.
        reconciliation: allocated vs. outside-district vs. unmatched totals.

    Args:
        out_path: Destination .xlsx path.
        district_df: Summary frame from allocate_ridership_to_districts.
        alloc_df: Detail frame from allocate_ridership_to_districts.
        recon_df: Frame from build_reconciliation_frame.
    """
    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    with pd.ExcelWriter(out_path) as writer:
        district_df.to_excel(writer, sheet_name="district_ridership", index=False)
        alloc_df.to_excel(writer, sheet_name="ridership_allocation", index=False)
        recon_df.to_excel(writer, sheet_name="reconciliation", index=False)
    apply_font_to_workbook(out_path, EXCEL_FONT_NAME)
    logging.info("Workbook written to: %s", out_path)


# =============================================================================
# RUN LOG
# =============================================================================


# Canonical version lives in utils/run_log.py — keep this copy in sync.
def extract_config_block(source_file: Path) -> str:
    r"""Return the text between the CONFIG markers in *source_file*.

    Reads ``source_file`` as UTF-8 text and slices out the lines strictly
    *between* the first occurrence of ``# === BEGIN CONFIG ===`` and the first
    subsequent occurrence of ``# === END CONFIG ===``.  The marker lines
    themselves are excluded; whitespace and inline comments inside the block
    are preserved verbatim.

    Args:
        source_file: Path to the Python source file to scan (typically
            ``Path(__file__)`` from the calling script).

    Returns:
        The verbatim text of the configuration block, joined with ``\n``.

    Raises:
        ValueError: If either marker is missing or they appear out of order.
        OSError: If ``source_file`` cannot be read.
    """
    _BEGIN = "# === BEGIN CONFIG ==="
    _END = "# === END CONFIG ==="

    lines: list[str] = source_file.read_text(encoding="utf-8").splitlines()

    begin_idx: int | None = None
    end_idx: int | None = None
    for i, line in enumerate(lines):
        stripped: str = line.strip()
        if begin_idx is None and stripped == _BEGIN:
            begin_idx = i
        elif begin_idx is not None and stripped == _END:
            end_idx = i
            break

    if begin_idx is None or end_idx is None:
        raise ValueError(
            f"Config markers not found in '{source_file}'. Expected '{_BEGIN}' and '{_END}'."
        )

    return "\n".join(lines[begin_idx + 1 : end_idx])


def _resolve_source_path() -> Optional[Path]:
    """Return a readable path to this script's source, or None.

    Prefers SOURCE_FILE_OVERRIDE (for notebook runs), then SELF_PATH (defined
    when running as a .py file). Returns None when neither is available.
    """
    override = SOURCE_FILE_OVERRIDE.strip()
    if override:
        p = Path(override)
        if p.is_file():
            return p
        logging.warning("SOURCE_FILE_OVERRIDE set but not found: %s", p)
    return SELF_PATH


def _live_config_snapshot() -> str:
    """Build a best-effort config record from live module globals.

    Used when the source file is unavailable (e.g. running in a Jupyter kernel),
    so the run log still carries a record of the configuration actually used.
    Captures UPPER_SNAKE_CASE globals holding simple scalar/sequence values.
    """
    g = globals()
    lines: list[str] = []
    for name in sorted(g):
        if name.startswith("_") or name != name.upper():
            continue
        val = g[name]
        if isinstance(val, (str, int, float, bool, list, tuple, type(None))):
            lines.append(f"{name} = {val!r}")
    return "\n".join(lines)


def input_fingerprints(gtfs_input: str, excel_file: str, districts_fc: str) -> List[str]:
    """Return SHA-256 fingerprint lines for the primary inputs, for provenance.

    A folder GTFS input fingerprints its stops.txt; a zip or stops.txt input
    fingerprints the file itself. The districts layer fingerprints the .shp
    and .dbf sidecars (geometry + attributes).

    Args:
        gtfs_input: The GTFS feed path actually used for this run.
        excel_file: The ridership Excel path actually used for this run.
        districts_fc: The district layer path actually used for this run.

    Returns:
        Formatted "label: path" / "sha256: digest" line pairs for the run log.
    """
    targets: list[tuple[str, str]] = []
    if os.path.isdir(gtfs_input):
        targets.append(("GTFS stops.txt", os.path.join(gtfs_input, "stops.txt")))
    else:
        targets.append(("GTFS input", gtfs_input))
    targets.append(("Ridership Excel", excel_file))
    targets.append(("Districts .shp", districts_fc))
    dbf = os.path.splitext(districts_fc)[0] + ".dbf"
    if os.path.isfile(dbf):
        targets.append(("Districts .dbf", dbf))

    lines: list[str] = []
    for label, path in targets:
        digest = sha256_of_file(path)
        lines.append(f"{label}: {path}")
        lines.append(f"  sha256: {digest if digest else 'unreadable'}")
    return lines


def write_run_log(
    output_folder: str,
    effective_settings: Sequence[str],
    fingerprints: Sequence[str],
) -> bool:
    """Write a run log of the configuration into *output_folder*.

    When the script's source is readable, the config block is captured
    verbatim. Otherwise (e.g. a Jupyter kernel with no __file__) it falls back
    to a live snapshot of the config globals so a run still produces a
    configuration record. The effective-settings section records the values
    actually used, which may come from CLI flags rather than the constants.

    Args:
        output_folder: Directory the run log is written into.
        effective_settings: Pre-formatted lines describing the resolved
            settings for this run.
        fingerprints: Pre-formatted input fingerprint lines (from
            :func:`input_fingerprints`).

    Returns:
        ``True`` if the log was written successfully, ``False`` otherwise.
    """
    log_path = Path(output_folder) / "district_ridership_share_runlog.txt"

    source_path = _resolve_source_path()
    if source_path is not None:
        try:
            config_text = extract_config_block(source_path)
            config_heading = "CONFIGURATION (verbatim from source)"
            source_label = str(source_path)
        except (OSError, ValueError) as exc:
            logging.warning(
                "Could not extract config block from '%s' (%s); falling back to a live snapshot.",
                source_path,
                exc,
            )
            config_text = _live_config_snapshot()
            config_heading = "CONFIGURATION (live snapshot — source unreadable)"
            source_label = f"{source_path} (unreadable)"
    else:
        config_text = _live_config_snapshot()
        config_heading = "CONFIGURATION (live snapshot — source file unavailable)"
        source_label = "unavailable (running without __file__, e.g. Jupyter)"

    lines: list[str] = [
        "=" * 72,
        "DISTRICT RIDERSHIP SHARE RUN LOG",
        "=" * 72,
        f"Run timestamp:    {datetime.now().isoformat(timespec='seconds')}",
        f"Output folder:    {output_folder}",
        f"Source script:    {source_label}",
        "",
        "-" * 72,
        "EFFECTIVE SETTINGS (constants or CLI flags, as resolved for this run)",
        "-" * 72,
        *effective_settings,
        "",
        "-" * 72,
        "INPUT FINGERPRINTS",
        "-" * 72,
        *fingerprints,
        "",
        "-" * 72,
        config_heading,
        "-" * 72,
        config_text,
        "=" * 72,
    ]

    try:
        log_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
        logging.info("Run log saved to '%s'.", log_path)
        return True
    except OSError as exc:
        logging.error("Error writing run log: %s", exc)
        return False


# =============================================================================
# PIPELINE
# =============================================================================


def run_analysis(args: argparse.Namespace) -> None:
    """Execute the full allocation pipeline with resolved settings.

    Args:
        args: Parsed CLI arguments (defaults mirror the CONFIGURATION block).

    Raises:
        OSError: An input path is missing or unreadable.
        ValueError: Invalid settings, empty inputs, or a failed/empty join
            (messages say which and how to fix it).
        RuntimeError: The run log could not be written and REQUIRE_RUN_LOG
            is True.
    """
    work_gdb = ensure_work_gdb(args.work_dir, WORK_GDB_NAME)
    arcpy.env.workspace = work_gdb
    logging.info("Workspace: %s", work_gdb)

    stops_df = filter_stops(load_gtfs_stops(args.gtfs_input), args.gtfs_join_key)
    stop_ridership = load_stop_ridership(
        args.excel_file, args.routes, args.routes_exclude, args.time_periods
    )

    validate_district_field(args.districts_fc, args.district_field)
    stops_wgs84 = stops_to_points(stops_df, work_gdb)
    stops_proj = project_stops_to_district_sr(
        stops_wgs84, args.districts_fc, work_gdb, args.geo_transformation
    )
    sj_fc = containment_join_stops_to_districts(stops_proj, args.districts_fc, work_gdb)

    stop_to_districts = extract_stop_districts(sj_fc, args.district_field)
    geocoded_stops = set(stops_df["stop_key"].astype(str))
    district_df, alloc_df, diagnostics = allocate_ridership_to_districts(
        stop_to_districts,
        stop_ridership,
        geocoded_stops,
        boundary_allocation=args.boundary_allocation,
        pct_base=args.pct_base,
        join_key=args.gtfs_join_key,
    )

    recon_df = build_reconciliation_frame(diagnostics)
    for _, row in recon_df.iterrows():
        logging.info(
            "Reconciliation | %-24s boardings=%.1f alightings=%.1f total=%.1f",
            row["category"],
            row["boardings"],
            row["alightings"],
            row["total"],
        )

    out_xlsx = os.path.join(args.output_dir, OUTPUT_EXCEL_NAME)
    write_output_workbook(out_xlsx, district_df, alloc_df, recon_df)

    effective_settings = [
        f"GTFS input:          {args.gtfs_input}",
        f"Ridership Excel:     {args.excel_file}",
        f"Districts layer:     {args.districts_fc}",
        f"District field:      {args.district_field}",
        f"GTFS join key:       {args.gtfs_join_key}",
        f"Boundary allocation: {args.boundary_allocation}",
        f"Percent base:        {args.pct_base}",
        f"Routes kept:         {list(args.routes) if args.routes else 'all'}",
        f"Routes dropped:      {list(args.routes_exclude) if args.routes_exclude else 'none'}",
        f"Time periods:        {list(args.time_periods) if args.time_periods else 'all'}",
        f"Geo transformation:  {args.geo_transformation or 'none'}",
    ]
    log_ok = write_run_log(
        args.output_dir,
        effective_settings=effective_settings,
        fingerprints=input_fingerprints(args.gtfs_input, args.excel_file, args.districts_fc),
    )
    if not log_ok and REQUIRE_RUN_LOG:
        raise RuntimeError(
            "Run log could not be written. Set REQUIRE_RUN_LOG = False to suppress "
            "this error when a sidecar file is genuinely impossible."
        )


# =============================================================================
# CLI / MAIN
# =============================================================================


# Canonical version lives in utils/cli_helpers.py — keep this copy in sync.
def notebook_safe_argv(argv: Optional[Sequence[str]]) -> Optional[List[str]]:
    """Return the argv to parse, shielding notebook kernels from stray flags.

    When a script's ``main()`` runs with no explicit ``argv`` inside a
    Jupyter/IPython kernel, ``sys.argv`` holds kernel plumbing (for example
    ``-f /path/kernel.json``) rather than flags meant for the script, and
    strict ``argparse.parse_args`` would reject it and abort.  This helper
    detects the notebook case and substitutes an empty argument list so the
    CONFIGURATION constants stay in charge, while shell runs keep strict
    parsing (a typo in a flag fails loudly instead of being silently ignored).

    Canonical implementation: ``utils/cli_helpers.py``.

    Args:
        argv: Explicit argument list passed to ``main()``, or ``None`` to
            fall back to ``sys.argv``.

    Returns:
        ``list(argv)`` when *argv* was provided; ``[]`` when running inside a
        notebook kernel; otherwise ``None`` so argparse reads ``sys.argv[1:]``.
    """
    if argv is not None:
        return list(argv)
    if "ipykernel" in sys.modules:
        return []
    return None


def build_arg_parser() -> argparse.ArgumentParser:
    """Create the command-line argument parser (defaults mirror CONFIGURATION)."""
    p = argparse.ArgumentParser(
        description="Allocate stop-level ridership to districts and report shares.",
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    p.add_argument(
        "--gtfs-input",
        default=GTFS_INPUT,
        help="GTFS feed: a folder, a stops.txt path, or a .zip archive.",
    )
    p.add_argument(
        "--excel-file",
        default=EXCEL_FILE,
        help="Stop-level ridership Excel workbook.",
    )
    p.add_argument(
        "--districts-fc",
        default=DISTRICTS_FC,
        help="District polygon layer (shapefile or feature class).",
    )
    p.add_argument(
        "--district-field",
        default=DISTRICT_FIELD,
        help="Field carrying the district identifier/name.",
    )
    p.add_argument(
        "--gtfs-join-key",
        default=GTFS_JOIN_KEY,
        choices=("stop_id", "stop_code"),
        help="stops.txt field the ridership STOP_ID column matches.",
    )
    p.add_argument(
        "--boundary-allocation",
        default=BOUNDARY_ALLOCATION,
        choices=("split", "full"),
        help="Boundary-coincident stops: split evenly, or count fully in each district.",
    )
    p.add_argument(
        "--pct-base",
        default=PCT_BASE,
        choices=("allocated", "geocoded", "grand_total"),
        help="Percent-column denominator: allocated ridership only, plus outside-district "
        "ridership, or the full ridership file.",
    )
    p.add_argument(
        "--routes",
        nargs="*",
        default=ROUTES,
        help="Route names to keep (empty keeps all).",
    )
    p.add_argument(
        "--routes-exclude",
        nargs="*",
        default=ROUTES_EXCLUDE,
        help="Route names to drop (empty drops none).",
    )
    p.add_argument(
        "--time-periods",
        nargs="*",
        default=TIME_PERIODS,
        help="TIME_PERIOD values to keep (empty sums all periods).",
    )
    p.add_argument(
        "--geo-transformation",
        default=GEO_TRANSFORMATION,
        help="Optional geographic (datum) transformation for the stops projection.",
    )
    p.add_argument(
        "--work-dir",
        default=WORK_DIR,
        help="Working directory for the intermediate file geodatabase.",
    )
    p.add_argument(
        "--output-dir",
        default=OUTPUT_DIR,
        help="Directory for the output workbook and run log.",
    )
    p.add_argument(
        "--log-dir",
        default=LOG_DIR,
        help="Directory for the .log file.",
    )
    return p


def main(argv: Optional[Sequence[str]] = None) -> int:
    """Entry point. Validates placeholder paths before doing any work.

    Args:
        argv: Optional explicit argument list; None reads ``sys.argv``.

    Returns:
        Process exit code: 0 on success, 1 on runtime failure, 2 if required
        CONFIGURATION values are still placeholders.
    """
    parser = build_arg_parser()
    args = parser.parse_args(notebook_safe_argv(argv))

    still_placeholder = [
        name
        for name, value in (
            ("GTFS_INPUT / --gtfs-input", args.gtfs_input),
            ("EXCEL_FILE / --excel-file", args.excel_file),
            ("DISTRICTS_FC / --districts-fc", args.districts_fc),
            ("OUTPUT_DIR / --output-dir", args.output_dir),
            ("LOG_DIR / --log-dir", args.log_dir),
        )
        if str(value).startswith(r"Path\To\Your")
    ]
    if still_placeholder:
        logging.basicConfig(level=LOG_LEVEL, force=True)
        logging.warning(
            "Placeholder value(s) still set for: %s. Update the CONFIGURATION "
            "section or pass the matching CLI flags before running.",
            "; ".join(still_placeholder),
        )
        return 2

    configure_logging(args.log_dir)
    arcpy.env.overwriteOutput = True

    try:
        run_analysis(args)
    except (OSError, ValueError, RuntimeError, arcpy.ExecuteError) as exc:
        logging.error("%s", exc)
        return 1

    logging.info("Script completed successfully.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
