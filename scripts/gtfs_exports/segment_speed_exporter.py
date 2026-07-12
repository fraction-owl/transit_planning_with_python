"""Export GTFS time-band speed tables to Excel workbooks.

This module ingests a General Transit Feed Specification (GTFS) folder, groups
trips by stop pattern and schedule, aggregates them into time-of-day bands, and
writes one workbook per (route × service ID).  Each workbook contains one sheet
per travel direction with per-segment speeds.

Typical use-cases
-----------------
* Rapid batch analysis of GTFS feeds in Jupyter Notebooks or ArcGIS Pro.
* Producing reviewer-friendly Excel outputs for operations or service planning.

Before running, **edit the paths and optional filters** in the *CONFIGURATION*
section.
"""

from __future__ import annotations

import logging
import os
import re
import zipfile
from collections.abc import Mapping, Sequence
from pathlib import Path
from typing import Any, Dict, Iterable, List, Literal, Optional, Tuple, TypedDict, Union, cast

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================

GTFS_FOLDER: Path = Path(r"Path\To\Your\GTFS_Folder")  # ←–– change me

OUTPUT_FOLDER: Path = Path(r"Path\To\Your\Output_Folder")  # ←–– change me

# Optional filters – leave empty to process everything
FILTER_IN_ROUTE_SHORT_NAMES: List[str] = ["101", "660"]
FILTER_OUT_ROUTE_SHORT_NAMES: List[str] = []
FILTER_IN_SERVICE_IDS: List[str] = ["3"]
FILTER_OUT_SERVICE_IDS: List[str] = []

EXPORT_TIMEPOINTS_ONLY: bool = True
# Unit of shape_dist_traveled in the GTFS feed: "feet", "meters", or "km".
# Output speeds are always mph, so distances are converted to miles internally.
INPUT_DISTANCE_UNIT: Literal["feet", "meters", "km"] = "meters"
MISSING_VAL: str = "–"

LOG_LEVEL: int = logging.INFO  # DEBUG / INFO / WARNING / ERROR

# -----------------------------------------------------------------------------
# CONSTANTS
# -----------------------------------------------------------------------------

REQ_FILES: Tuple[str, ...] = ("trips.txt", "stop_times.txt", "routes.txt", "stops.txt")

# -----------------------------------------------------------------------------
# TYPE ALIASES
# -----------------------------------------------------------------------------

Pattern = Tuple[str, ...]
SegSpeeds = Tuple[Union[float, str], ...]


class SpeedRecord(TypedDict):
    """Structured record for segment speeds and summary metrics."""

    seg_speeds: SegSpeeds
    mean_mph: float | str


# =============================================================================
# FUNCTIONS
# =============================================================================


def parse_time_to_minutes(time_value: Optional[str]) -> Optional[int]:
    """Convert an ``HH:MM[:SS]`` time string to integer minutes past midnight.

    GTFS times may exceed 24:00 (e.g. ``"25:30:00"`` for a 1:30 AM trip on
    the following calendar day); those values are preserved as integers
    greater than or equal to 1440. Seconds, when present, are rounded to the
    nearest minute.

    Args:
        time_value: Time string such as ``"7:05"``, ``"07:05:00"``, or
            ``"26:30:00"``. Leading/trailing whitespace is ignored.
            Non-string or malformed values yield ``None``.

    Returns:
        Minutes since midnight, or ``None`` if the value cannot be parsed.
    """
    if not isinstance(time_value, str):
        return None
    parts = time_value.strip().split(":")
    if len(parts) not in (2, 3):
        return None
    try:
        hours = int(parts[0])
        minutes = int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
    except ValueError:
        return None
    if hours < 0 or not 0 <= minutes < 60 or not 0 <= seconds < 60:
        return None
    return hours * 60 + minutes + round(seconds / 60)


def minutes_to_hhmm(minutes: Optional[float], missing: str = "") -> str:
    """Convert minutes past midnight to a zero-padded ``HH:MM`` string.

    GTFS service days may exceed 24 hours, so values of 1440 minutes or more
    format with hours >= 24 (e.g. ``1590`` -> ``"26:30"``).

    Args:
        minutes: Minutes since midnight (may be fractional; rounded to the
            nearest minute). ``None`` and NaN yield ``missing``.
        missing: String returned for missing values, e.g. ``""`` or a
            sentinel such as ``"–"``.

    Returns:
        Zero-padded ``HH:MM`` string, or ``missing`` when *minutes* is
        ``None``/NaN.
    """
    if minutes is None or pd.isna(minutes):
        return missing
    hours, mins = divmod(int(round(minutes)), 60)
    return f"{hours:02d}:{mins:02d}"


def convert_distance(
    value: Any,
    input_unit: str,
    output_unit: Literal["miles", "km"] = "miles",
) -> Optional[float]:
    """Convert a distance value between transit-planning units.

    Args:
        value: Distance as a number or numeric string. ``None``, NaN, and
            empty/whitespace strings yield ``None``.
        input_unit: Unit of *value*: ``"feet"``, ``"meters"``, ``"km"``, or
            ``"miles"`` (case-insensitive).
        output_unit: Unit to convert to: ``"miles"`` or ``"km"``.

    Returns:
        The converted distance as a float, or ``None`` when *value* is
        missing or cannot be interpreted as a number.

    Raises:
        ValueError: If *input_unit* or *output_unit* is not a supported unit.
    """
    meters_per_input_unit = {"feet": 0.3048, "meters": 1.0, "km": 1000.0, "miles": 1609.344}
    meters_per_output_unit = {"miles": 1609.344, "km": 1000.0}

    input_factor = meters_per_input_unit.get(str(input_unit).strip().lower())
    if input_factor is None:
        raise ValueError(
            f"Unsupported input_unit {input_unit!r}; "
            f"expected one of {sorted(meters_per_input_unit)}."
        )
    output_factor = meters_per_output_unit.get(str(output_unit).strip().lower())
    if output_factor is None:
        raise ValueError(
            f"Unsupported output_unit {output_unit!r}; "
            f"expected one of {sorted(meters_per_output_unit)}."
        )

    if value is None or (isinstance(value, str) and not value.strip()):
        return None
    if pd.isna(value):
        return None
    try:
        numeric = float(value)
    except (TypeError, ValueError):
        return None
    return numeric * input_factor / output_factor


def mph(dist_miles: Optional[float], runtime_min: Optional[int]) -> Union[float, str]:
    """Compute miles-per-hour, guarding against divide-by-zero.

    Args:
        dist_miles: Segment length in miles.
        runtime_min: Segment runtime in minutes.

    Returns:
        Speed rounded to one decimal place, or :data:`MISSING_VAL` if inputs are
        not computable.
    """
    if dist_miles is None or runtime_min in (None, 0):
        return MISSING_VAL
    return round(dist_miles / (runtime_min / 60), 1)


def safe_sheet(name: str) -> str:
    """Sanitise an arbitrary string into a valid Excel worksheet name.

    Disallowed characters are replaced with underscores and the result is
    truncated to the Excel hard limit (31 chars).  An empty string becomes
    ``"Sheet"``.
    """
    cleaned = re.sub(r"[:\\/*?\[\]]", "_", name)[:31]
    return cleaned or "Sheet"


def _sum_numeric(seq: Iterable[object]) -> float:
    """Return the sum of numeric items, ignoring non-numeric elements."""
    return float(sum(x for x in seq if isinstance(x, (int, float))))


def load_gtfs_data(
    gtfs_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
    logger: Optional[logging.Logger] = None,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_path: Absolute or relative path to the folder containing the
            GTFS feed, or to a ``.zip`` archive of it — the form GTFS
            producers and most open-data portals distribute feeds in. Zip
            members may sit at the archive root or nested one level inside
            a single wrapper folder; both layouts are handled.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.
        logger: Logger for progress messages. Defaults to this module's
            logger (``logging.getLogger(__name__)``) rather than the root
            logger, so callers keep control of handler configuration.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Path missing, one of *files* not present in the feed, or
            an OS-level failure while reading a file.
        ValueError: *gtfs_path* is neither a directory nor a valid ``.zip``
            file, a requested file matches more than one location inside
            the zip, a file is empty, or the CSV parser fails.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    log = logger if logger is not None else logging.getLogger(__name__)

    if not os.path.exists(gtfs_path):
        raise OSError(f"The path '{gtfs_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    is_zip = os.path.isfile(gtfs_path) and gtfs_path.lower().endswith(".zip")
    if not is_zip and not os.path.isdir(gtfs_path):
        raise ValueError(f"'{gtfs_path}' is neither a directory nor a .zip file.")

    archive: zipfile.ZipFile | None = None
    members_by_name: dict[str, list[str]] = {}
    if is_zip:
        try:
            archive = zipfile.ZipFile(gtfs_path)
        except zipfile.BadZipFile as exc:
            raise ValueError(f"'{gtfs_path}' is not a valid zip archive.") from exc
        for name in archive.namelist():
            members_by_name.setdefault(os.path.basename(name), []).append(name)

    try:
        missing: list[str] = []
        ambiguous: list[str] = []
        resolved: dict[str, str] = {}
        for file_name in files:
            if archive is None:
                if not os.path.exists(os.path.join(gtfs_path, file_name)):
                    missing.append(file_name)
                continue
            candidates = members_by_name.get(file_name, [])
            if not candidates:
                missing.append(file_name)
            elif len(candidates) > 1:
                ambiguous.append(file_name)
            else:
                resolved[file_name] = candidates[0]

        if ambiguous:
            raise ValueError(
                f"Ambiguous GTFS files in '{gtfs_path}' (found in multiple "
                f"locations): {', '.join(ambiguous)}"
            )
        if missing:
            raise OSError(f"Missing GTFS files in '{gtfs_path}': {', '.join(missing)}")

        data: dict[str, pd.DataFrame] = {}
        for file_name in files:
            key = file_name.replace(".txt", "")
            try:
                if archive is None:
                    df = pd.read_csv(
                        os.path.join(gtfs_path, file_name), dtype=dtype, low_memory=False
                    )
                else:
                    with archive.open(resolved[file_name]) as handle:
                        df = pd.read_csv(handle, dtype=dtype, low_memory=False)
                data[key] = df
                log.info("Loaded %s (%d records).", file_name, len(df))

            except pd.errors.EmptyDataError as exc:
                raise ValueError(f"File '{file_name}' in '{gtfs_path}' is empty.") from exc

            except pd.errors.ParserError as exc:
                raise ValueError(f"Parser error in '{file_name}' in '{gtfs_path}': {exc}") from exc

        return data
    finally:
        if archive is not None:
            archive.close()


def segment_metrics(grp: pd.DataFrame) -> Tuple[SegSpeeds, float, int]:
    """Derive per-segment speeds and trip totals for one trip.

    Args:
        grp: Consecutive ``stop_times`` rows for a single trip, sorted by
            ``stop_sequence``.

    Returns:
        Tuple ``(seg_speeds, dist_total, runtime_total)`` where

        * **seg_speeds** – tuple of per-segment mph values (sentinel for gaps).
        * **dist_total** – total distance (miles).
        * **runtime_total** – total runtime (minutes).
    """
    times: List[Optional[int]] = []
    dists: List[Optional[float]] = []

    for _, row in grp.iterrows():
        times.append(
            parse_time_to_minutes(cast("str", row.get("departure_time") or row.get("arrival_time")))
        )
        dists.append(convert_distance(row.get("shape_dist_traveled"), INPUT_DISTANCE_UNIT, "miles"))

    run_min: List[Union[int, str]] = [MISSING_VAL]
    seg_mi: List[Union[float, str]] = [MISSING_VAL]
    seg_mph: List[Union[float, str]] = [MISSING_VAL]

    for (t0, t1), (d0, d1) in zip(
        zip(times, times[1:]),
        zip(dists, dists[1:]),
    ):
        run = None if t0 is None or t1 is None else t1 - t0
        dist = None if d0 is None or d1 is None else d1 - d0

        run_min.append(run if run is not None else MISSING_VAL)
        seg_mi.append(round(dist, 3) if dist is not None else MISSING_VAL)
        seg_mph.append(mph(dist, run))

    dist_tot = _sum_numeric(seg_mi[1:])
    time_tot = int(_sum_numeric(run_min[1:]))

    return tuple(seg_mph), dist_tot, time_tot


def build_index(
    gtfs: Dict[str, pd.DataFrame],
) -> Tuple[
    pd.DataFrame,
    Dict[int, Pattern],
    Dict[int, SpeedRecord],
    Dict[int, List[str]],
]:
    """Create lookup tables linking trips → patterns → speed signatures.

    The heavy-lifting step that

    1. Applies user-defined filters.
    2. Identifies unique stop-patterns and speed arrays.
    3. Generates a long trip index with start times for later banding.

    Args:
        gtfs: GTFS tables from :func:`load_gtfs`.

    Returns:
        ``index_df, pattern_lut, speed_lut, header_lut``.
    """
    trips = gtfs["trips"].merge(
        gtfs["routes"][["route_id", "route_short_name"]],
        on="route_id",
        how="left",
    )

    if FILTER_IN_ROUTE_SHORT_NAMES:
        trips = trips[trips["route_short_name"].isin(FILTER_IN_ROUTE_SHORT_NAMES)]
    if FILTER_OUT_ROUTE_SHORT_NAMES:
        trips = trips[~trips["route_short_name"].isin(FILTER_OUT_ROUTE_SHORT_NAMES)]
    if FILTER_IN_SERVICE_IDS:
        trips = trips[trips["service_id"].isin(FILTER_IN_SERVICE_IDS)]
    if FILTER_OUT_SERVICE_IDS:
        trips = trips[~trips["service_id"].isin(FILTER_OUT_SERVICE_IDS)]

    if trips.empty:
        return pd.DataFrame(), {}, {}, {}

    st = (
        gtfs["stop_times"]
        .merge(
            trips[["trip_id", "route_id", "service_id", "direction_id"]],
            on="trip_id",
            how="inner",
        )
        .sort_values(["trip_id", "stop_sequence"])
    )

    if EXPORT_TIMEPOINTS_ONLY and "timepoint" in st.columns:
        st = st[st["timepoint"] == 1]

    stop_meta = (
        gtfs["stops"][["stop_id", "stop_name", "stop_code"]].set_index("stop_id").to_dict("index")
    )

    pat_lut: Dict[int, Pattern] = {}
    speed_lut: Dict[int, SpeedRecord] = {}
    header_lut: Dict[int, List[str]] = {}
    rows: List[Dict[str, Union[int, str, float, None]]] = []

    for _trip_id, grp in st.groupby("trip_id"):
        pattern = tuple(cast("str", sid) for sid in grp["stop_id"])
        pat_hash = hash(pattern)
        pat_lut.setdefault(pat_hash, pattern)

        seg_speeds, dist_tot, time_tot = segment_metrics(grp)
        spd_hash = hash(seg_speeds)
        speed_lut.setdefault(
            spd_hash,
            {"seg_speeds": seg_speeds, "mean_mph": mph(dist_tot, time_tot)},
        )

        header_lut.setdefault(
            pat_hash,
            [
                f"{meta['stop_name']} ({meta['stop_code']})"
                if (meta := stop_meta.get(sid))
                else sid
                for sid in pattern
            ],
        )

        first_seg = seg_speeds[1] if len(seg_speeds) > 1 else MISSING_VAL
        start_min = parse_time_to_minutes(
            cast(
                "str",
                grp.iloc[0].get("departure_time") or grp.iloc[0].get("arrival_time"),
            )
        )

        rows.append(
            {
                "route_id": grp.iloc[0]["route_id"],
                "service_id": grp.iloc[0]["service_id"],
                "direction_id": grp.iloc[0]["direction_id"],
                "pattern_hash": pat_hash,
                "speed_hash": spd_hash,
                "first_seg_mph": None if first_seg == MISSING_VAL else cast("float", first_seg),
                "start": start_min,
            }
        )

    return pd.DataFrame(rows), pat_lut, speed_lut, header_lut


def band_rows(index_df: pd.DataFrame) -> pd.DataFrame:
    """Collapse individual trips into time-of-day bands.

    Args:
        index_df: Output from :func:`build_index`.

    Returns:
        One row per unique (route, service_id, direction, pattern_hash,
        speed_hash) × time-band, with first/last departure and trip count.
    """
    grouped = index_df.groupby(
        ["route_id", "service_id", "direction_id", "pattern_hash", "speed_hash"],
        dropna=False,
    )

    out = (
        grouped.agg(
            FrTime_min=("start", "min"),
            ToTime_min=("start", "max"),
            TripCount=("start", "count"),
        )
        .reset_index()
        .sort_values(["direction_id", "FrTime_min"])
    )

    out["FrTime"] = out["FrTime_min"].apply(lambda m: minutes_to_hhmm(m, MISSING_VAL))
    out["ToTime"] = out["ToTime_min"].apply(lambda m: minutes_to_hhmm(m, MISSING_VAL))
    return out.drop(columns=["FrTime_min", "ToTime_min"])


def export_excel(
    bands: pd.DataFrame,
    pat_lut: Dict[int, Pattern],
    speed_lut: Dict[int, SpeedRecord],
    header_lut: Dict[int, List[str]],
    routes: pd.DataFrame,
) -> None:
    """Write each (route × service_id) bundle to its own workbook.

    Side effects:
        * Creates :data:`OUTPUT_FOLDER` if missing.
        * Writes ``.xlsx`` files.
        * Logs progress via :pymod:`logging`.
    """
    OUTPUT_FOLDER.mkdir(parents=True, exist_ok=True)
    route_short = routes.set_index("route_id")["route_short_name"].to_dict()

    for (rid, sid), grp_rs in bands.groupby(["route_id", "service_id"], sort=False):
        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for did, grp_dir in grp_rs.groupby("direction_id", sort=False):
            ws = wb.create_sheet(safe_sheet(f"Dir_{did or 'X'}"))

            first_pat = cast("int", grp_dir.iloc[0]["pattern_hash"])
            header = ["Pattern", "FrTime", "ToTime", "Mean_mph", *header_lut[first_pat]]
            ws.append(header)

            for _, row in grp_dir.iterrows():
                speed_rec = speed_lut[cast("int", row["speed_hash"])]
                ws.append(
                    [
                        row["pattern_hash"],
                        row["FrTime"],
                        row["ToTime"],
                        speed_rec["mean_mph"],
                        *speed_rec["seg_speeds"],
                    ]
                )

            for col in range(1, len(header) + 1):
                ws.column_dimensions[get_column_letter(col)].width = 14

        out_path = OUTPUT_FOLDER / f"route_{route_short.get(rid, rid)}_cal{sid}_speed_table.xlsx"
        wb.save(out_path)
        logging.info("Wrote %s", out_path)


# =============================================================================
# MAIN
# =============================================================================


def main() -> None:
    """CLI entry-point – orchestrates GTFS load, processing, and Excel export."""
    logging.basicConfig(
        level=LOG_LEVEL,
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    if GTFS_FOLDER == Path(r"Path\To\Your\GTFS_Folder") or OUTPUT_FOLDER == Path(
        r"Path\To\Your\Output_Folder"
    ):
        logging.warning(
            "GTFS_FOLDER and/or OUTPUT_FOLDER are still set to their default placeholder values. "
            "Please update them in the CONFIGURATION section before running."
        )
        return

    logging.info("GTFS folder:   %s", GTFS_FOLDER)
    logging.info("Output folder: %s", OUTPUT_FOLDER)

    gtfs = load_gtfs_data(str(GTFS_FOLDER))
    index_df, pat_lut, speed_lut, header_lut = build_index(gtfs)

    if index_df.empty:
        logging.warning("No trips after filters – nothing to export.")
        return

    bands = band_rows(index_df)
    export_excel(bands, pat_lut, speed_lut, header_lut, gtfs["routes"])
    logging.info("Finished – %d band rows written.", len(bands))
    logging.info("Script completed successfully.")


if __name__ == "__main__":  # pragma: no cover
    main()
