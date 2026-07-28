"""Generate printable Excel schedules for each vehicle block in a GTFS feed.

The script reads the five core GTFS tables—``trips``, ``stop_times``, ``stops``,
``routes`` and ``calendar``—and optionally filters them by *service ID* and/or
*route short name*.  For every vehicle ``block_id`` that survives filtering it
produces a nicely-formatted ``.xlsx`` file ready for field auditing.

Outputs
-------
- ``block_<block_id>_schedule_printable.xlsx`` (one per surviving block, written
  to ``BASE_OUTPUT_PATH``): the block's stop-by-stop schedule with placeholder
  columns for handwritten field notes.
- ``printable_block_schedules_runlog.txt``: run-log sidecar capturing the
  verbatim CONFIGURATION block, the effective (CLI-resolved) settings, and a
  run summary.

Typical usage
-------------
Update the paths in the CONFIGURATION section (or pass ``--gtfs-dir`` /
``--output-dir``) and run from the command line, an ArcGIS Pro Python
toolbox, or a notebook. The effective configuration — config-block values
plus any CLI overrides — is echoed at INFO at startup so unintended
settings are visible before any work starts.

Key Features
------------
- Loads GTFS text files into ``pandas`` DataFrames with robust error handling.
- Converts ``HH:MM(:SS)`` time strings to seconds (and back) safely.
- Applies ergonomic Excel formatting via ``openpyxl`` (column widths, wrapping).
- Inserts placeholders for handwritten field notes (actual time, boardings, etc.).
"""

from __future__ import annotations

import argparse
import logging
import math
import os
import sys
import zipfile
from collections.abc import Mapping, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any, List, Optional, Union

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================
# === BEGIN CONFIG ===

_DEFAULT_GTFS_FOLDER_PATH = r"Path\To\Your\Input\Folder"
_DEFAULT_BASE_OUTPUT_PATH = r"Path\To\Your\Output\Folder"

GTFS_FOLDER_PATH = _DEFAULT_GTFS_FOLDER_PATH  # <<< EDIT HERE
BASE_OUTPUT_PATH = _DEFAULT_BASE_OUTPUT_PATH  # <<< EDIT HERE

REQUIRED_GTFS_FILES = [
    "trips.txt",
    "stop_times.txt",
    "stops.txt",
    "routes.txt",
    "calendar.txt",
]

# If you only want certain service IDs or route short names, specify them here:
FILTER_SERVICE_IDS: list[str] = []  # e.g. ["WKD", "SAT"]
FILTER_ROUTE_SHORT_NAMES: list[str] = []  # e.g. ["101", "202"]

# Placeholder values for printing:
MISSING_TIME = "________"
MISSING_VALUE = "_____"

# Maximum column width for neat Excel formatting:
MAX_COLUMN_WIDTH = 35

# When True, a failed run-log write aborts the script so outputs are never
# left untraced. Set False only for genuinely read-only output locations.
REQUIRE_RUN_LOG: bool = True

LOG_LEVEL: int = logging.INFO  # DEBUG / INFO / WARNING / ERROR

# === END CONFIG ===

# =============================================================================
# FUNCTIONS
# =============================================================================


def time_to_seconds(time_str: str) -> Union[float, int]:
    """Convert a ``HH:MM`` or ``HH:MM:SS`` string to total seconds.

    Args:
        time_str: Time string *or* ``NaN``; may exceed 24 h (e.g.,
            ``'25:10:00'`` → 1:10 a.m. next day).

    Returns:
        Non-negative number of seconds, or :pydata:`math.nan` on failure.
    """
    if pd.isna(time_str):
        return math.nan

    parts = time_str.strip().split(":")
    if len(parts) < 2:
        return math.nan

    try:
        hours = int(parts[0]) % 24  # Roll over hours >= 24
        minutes = int(parts[1])
        seconds = int(parts[2]) if len(parts) == 3 else 0
    except ValueError:
        return math.nan

    return hours * 3600 + minutes * 60 + seconds


def format_hhmm(total_seconds: Union[int, float]) -> str:
    """Render seconds since midnight as a ``HH:MM`` string.

    Args:
        total_seconds: Seconds since 00:00.  Negative or ``NaN`` returns
            an empty string.

    Returns:
        Two-digit hour and minute representation (24-hour clock).
    """
    if pd.isna(total_seconds) or total_seconds < 0:
        return ""
    hours = int(total_seconds // 3600)
    minutes = int((total_seconds % 3600) // 60)
    return f"{hours:02d}:{minutes:02d}"


# -----------------------------------------------------------------------------
# OTHER FUNCTIONS
# -----------------------------------------------------------------------------


def export_to_excel(data_frame: pd.DataFrame, output_file: str) -> None:
    """Write *data_frame* to an Excel file with basic styling.

    The sheet is named **Schedule** and receives:

    * Left-aligned cells.
    * Word-wrapped headers.
    * Column widths sized to longest cell (capped by ``MAX_COLUMN_WIDTH``).

    Args:
        data_frame: Tidy table to export; must be non-empty.
        output_file: Full path of the ``.xlsx`` file to create.

    Notes:
        ``os.makedirs`` is called with *exist_ok=True* so nested output
        folders are created automatically.
    """
    if data_frame.empty:
        logging.info("No data to export to %s", output_file)
        return

    os.makedirs(os.path.dirname(output_file), exist_ok=True)

    # Write DataFrame to Excel and access the worksheet object for formatting
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        data_frame.to_excel(writer, index=False, sheet_name="Schedule")
        worksheet = writer.sheets["Schedule"]

        # Adjust columns
        for col_i, col_name in enumerate(data_frame.columns, 1):
            col_letter = get_column_letter(col_i)

            # Header alignment and text wrap
            header_cell = worksheet[f"{col_letter}1"]
            header_cell.alignment = Alignment(horizontal="left", wrap_text=True)

            # Data alignment
            for row_i in range(2, worksheet.max_row + 1):
                cell = worksheet[f"{col_letter}{row_i}"]
                cell.alignment = Alignment(horizontal="left")

            # Set column width based on max content length, capped at MAX_COLUMN_WIDTH
            max_len = max(len(str(col_name)), 10)  # Minimum width
            for row_i in range(2, worksheet.max_row + 1):
                val = worksheet[f"{col_letter}{row_i}"].value
                if val is not None:
                    max_len = max(max_len, len(str(val)))
            worksheet.column_dimensions[col_letter].width = min(max_len + 2, MAX_COLUMN_WIDTH)

    logging.info("Exported: %s", output_file)


def filter_data(
    trips_df: pd.DataFrame,
    stop_times_df: pd.DataFrame,
    routes_df: pd.DataFrame,
    filter_route_short_names: Optional[Sequence[str]] = None,
    filter_service_ids: Optional[Sequence[str]] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Apply route and service filters, propagating them to stop times.

    Args:
        trips_df: Parsed **trips.txt** table.
        stop_times_df: Parsed **stop_times.txt** table.
        routes_df: Parsed **routes.txt** table (for ``route_short_name``).
        filter_route_short_names: Route short names whose blocks to keep;
            ``None`` falls back to the ``FILTER_ROUTE_SHORT_NAMES`` config
            constant, and an empty sequence keeps every route.
        filter_service_ids: Service IDs to keep; ``None`` falls back to the
            ``FILTER_SERVICE_IDS`` config constant, and an empty sequence
            keeps every service.

    Returns:
        ``(filtered_trips, filtered_stop_times)``.

    Raises:
        KeyError: If required columns are missing.

    Warning:
        If the filters remove every trip, the function returns two
        **empty** DataFrames.
    """
    if filter_route_short_names is None:
        filter_route_short_names = FILTER_ROUTE_SHORT_NAMES
    if filter_service_ids is None:
        filter_service_ids = FILTER_SERVICE_IDS

    # Merge route_short_name into trips
    routes_subset = routes_df[["route_id", "route_short_name"]]
    trips_df = trips_df.merge(routes_subset, on="route_id", how="left")

    # Apply Route Filtering
    if filter_route_short_names:
        blocks_for_selected_routes = (
            trips_df[trips_df["route_short_name"].isin(filter_route_short_names)]["block_id"]
            .dropna()
            .unique()
        )
        if len(blocks_for_selected_routes) == 0:
            logging.info("No blocks found with the specified route short names.")
            return pd.DataFrame(), pd.DataFrame()

        trips_df = trips_df[trips_df["block_id"].isin(blocks_for_selected_routes)]

    # Apply Service ID Filtering
    if filter_service_ids:
        trips_df = trips_df[trips_df["service_id"].isin(filter_service_ids)]

    # Filter stop_times to only include relevant trips
    stop_times_df = stop_times_df[stop_times_df["trip_id"].isin(trips_df["trip_id"])]

    return trips_df, stop_times_df


def prepare_stop_times(
    trips_df: pd.DataFrame, stop_times_df: pd.DataFrame, stops_df: pd.DataFrame
) -> pd.DataFrame:
    """Enrich and tidy ``stop_times`` for Excel export.

    Steps
    -----
    1. Ensure a numeric ``timepoint`` column (create if absent).
    2. Attach ``block_id``, ``route_short_name`` and ``direction_id``.
    3. Convert arrival/departure times → seconds → ``HH:MM`` format.
    4. Map ``stop_id`` → human-readable stop names.
    5. Sort by ``block_id``, ``trip_id``, ``stop_sequence``.

    Args:
        trips_df: Output of :pyfunc:`filter_data`.
        stop_times_df: Ditto.
        stops_df: Parsed **stops.txt** table.

    Returns:
        Cleaned ``stop_times`` DataFrame ready for grouping by block.
    """
    # If 'timepoint' does not exist, create a new column with 0.
    if "timepoint" not in stop_times_df.columns:
        stop_times_df["timepoint"] = 0
    else:
        # Convert to numeric, fill NaN with 0
        stop_times_df["timepoint"] = (
            pd.to_numeric(stop_times_df["timepoint"], errors="coerce").fillna(0).astype(int)
        )

    # Merge essential trip columns into stop_times
    needed_trip_cols = ["trip_id", "block_id", "route_short_name", "direction_id"]
    stop_times_df = stop_times_df.merge(trips_df[needed_trip_cols], on="trip_id", how="left")

    # Convert arrival/departure times to seconds and format
    stop_times_df["arrival_seconds"] = stop_times_df["arrival_time"].apply(time_to_seconds)
    stop_times_df["departure_seconds"] = stop_times_df["departure_time"].apply(time_to_seconds)
    stop_times_df["scheduled_time_hhmm"] = stop_times_df["departure_seconds"].apply(format_hhmm)

    # Merge in stop names
    stop_name_map = stops_df.set_index("stop_id")["stop_name"].to_dict()
    stop_times_df["stop_name"] = stop_times_df["stop_id"].map(stop_name_map).fillna("Unknown Stop")

    # Sort by block, trip, and stop_sequence
    stop_times_df = stop_times_df.dropna(subset=["block_id"])
    stop_times_df["stop_sequence"] = pd.to_numeric(stop_times_df["stop_sequence"], errors="coerce")
    stop_times_df = stop_times_df.dropna(subset=["stop_sequence"])
    stop_times_df = stop_times_df.sort_values(["block_id", "trip_id", "stop_sequence"])

    return stop_times_df


def export_blocks(stop_times_df: pd.DataFrame, base_output_path: Optional[str] = None) -> None:
    """Generate one Excel schedule per vehicle block.

    Args:
        stop_times_df: Prepared stop times (see
            :pyfunc:`prepare_stop_times`).  Must include the columns
            produced earlier (``block_id``, ``scheduled_time_hhmm``,
            etc.).
        base_output_path: Folder to write the per-block XLSX files into;
            ``None`` falls back to the ``BASE_OUTPUT_PATH`` config constant.

    Side Effects:
        Writes ``block_<id>_schedule_printable.xlsx`` to the output folder;
        creates the folder tree if needed.
    """
    if base_output_path is None:
        base_output_path = BASE_OUTPUT_PATH
    all_blocks = stop_times_df["block_id"].unique()
    logging.info("Found %d blocks to export.\n", len(all_blocks))

    for block_id in all_blocks:
        block_subset = stop_times_df[stop_times_df["block_id"] == block_id].copy()
        if block_subset.empty:
            continue

        # For each trip_id within this block, find earliest departure
        first_departures = (
            block_subset.groupby("trip_id")["departure_seconds"]
            .min()
            .reset_index(name="trip_start_seconds")
        )
        first_departures["trip_start_hhmm"] = first_departures["trip_start_seconds"].apply(
            format_hhmm
        )
        block_subset = block_subset.merge(first_departures, on="trip_id", how="left")

        block_subset["Trip Start Time"] = block_subset["trip_start_hhmm"]

        # Select and rename columns for clarity
        out_cols = [
            "block_id",
            "route_short_name",
            "direction_id",
            "trip_id",
            "Trip Start Time",
            "stop_sequence",
            "timepoint",
            "stop_id",
            "stop_name",
            "scheduled_time_hhmm",
        ]
        final_df = block_subset[out_cols].copy()
        final_df = final_df.rename(
            columns={
                "block_id": "Block ID",
                "route_short_name": "Route",
                "direction_id": "Direction",
                "trip_id": "Trip ID",
                "stop_sequence": "Stop Sequence",
                "timepoint": "Timepoint",
                "stop_id": "Stop ID",
                "stop_name": "Stop Name",
                "scheduled_time_hhmm": "Scheduled Time",
            },
        )

        # Insert placeholders
        final_df["Actual Time"] = MISSING_TIME
        final_df["Boardings"] = MISSING_VALUE
        final_df["Alightings"] = MISSING_VALUE
        final_df["Comments"] = MISSING_VALUE

        # Reorder columns to place 'Timepoint' after 'Stop Sequence'
        final_df = final_df[
            [
                "Block ID",
                "Route",
                "Direction",
                "Trip ID",
                "Trip Start Time",
                "Stop Sequence",
                "Timepoint",
                "Stop ID",
                "Stop Name",
                "Scheduled Time",
                "Actual Time",
                "Boardings",
                "Alightings",
                "Comments",
            ]
        ]

        final_df = final_df.sort_values(by=["Trip Start Time", "Trip ID", "Stop Sequence"])

        filename = f"block_{block_id}_schedule_printable.xlsx"
        output_path = os.path.join(base_output_path, filename)
        export_to_excel(final_df, output_path)


# -----------------------------------------------------------------------------
# REUSABLE FUNCTIONS
# -----------------------------------------------------------------------------


def load_gtfs_data(
    gtfs_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
    logger: Optional[logging.Logger] = None,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_path: Absolute or relative path to the folder containing the
            GTFS feed, or to a ``.zip`` archive of it — the form GTFS
            producers and most open-data portals distribute feeds in. Zip
            members may sit at the archive root or nested one level inside
            a single wrapper folder; both layouts are handled.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.
        logger: Logger for progress messages. Defaults to this module's
            logger (``logging.getLogger(__name__)``) rather than the root
            logger, so callers keep control of handler configuration.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Path missing, one of *files* not present in the feed, or
            an OS-level failure while reading a file.
        ValueError: *gtfs_path* is neither a directory nor a valid ``.zip``
            file, a requested file matches more than one location inside
            the zip, a file is empty, or the CSV parser fails.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    log = logger if logger is not None else logging.getLogger(__name__)

    if not os.path.exists(gtfs_path):
        raise OSError(f"The path '{gtfs_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    is_zip = os.path.isfile(gtfs_path) and gtfs_path.lower().endswith(".zip")
    if not is_zip and not os.path.isdir(gtfs_path):
        raise ValueError(f"'{gtfs_path}' is neither a directory nor a .zip file.")

    archive: zipfile.ZipFile | None = None
    members_by_name: dict[str, list[str]] = {}
    if is_zip:
        try:
            archive = zipfile.ZipFile(gtfs_path)
        except zipfile.BadZipFile as exc:
            raise ValueError(f"'{gtfs_path}' is not a valid zip archive.") from exc
        for name in archive.namelist():
            members_by_name.setdefault(os.path.basename(name), []).append(name)

    try:
        missing: list[str] = []
        ambiguous: list[str] = []
        resolved: dict[str, str] = {}
        for file_name in files:
            if archive is None:
                if not os.path.exists(os.path.join(gtfs_path, file_name)):
                    missing.append(file_name)
                continue
            candidates = members_by_name.get(file_name, [])
            if not candidates:
                missing.append(file_name)
            elif len(candidates) > 1:
                ambiguous.append(file_name)
            else:
                resolved[file_name] = candidates[0]

        if ambiguous:
            raise ValueError(
                f"Ambiguous GTFS files in '{gtfs_path}' (found in multiple "
                f"locations): {', '.join(ambiguous)}"
            )
        if missing:
            raise OSError(f"Missing GTFS files in '{gtfs_path}': {', '.join(missing)}")

        data: dict[str, pd.DataFrame] = {}
        for file_name in files:
            key = file_name.replace(".txt", "")
            try:
                if archive is None:
                    df = pd.read_csv(
                        os.path.join(gtfs_path, file_name), dtype=dtype, low_memory=False
                    )
                else:
                    with archive.open(resolved[file_name]) as handle:
                        df = pd.read_csv(handle, dtype=dtype, low_memory=False)
                data[key] = df
                log.info("Loaded %s (%d records).", file_name, len(df))

            except pd.errors.EmptyDataError as exc:
                raise ValueError(f"File '{file_name}' in '{gtfs_path}' is empty.") from exc

            except pd.errors.ParserError as exc:
                raise ValueError(f"Parser error in '{file_name}' in '{gtfs_path}': {exc}") from exc

        return data
    finally:
        if archive is not None:
            archive.close()


# ---- REUSABLE HELPERS (copied from utils/run_log.py) -----------------------


def extract_config_block(source_file: Path) -> str:
    r"""Return the text between the CONFIG markers in *source_file*.

    Reads ``source_file`` as UTF-8 text and slices out the lines strictly
    *between* the first occurrence of ``# === BEGIN CONFIG ===`` and the first
    subsequent occurrence of ``# === END CONFIG ===``.  The marker lines
    themselves are excluded; whitespace and inline comments inside the block
    are preserved verbatim.

    Args:
        source_file: Path to the Python source file to scan (typically
            ``Path(__file__)`` from the calling script).

    Returns:
        The verbatim text of the configuration block, joined with ``\n``.

    Raises:
        ValueError: If either marker is missing or they appear out of order.
        OSError: If ``source_file`` cannot be read.
    """
    _BEGIN = "# === BEGIN CONFIG ==="
    _END = "# === END CONFIG ==="

    lines: list[str] = source_file.read_text(encoding="utf-8").splitlines()

    begin_idx: int | None = None
    end_idx: int | None = None
    for i, line in enumerate(lines):
        stripped: str = line.strip()
        if begin_idx is None and stripped == _BEGIN:
            begin_idx = i
        elif begin_idx is not None and stripped == _END:
            end_idx = i
            break

    if begin_idx is None or end_idx is None:
        raise ValueError(
            f"Config markers not found in '{source_file}'. Expected '{_BEGIN}' and '{_END}'."
        )

    return "\n".join(lines[begin_idx + 1 : end_idx])


# ---- RUN LOG ----------------------------------------------------------------


def write_run_log(output_dir: Path, summary_lines: List[str]) -> bool:
    """Write the verbatim config block plus a run summary into *output_dir*.

    Returns:
        ``True`` if the log was written successfully, ``False`` otherwise.
    """
    log_path = output_dir / "printable_block_schedules_runlog.txt"
    try:
        config_text = extract_config_block(Path(__file__))
    except (OSError, ValueError) as exc:
        logging.error("Could not extract config block for run log: %s", exc)
        return False

    lines: List[str] = [
        "=" * 72,
        "PRINTABLE BLOCK SCHEDULES RUN LOG",
        "=" * 72,
        f"Run timestamp:    {datetime.now().isoformat(timespec='seconds')}",
        f"Output directory: {output_dir}",
        f"Source script:    {Path(__file__).resolve()}",
        "",
        "-" * 72,
        "RUN SUMMARY",
        "-" * 72,
        *summary_lines,
        "",
        "-" * 72,
        "CONFIGURATION (verbatim)",
        "-" * 72,
        "# === BEGIN CONFIG ===",
        config_text,
        "# === END CONFIG ===",
        "",
    ]
    try:
        log_path.write_text("\n".join(lines), encoding="utf-8")
    except OSError as exc:
        logging.error("Could not write run log '%s': %s", log_path, exc)
        return False
    logging.info("Run log written → %s", log_path)
    return True


# ---- REUSABLE HELPERS (copied from utils/cli_helpers.py) -------------------


def notebook_safe_argv(argv: Optional[Sequence[str]]) -> Optional[List[str]]:
    """Return the argv to parse, shielding notebook kernels from stray flags.

    When a script's ``main()`` runs with no explicit ``argv`` inside a
    Jupyter/IPython kernel, ``sys.argv`` holds kernel plumbing (for example
    ``-f /path/kernel.json``) rather than flags meant for the script, and
    strict ``argparse.parse_args`` would reject it and abort.  This helper
    detects the notebook case and substitutes an empty argument list so the
    CONFIGURATION constants stay in charge, while shell runs keep strict
    parsing (a typo in a flag fails loudly instead of being silently ignored).

    Canonical implementation: ``utils/cli_helpers.py``.

    Args:
        argv: Explicit argument list passed to ``main()``, or ``None`` to
            fall back to ``sys.argv``.

    Returns:
        ``list(argv)`` when *argv* was provided; ``[]`` when running inside a
        notebook kernel; otherwise ``None`` so argparse reads ``sys.argv[1:]``.
    """
    if argv is not None:
        return list(argv)
    if "ipykernel" in sys.modules:
        return []
    return None


def log_effective_config(parser: argparse.ArgumentParser, args: argparse.Namespace) -> None:
    """Log the resolved settings this run will actually use, at INFO.

    Prints one line per flag-exposed setting with its origin — ``config`` when
    the value is the CONFIGURATION-block default, ``CLI`` when a command-line
    flag changed it — so a stale config edit or a forgotten flag is visible in
    the console (and in notebook output) before any work starts. A flag passed
    with a value equal to its default is indistinguishable from the default and
    is labelled ``config``; the effective value is identical either way.

    Canonical implementation: ``utils/cli_helpers.py``.

    Args:
        parser: The parser that produced *args*; supplies the defaults.
        args: The parsed namespace holding the resolved values.
    """
    logging.info("Effective configuration (CONFIGURATION block + CLI overrides):")
    for name in sorted(vars(args)):
        value = getattr(args, name)
        origin = "CLI" if value != parser.get_default(name) else "config"
        logging.info("  --%-20s %-40s [%s]", name.replace("_", "-"), value, origin)


# =============================================================================
# MAIN
# =============================================================================


def build_arg_parser() -> argparse.ArgumentParser:
    """Build the CLI parser; every flag defaults to its CONFIGURATION constant."""
    parser = argparse.ArgumentParser(
        description=(
            "Generate one printable XLSX schedule per vehicle block in a GTFS feed, "
            "with placeholder columns for handwritten field notes. Defaults come "
            "from the configuration block at the top of this file."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--gtfs-dir",
        default=GTFS_FOLDER_PATH,
        help="Folder (or .zip) containing the GTFS feed.",
    )
    parser.add_argument(
        "--output-dir",
        default=BASE_OUTPUT_PATH,
        help="Folder for the per-block XLSX files.",
    )
    parser.add_argument(
        "--routes",
        nargs="*",
        default=FILTER_ROUTE_SHORT_NAMES,
        help="Route short names whose blocks to keep (default: all routes).",
    )
    parser.add_argument(
        "--service-ids",
        nargs="*",
        default=FILTER_SERVICE_IDS,
        help="Service IDs to keep (default: all services).",
    )
    parser.add_argument(
        "--log-level",
        default=logging.getLevelName(LOG_LEVEL),
        help="DEBUG / INFO / WARNING / ERROR.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    """Command-line entry point.

    Orchestrates:

    * Logging configuration.
    * Data ingestion via :pyfunc:`load_gtfs_data`.
    * Optional filtering (:pyfunc:`filter_data`).
    * Data preparation (:pyfunc:`prepare_stop_times`).
    * Per-block Excel export (:pyfunc:`export_blocks`).

    The function traps anticipated exceptions and logs them with useful
    context before exiting with a non-zero status. Defaults fall back to the
    config block at the top of this file.

    Returns:
        Process exit code: 0 on success, 1 on failure, 2 if required
        CONFIGURATION values are still placeholders.
    """
    parser = build_arg_parser()
    args = parser.parse_args(notebook_safe_argv(argv))
    logging.basicConfig(
        level=getattr(logging, str(args.log_level).upper(), LOG_LEVEL),
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    using_defaults = False
    if args.gtfs_dir == _DEFAULT_GTFS_FOLDER_PATH:
        logging.warning(
            "GTFS_FOLDER_PATH is still the default placeholder – update it in the "
            "CONFIGURATION section or pass --gtfs-dir before running: %s",
            _DEFAULT_GTFS_FOLDER_PATH,
        )
        using_defaults = True
    if args.output_dir == _DEFAULT_BASE_OUTPUT_PATH:
        logging.warning(
            "BASE_OUTPUT_PATH is still the default placeholder – update it in the "
            "CONFIGURATION section or pass --output-dir before running: %s",
            _DEFAULT_BASE_OUTPUT_PATH,
        )
        using_defaults = True
    if using_defaults:
        logging.info("No processing performed. Update the placeholder paths above and re-run.")
        return 2

    logging.info("========================================================")
    logging.info("GTFS Block Schedule Printable Generator")
    log_effective_config(parser, args)

    try:
        gtfs_data = load_gtfs_data(
            gtfs_path=args.gtfs_dir,
            files=REQUIRED_GTFS_FILES,
            dtype=str,
        )

        trips_df = gtfs_data["trips"]
        stop_times_df = gtfs_data["stop_times"]
        stops_df = gtfs_data["stops"]
        routes_df = gtfs_data["routes"]

        trips_df, stop_times_df = filter_data(
            trips_df,
            stop_times_df,
            routes_df,
            filter_route_short_names=args.routes,
            filter_service_ids=args.service_ids,
        )
        if trips_df.empty or stop_times_df.empty:
            logging.warning("No data remains after filtering – no files generated.")
            return 1

        prepared = prepare_stop_times(trips_df, stop_times_df, stops_df)
        if prepared.empty:
            logging.warning("No data remains after preparation – no files generated.")
            return 1

        export_blocks(prepared, base_output_path=args.output_dir)

        effective_lines = [
            f"  --{name.replace('_', '-'):<20} {getattr(args, name)!s:<40} "
            f"[{'CLI' if getattr(args, name) != parser.get_default(name) else 'config'}]"
            for name in sorted(vars(args))
        ]
        summary_lines = [
            f"GTFS feed:       {args.gtfs_dir}",
            f"Blocks exported: {prepared['block_id'].nunique()}",
            f"Output folder:   {args.output_dir}",
            "",
            "Effective configuration (CONFIGURATION block + CLI overrides):",
            *effective_lines,
        ]
        if not write_run_log(Path(args.output_dir), summary_lines) and REQUIRE_RUN_LOG:
            logging.error(
                "Run log could not be written to '%s' and REQUIRE_RUN_LOG is True.",
                args.output_dir,
            )
            return 1

        logging.info("Script finished successfully.")
        logging.info("Script completed successfully.")
        return 0

    except (OSError, ValueError) as err:
        logging.error("%s", err)
        return 1
    except Exception as err:  # catch-all for unforeseen issues
        logging.exception("Unexpected error: %s", err)
        return 1
    finally:
        logging.info("Exiting script.")


# Strict parsing; in a notebook, notebook_safe_argv() keeps the kernel's
# injected argv away from argparse so the config block stays in charge.
if __name__ == "__main__":
    raise SystemExit(main())
