"""Build a printable quick reference of GTFS service change dates from a feed archive.

Point ``FEEDS_DIR`` at a folder of GTFS feeds — zipped or unzipped, mixed
freely, e.g. every feed snapshot an agency has published — and the script
expands each feed's real active service dates, finds the dates where the
operated weekly service pattern durably changes (markups / picks / board
changes), reconciles those dates across overlapping feed snapshots, and
exports a formatted, print-ready XLSX quick reference a planner can pin to
a wall.

Detection is calendar-based. Each week of each feed is fingerprinted by
which service_ids run on which days of the week; a change is reported only
when a new fingerprint persists for ``MIN_STABLE_WEEKS``, so one-week
variations (holidays, special events) are logged as transient rather than
reported as service changes. A pick that changes only trip times never
touches calendar structure, so it is visible only when the archive contains
a new feed for it — the feed-succession boundary is then reported, with a
log note about the ambiguity. The script also cross-checks the feeds
against each other — agency names, timezones, declared ``feed_info`` ranges
vs. real active dates, coverage gaps, and overlapping snapshots that
disagree — and logs an actionable warning for every inconsistency it finds.

Inputs
------
- ``FEEDS_DIR``: a folder scanned (non-recursively) for GTFS feeds: ``.zip``
  archives and immediate subfolders holding GTFS text files. The folder may
  also itself be a single unzipped feed, or a single ``.zip``.
- Each feed needs ``calendar.txt`` and/or ``calendar_dates.txt``;
  ``agency.txt`` and ``feed_info.txt``, when present, feed the same-system
  and declared-date-range checks.

Outputs
-------
- ``service_change_quick_reference.xlsx``: sheet *Service Changes* (one row
  per service change date — day of week, services added/removed, source
  feeds, caveats) and sheet *Feeds* (one row per discovered feed — agency,
  version, declared vs. real active date range, data issues).
- ``service_changes.csv`` / ``service_change_feeds.csv``: the same two
  tables as machine-readable CSVs — full ``;``-separated service_id lists
  with no display truncation — for joining into other analyses.
- ``gtfs_service_change_dates_runlog.txt``: sidecar capturing the verbatim
  CONFIGURATION block, run timestamp, and a run summary.

Typical usage
-------------
Update the paths in the CONFIGURATION section (or pass ``--feeds-dir`` /
``--output-dir``) and run from a shell, ArcGIS Pro's Python window, or a
notebook. Set ``MAX_CHANGES`` / ``MAX_YEARS`` (or ``--max-changes`` /
``--max-years``) to limit the reference to, say, the last 10 changes or
the last five years.
"""

from __future__ import annotations

import argparse
import datetime as dt
import logging
import os
import sys
import zipfile
from collections import Counter
from collections.abc import Iterable, Mapping, Sequence
from datetime import datetime
from pathlib import Path
from typing import Any, List, NamedTuple, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# =============================================================================
# CONFIGURATION
# =============================================================================
# === BEGIN CONFIG ===

FEEDS_DIR: Path = Path(r"Path\To\Your\GTFS_Feeds_Folder")  # ←–– change me
OUTPUT_DIR: Path = Path(r"Path\To\Your\Output_Folder")  # ←–– change me
OUTPUT_FILENAME: str = r"service_change_quick_reference.xlsx"

# Alongside the printable XLSX, the same two tables are written as
# machine-readable CSVs: full service_id lists with no display truncation,
# and disputes in their own column — ready to join into other analyses.
CHANGES_CSV_FILENAME: str = r"service_changes.csv"
FEEDS_CSV_FILENAME: str = r"service_change_feeds.csv"

# Optional cutoffs for the printable reference. Leave both as None to list
# every service change detected in the archive.
#   MAX_CHANGES – keep only the most recent N service changes.
#   MAX_YEARS   – keep only changes within the last X years, anchored to the
#                 newest active service date found in the archive (not to
#                 today), so purely historical archives stay analyzable.
MAX_CHANGES: Optional[int] = None
MAX_YEARS: Optional[float] = None

# A new weekly service pattern must persist this many consecutive weeks to
# count as a service change. At the default of 2, one-week variations
# (holiday weeks, special events, storm shutdowns) are logged as transient
# and excluded. Raise it if multi-week specials (e.g. a two-week festival
# schedule) show up as changes; lower it to 1 to list every weekly
# variation, holidays included.
MIN_STABLE_WEEKS: int = 2

# When consecutive feeds leave more than this many days uncovered between
# one feed's last active date and the next feed's first, the boundary is
# reported as a coverage gap instead of a seamless hand-off. The default of
# 7 keeps an ordinary weekend (or a short holiday shutdown) between
# snapshots from being called a gap.
COVERAGE_GAP_DAYS: int = 7

# When True, a failed run-log write aborts the script so outputs are never
# left untraced. Set False only for genuinely read-only output locations.
REQUIRE_RUN_LOG: bool = True

LOG_LEVEL: int = logging.INFO  # DEBUG / INFO / WARNING / ERROR

# === END CONFIG ===

# Change events from different feeds within this many days of each other are
# treated as the same service change (feed exports often disagree by a day
# or two on when a pick formally begins).
_MERGE_TOLERANCE_DAYS: int = 3

# A feed "covers" a change date — and can therefore dispute it — only when
# its active range extends at least this many days on both sides of it.
_DISPUTE_COVERAGE_BUFFER_DAYS: int = 7

# Files whose presence marks a folder or zip as a candidate GTFS feed.
_GTFS_PROBE_FILES: tuple[str, ...] = (
    "agency.txt",
    "stops.txt",
    "routes.txt",
    "trips.txt",
    "stop_times.txt",
    "calendar.txt",
    "calendar_dates.txt",
)

_DAY_NAMES: tuple[str, ...] = (
    "Monday",
    "Tuesday",
    "Wednesday",
    "Thursday",
    "Friday",
    "Saturday",
    "Sunday",
)

# =============================================================================
# FUNCTIONS
# =============================================================================

# ---- REUSABLE HELPERS (copied from utils/gtfs_helpers.py) ------------------


def load_gtfs_data(
    gtfs_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
    logger: Optional[logging.Logger] = None,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_path: Absolute or relative path to the folder containing the
            GTFS feed, or to a ``.zip`` archive of it — the form GTFS
            producers and most open-data portals distribute feeds in. Zip
            members may sit at the archive root or nested one level inside
            a single wrapper folder; both layouts are handled.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.
        logger: Logger for progress messages. Defaults to this module's
            logger (``logging.getLogger(__name__)``) rather than the root
            logger, so callers keep control of handler configuration.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Path missing, one of *files* not present in the feed, or
            an OS-level failure while reading a file.
        ValueError: *gtfs_path* is neither a directory nor a valid ``.zip``
            file, a requested file matches more than one location inside
            the zip, a file is empty, or the CSV parser fails.

    Notes:
        All columns default to ``str`` to avoid pandas’ type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    log = logger if logger is not None else logging.getLogger(__name__)

    if not os.path.exists(gtfs_path):
        raise OSError(f"The path '{gtfs_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    is_zip = os.path.isfile(gtfs_path) and gtfs_path.lower().endswith(".zip")
    if not is_zip and not os.path.isdir(gtfs_path):
        raise ValueError(f"'{gtfs_path}' is neither a directory nor a .zip file.")

    archive: zipfile.ZipFile | None = None
    members_by_name: dict[str, list[str]] = {}
    if is_zip:
        try:
            archive = zipfile.ZipFile(gtfs_path)
        except zipfile.BadZipFile as exc:
            raise ValueError(f"'{gtfs_path}' is not a valid zip archive.") from exc
        for name in archive.namelist():
            members_by_name.setdefault(os.path.basename(name), []).append(name)

    try:
        missing: list[str] = []
        ambiguous: list[str] = []
        resolved: dict[str, str] = {}
        for file_name in files:
            if archive is None:
                if not os.path.exists(os.path.join(gtfs_path, file_name)):
                    missing.append(file_name)
                continue
            candidates = members_by_name.get(file_name, [])
            if not candidates:
                missing.append(file_name)
            elif len(candidates) > 1:
                ambiguous.append(file_name)
            else:
                resolved[file_name] = candidates[0]

        if ambiguous:
            raise ValueError(
                f"Ambiguous GTFS files in '{gtfs_path}' (found in multiple "
                f"locations): {', '.join(ambiguous)}"
            )
        if missing:
            raise OSError(f"Missing GTFS files in '{gtfs_path}': {', '.join(missing)}")

        data: dict[str, pd.DataFrame] = {}
        for file_name in files:
            key = file_name.replace(".txt", "")
            try:
                if archive is None:
                    df = pd.read_csv(
                        os.path.join(gtfs_path, file_name), dtype=dtype, low_memory=False
                    )
                else:
                    with archive.open(resolved[file_name]) as handle:
                        df = pd.read_csv(handle, dtype=dtype, low_memory=False)
                data[key] = df
                log.info("Loaded %s (%d records).", file_name, len(df))

            except pd.errors.EmptyDataError as exc:
                raise ValueError(f"File '{file_name}' in '{gtfs_path}' is empty.") from exc

            except pd.errors.ParserError as exc:
                raise ValueError(f"Parser error in '{file_name}' in '{gtfs_path}': {exc}") from exc

        return data
    finally:
        if archive is not None:
            archive.close()


# ---- REUSABLE HELPERS (copied from utils/calendar_helpers.py) --------------


def expand_service_active_dates(
    calendar_df: Optional[pd.DataFrame],
    calendar_dates_df: Optional[pd.DataFrame] = None,
    max_days_per_service: int = 1830,
    today: Optional[dt.date] = None,
) -> dict[str, set[dt.date]]:
    """Expand each service_id into its real set of active calendar dates.

    Builds the base date set from each ``calendar.txt`` row (day-of-week
    pattern × ``start_date``–``end_date`` range), then applies
    ``calendar_dates.txt`` exceptions (``exception_type`` 1 adds a date,
    2 removes it). Handles calendar_dates-only feeds (*calendar_df* empty or
    ``None``), redundant additions, and fully negated base patterns — the
    returned sets reflect only the dates a service truly operates.

    Rows with unparseable or reversed dates are skipped with a warning.
    A date range longer than *max_days_per_service* (a common placeholder
    pattern, e.g. 2000–2099) is clamped to a window of that length centred
    on *today* and logged, so expansion stays fast and downstream per-year
    statistics stay meaningful.

    Args:
        calendar_df: Parsed ``calendar.txt``, or ``None`` if the feed has
            none. Expected columns: ``service_id``, the seven day-of-week
            flags, ``start_date``, ``end_date``.
        calendar_dates_df: Parsed ``calendar_dates.txt`` or ``None``.
            Expected columns: ``service_id``, ``date``, ``exception_type``.
        max_days_per_service: Longest date range expanded per service before
            clamping kicks in. The default (1830 ≈ 5 years) is far beyond
            any real service span but well short of placeholder ranges.
        today: Anchor date for clamping oversized ranges. Defaults to the
            current date; pass a fixed date for deterministic tests.

    Returns:
        Mapping of ``service_id`` (as ``str``) to the set of dates the
        service operates. Services whose dates never parse map to an empty
        set rather than being dropped, so callers can report them.

    Raises:
        ValueError: If *calendar_df* is provided but lacks ``service_id``,
            ``start_date``, or ``end_date`` columns.
    """
    day_cols = ("monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday")
    anchor = dt.date.today() if today is None else today
    active: dict[str, set[dt.date]] = {}

    if calendar_df is not None and not calendar_df.empty:
        required = {"service_id", "start_date", "end_date"}
        missing = required - set(calendar_df.columns)
        if missing:
            raise ValueError(f"calendar.txt is missing required column(s): {sorted(missing)}")
        for _, row in calendar_df.iterrows():
            sid = str(row["service_id"]).strip()
            try:
                start = dt.datetime.strptime(str(row["start_date"]).strip(), "%Y%m%d").date()
                end = dt.datetime.strptime(str(row["end_date"]).strip(), "%Y%m%d").date()
            except ValueError:
                logging.warning("Service %s: unparseable start/end date — skipping row.", sid)
                active.setdefault(sid, set())
                continue
            if end < start:
                logging.warning(
                    "Service %s: end_date %s precedes start_date %s — skipping row.",
                    sid,
                    end,
                    start,
                )
                active.setdefault(sid, set())
                continue
            if (end - start).days + 1 > max_days_per_service:
                half = max_days_per_service // 2
                clamped_start = max(start, anchor - dt.timedelta(days=half))
                clamped_end = min(end, anchor + dt.timedelta(days=half))
                logging.warning(
                    "Service %s: date range %s–%s looks like a placeholder; "
                    "clamping expansion to %s–%s.",
                    sid,
                    start,
                    end,
                    clamped_start,
                    clamped_end,
                )
                start, end = clamped_start, clamped_end
            pattern = [str(row.get(c, "0")).strip() == "1" for c in day_cols]
            dates = active.setdefault(sid, set())
            d = start
            while d <= end:
                if pattern[d.weekday()]:
                    dates.add(d)
                d += dt.timedelta(days=1)

    if calendar_dates_df is not None and not calendar_dates_df.empty:
        bad_rows = 0
        for _, row in calendar_dates_df.iterrows():
            sid = str(row["service_id"]).strip()
            try:
                d = dt.datetime.strptime(str(row["date"]).strip(), "%Y%m%d").date()
            except ValueError:
                bad_rows += 1
                continue
            etype = str(row.get("exception_type", "")).strip()
            dates = active.setdefault(sid, set())
            if etype == "1":
                dates.add(d)
            elif etype == "2":
                dates.discard(d)
            else:
                bad_rows += 1
        if bad_rows:
            logging.warning(
                "calendar_dates.txt: skipped %d row(s) with unparseable date/exception_type.",
                bad_rows,
            )

    return active


# ---- REUSABLE HELPERS (copied from utils/run_log.py) -----------------------


def extract_config_block(source_file: Path) -> str:
    r"""Return the text between the CONFIG markers in *source_file*.

    Reads ``source_file`` as UTF-8 text and slices out the lines strictly
    *between* the first occurrence of ``# === BEGIN CONFIG ===`` and the first
    subsequent occurrence of ``# === END CONFIG ===``.  The marker lines
    themselves are excluded; whitespace and inline comments inside the block
    are preserved verbatim.

    Args:
        source_file: Path to the Python source file to scan (typically
            ``Path(__file__)`` from the calling script).

    Returns:
        The verbatim text of the configuration block, joined with ``\n``.

    Raises:
        ValueError: If either marker is missing or they appear out of order.
        OSError: If ``source_file`` cannot be read.
    """
    _BEGIN = "# === BEGIN CONFIG ==="
    _END = "# === END CONFIG ==="

    lines: list[str] = source_file.read_text(encoding="utf-8").splitlines()

    begin_idx: int | None = None
    end_idx: int | None = None
    for i, line in enumerate(lines):
        stripped: str = line.strip()
        if begin_idx is None and stripped == _BEGIN:
            begin_idx = i
        elif begin_idx is not None and stripped == _END:
            end_idx = i
            break

    if begin_idx is None or end_idx is None:
        raise ValueError(
            f"Config markers not found in '{source_file}'. Expected '{_BEGIN}' and '{_END}'."
        )

    return "\n".join(lines[begin_idx + 1 : end_idx])


# ---- REUSABLE HELPERS (copied from utils/cli_helpers.py) -------------------


def notebook_safe_argv(argv: Optional[Sequence[str]]) -> Optional[List[str]]:
    """Return the argv to parse, shielding notebook kernels from stray flags.

    When a script's ``main()`` runs with no explicit ``argv`` inside a
    Jupyter/IPython kernel, ``sys.argv`` holds kernel plumbing (for example
    ``-f /path/kernel.json``) rather than flags meant for the script, and
    strict ``argparse.parse_args`` would reject it and abort.  This helper
    detects the notebook case and substitutes an empty argument list so the
    CONFIGURATION constants stay in charge, while shell runs keep strict
    parsing (a typo in a flag fails loudly instead of being silently ignored).

    Canonical implementation: ``utils/cli_helpers.py``.

    Args:
        argv: Explicit argument list passed to ``main()``, or ``None`` to
            fall back to ``sys.argv``.

    Returns:
        ``list(argv)`` when *argv* was provided; ``[]`` when running inside a
        notebook kernel; otherwise ``None`` so argparse reads ``sys.argv[1:]``.
    """
    if argv is not None:
        return list(argv)
    if "ipykernel" in sys.modules:
        return []
    return None


# ---- SCRIPT-SPECIFIC TYPES --------------------------------------------------


class FeedSummary(NamedTuple):
    """Identity, active-date range, and data issues for one discovered feed."""

    label: str
    path: Path
    usable: bool
    active: dict[str, set[dt.date]]
    first_active: Optional[dt.date]
    last_active: Optional[dt.date]
    agency_names: tuple[str, ...]
    timezones: tuple[str, ...]
    feed_version: str
    feed_publisher: str
    declared_start: Optional[dt.date]
    declared_end: Optional[dt.date]
    issues: list[str]


class Regime(NamedTuple):
    """A run of consecutive weeks sharing one weekly service pattern."""

    start_week: dt.date
    end_week: dt.date
    weeks: int
    pattern: frozenset[tuple[str, int]]


class ServiceChange(NamedTuple):
    """One detected service change, within a feed or at a feed boundary."""

    date: dt.date
    feeds: tuple[str, ...]
    kind: str
    added: frozenset[str]
    removed: frozenset[str]
    note: str = ""


class MergedChange(NamedTuple):
    """A service change reconciled across every feed that evidences it."""

    date: dt.date
    kind: str
    added: frozenset[str]
    removed: frozenset[str]
    feeds: tuple[str, ...]
    notes: tuple[str, ...]
    disputed_by: tuple[str, ...] = ()


# ---- FEED DISCOVERY ---------------------------------------------------------


def _looks_like_gtfs_dir(path: Path) -> bool:
    """Return ``True`` when *path* directly contains at least one core GTFS file."""
    return any((path / name).is_file() for name in _GTFS_PROBE_FILES)


def _looks_like_gtfs_zip(path: Path) -> bool:
    """Return ``True`` when the zip contains at least one core GTFS file (any depth)."""
    try:
        with zipfile.ZipFile(path) as archive:
            members = {os.path.basename(name) for name in archive.namelist()}
    except zipfile.BadZipFile:
        logging.warning("Skipping '%s': not a valid zip archive.", path.name)
        return False
    return any(name in members for name in _GTFS_PROBE_FILES)


def discover_feeds(feeds_dir: Path) -> list[Path]:
    """Find candidate GTFS feeds (zip archives or folders) inside *feeds_dir*.

    The scan is non-recursive: immediate subfolders and ``.zip`` files are
    considered. A folder or zip qualifies as a candidate when it contains at
    least one core GTFS file, so feeds with missing calendars are still
    surfaced (and reported) rather than silently ignored. When *feeds_dir*
    itself is a GTFS feed (or a single zip), it is analyzed directly.

    Args:
        feeds_dir: Folder to scan, or a single feed folder / ``.zip``.

    Returns:
        Sorted list of feed paths (each a folder or a ``.zip``).

    Raises:
        OSError: If *feeds_dir* does not exist.
        ValueError: If no candidate GTFS feed is found.
    """
    if not feeds_dir.exists():
        raise OSError(f"The path '{feeds_dir}' does not exist.")

    if feeds_dir.is_file():
        if feeds_dir.suffix.lower() == ".zip" and _looks_like_gtfs_zip(feeds_dir):
            logging.info("'%s' is itself a GTFS zip — analyzing it as a single feed.", feeds_dir)
            return [feeds_dir]
        raise ValueError(f"'{feeds_dir}' is neither a folder nor a GTFS .zip archive.")

    if _looks_like_gtfs_dir(feeds_dir):
        logging.info("'%s' itself contains GTFS files — analyzing it as a single feed.", feeds_dir)
        return [feeds_dir]

    feeds: list[Path] = []
    for child in sorted(feeds_dir.iterdir(), key=lambda p: p.name.lower()):
        if child.is_dir():
            if _looks_like_gtfs_dir(child):
                feeds.append(child)
            else:
                logging.info("Skipping folder '%s': no GTFS files at its top level.", child.name)
        elif child.suffix.lower() == ".zip":
            if _looks_like_gtfs_zip(child):
                feeds.append(child)
            else:
                logging.warning("Skipping '%s': contains no GTFS files.", child.name)
    if not feeds:
        raise ValueError(
            f"No GTFS feeds (zip archives or folders with GTFS files) found in '{feeds_dir}'."
        )
    logging.info("Discovered %d candidate feed(s) in '%s'.", len(feeds), feeds_dir)
    return feeds


# ---- PER-FEED INSPECTION ----------------------------------------------------


def _load_optional(gtfs_path: Path, file_name: str) -> Optional[pd.DataFrame]:
    """Load one GTFS file via ``load_gtfs_data``, returning ``None`` if absent."""
    try:
        return load_gtfs_data(str(gtfs_path), files=(file_name,))[file_name[:-4]]
    except (OSError, ValueError) as exc:
        logging.info(
            "%s not usable in '%s' (%s) — continuing without it.", file_name, gtfs_path, exc
        )
        return None


def _cell_str(row: pd.Series, column: str) -> str:
    """Return a stripped string cell value, mapping missing/NaN to ``''``."""
    value = row.get(column, "")
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    return str(value).strip()


def _parse_gtfs_date(value: Any) -> Optional[dt.date]:
    """Parse a GTFS ``YYYYMMDD`` date; return ``None`` when blank or unparseable."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return dt.datetime.strptime(text, "%Y%m%d").date()
    except ValueError:
        return None


def inspect_feed(feed_path: Path) -> FeedSummary:
    """Summarize one feed: identity, real active dates, and data issues.

    Loads ``calendar.txt`` / ``calendar_dates.txt`` (either may be absent),
    expands every service_id's true active dates, and cross-checks the
    result against the range ``feed_info.txt`` declares. Every problem found
    is logged as a warning *and* recorded on the summary's ``issues`` list so
    it lands in the Feeds sheet of the printed reference.

    Args:
        feed_path: Path to a feed folder or ``.zip``.

    Returns:
        A :class:`FeedSummary`; ``usable`` is ``False`` when no active
        service dates could be derived.
    """
    label = feed_path.stem if feed_path.suffix.lower() == ".zip" else feed_path.name
    issues: list[str] = []

    calendar = _load_optional(feed_path, "calendar.txt")
    calendar_dates = _load_optional(feed_path, "calendar_dates.txt")
    agency = _load_optional(feed_path, "agency.txt")
    feed_info = _load_optional(feed_path, "feed_info.txt")

    agency_names: tuple[str, ...] = ()
    timezones: tuple[str, ...] = ()
    if agency is not None and "agency_name" in agency.columns:
        agency_names = tuple(
            sorted({str(n).strip() for n in agency["agency_name"].dropna() if str(n).strip()})
        )
    if agency is not None and "agency_timezone" in agency.columns:
        timezones = tuple(
            sorted({str(z).strip() for z in agency["agency_timezone"].dropna() if str(z).strip()})
        )

    feed_version = ""
    feed_publisher = ""
    declared_start: Optional[dt.date] = None
    declared_end: Optional[dt.date] = None
    if feed_info is not None and not feed_info.empty:
        first_row = feed_info.iloc[0]
        feed_version = _cell_str(first_row, "feed_version")
        feed_publisher = _cell_str(first_row, "feed_publisher_name")
        declared_start = _parse_gtfs_date(first_row.get("feed_start_date"))
        declared_end = _parse_gtfs_date(first_row.get("feed_end_date"))

    if calendar is None and calendar_dates is None:
        issues.append("no calendar.txt or calendar_dates.txt — active service dates unavailable")
        logging.warning("Feed '%s': %s.", label, issues[-1])
        return FeedSummary(
            label,
            feed_path,
            False,
            {},
            None,
            None,
            agency_names,
            timezones,
            feed_version,
            feed_publisher,
            declared_start,
            declared_end,
            issues,
        )

    active = expand_service_active_dates(calendar, calendar_dates)
    all_dates: set[dt.date] = set()
    for dates in active.values():
        all_dates |= dates
    if not all_dates:
        issues.append("calendar files are present but contain no parseable active service dates")
        logging.warning("Feed '%s': %s.", label, issues[-1])
        return FeedSummary(
            label,
            feed_path,
            False,
            active,
            None,
            None,
            agency_names,
            timezones,
            feed_version,
            feed_publisher,
            declared_start,
            declared_end,
            issues,
        )

    first_active, last_active = min(all_dates), max(all_dates)
    if declared_start is not None and first_active < declared_start:
        issues.append(
            f"active service starts {first_active}, before the declared "
            f"feed_start_date {declared_start}"
        )
    if declared_end is not None and last_active > declared_end:
        issues.append(
            f"active service ends {last_active}, after the declared feed_end_date {declared_end}"
        )
    if declared_end is not None and (declared_end - last_active).days > 30:
        issues.append(
            f"feed_info declares service through {declared_end} but the last active date "
            f"is {last_active} — the calendar may end early"
        )
    for issue in issues:
        logging.warning("Feed '%s': %s.", label, issue)
    logging.info(
        "Feed '%s': %d service_id(s), active %s → %s%s.",
        label,
        len(active),
        first_active,
        last_active,
        f", agency: {', '.join(agency_names)}" if agency_names else "",
    )
    return FeedSummary(
        label,
        feed_path,
        True,
        active,
        first_active,
        last_active,
        agency_names,
        timezones,
        feed_version,
        feed_publisher,
        declared_start,
        declared_end,
        issues,
    )


def check_system_consistency(summaries: Sequence[FeedSummary]) -> str:
    """Cross-check agency identity across feeds and warn when they diverge.

    Feeds whose agency names share nothing with the rest of the folder — and
    folders whose feeds span several agency timezones — very likely mix
    different transit systems, which would make the merged change timeline
    meaningless. Each suspect feed gets a warning and a Feeds-sheet issue.

    Args:
        summaries: All inspected feeds (unusable ones are ignored).

    Returns:
        Display name for the system (the most common agency name), used in
        the printed title; ``"GTFS archive"`` when no agency name exists.
    """
    usable = [s for s in summaries if s.usable]
    named = {
        s.label: {" ".join(n.lower().split()) for n in s.agency_names}
        for s in usable
        if s.agency_names
    }
    for summary in usable:
        mine = named.get(summary.label, set())
        if not mine:
            summary.issues.append(
                "no agency_name available — cannot confirm this feed belongs to the same system"
            )
            logging.warning("Feed '%s': %s.", summary.label, summary.issues[-1])
            continue
        others: set[str] = set()
        for other_label, names in named.items():
            if other_label != summary.label:
                others |= names
        if others and mine.isdisjoint(others):
            summary.issues.append(
                f"agency name(s) {', '.join(summary.agency_names)} match nothing in the rest "
                "of the folder — this feed may not be for the same system"
            )
            logging.warning("Feed '%s': %s.", summary.label, summary.issues[-1])
    zones = {zone for s in usable for zone in s.timezones}
    if len(zones) > 1:
        logging.warning(
            "Feeds span multiple agency_timezones (%s) — they may not describe the same system.",
            ", ".join(sorted(zones)),
        )
    name_counts = Counter(name for s in usable for name in s.agency_names)
    if name_counts:
        return name_counts.most_common(1)[0][0]
    return "GTFS archive"


# ---- SERVICE CHANGE DETECTION -----------------------------------------------


def _service_ids_by_date(active_dates: Mapping[str, set[dt.date]]) -> dict[dt.date, frozenset[str]]:
    """Invert service_id → active dates into date → active service_ids."""
    mutable: dict[dt.date, set[str]] = {}
    for sid, dates in active_dates.items():
        for day in dates:
            mutable.setdefault(day, set()).add(sid)
    return {day: frozenset(sids) for day, sids in mutable.items()}


def weekly_service_patterns(
    services_by_date: Mapping[dt.date, frozenset[str]],
) -> dict[dt.date, frozenset[tuple[str, int]]]:
    """Fingerprint each Monday-aligned week as its set of (service_id, weekday) pairs.

    Comparing whole-week fingerprints instead of day-by-day service sets is
    what keeps ordinary weekday/Saturday/Sunday cycling from reading as
    seven "changes" a week: two weeks match only when the same services run
    on the same days of the week.

    Args:
        services_by_date: Output of :func:`_service_ids_by_date`.

    Returns:
        Mapping of week start (a Monday) → fingerprint, covering every week
        from the first to the last active date. Weeks with no service at all
        map to an empty fingerprint, so suspensions are visible.
    """
    if not services_by_date:
        return {}
    first, last = min(services_by_date), max(services_by_date)
    week = first - dt.timedelta(days=first.weekday())
    last_week = last - dt.timedelta(days=last.weekday())
    patterns: dict[dt.date, frozenset[tuple[str, int]]] = {}
    while week <= last_week:
        entries: set[tuple[str, int]] = set()
        for offset in range(7):
            day = week + dt.timedelta(days=offset)
            for sid in services_by_date.get(day, frozenset()):
                entries.add((sid, offset))
        patterns[week] = frozenset(entries)
        week += dt.timedelta(days=7)
    return patterns


def stable_regimes(
    weekly_patterns: Mapping[dt.date, frozenset[tuple[str, int]]],
    min_stable_weeks: int = MIN_STABLE_WEEKS,
    feed_label: str = "feed",
) -> list[Regime]:
    """Compress weekly fingerprints into runs and keep the lasting ones.

    Runs shorter than *min_stable_weeks* — holiday weeks, special events,
    and the partial weeks at a feed's edges — are logged as transient and
    dropped, so they never masquerade as service changes.

    Args:
        weekly_patterns: Output of :func:`weekly_service_patterns`.
        min_stable_weeks: Minimum run length (in weeks) to keep.
        feed_label: Feed name used in log messages.

    Returns:
        Chronological list of stable :class:`Regime` runs.
    """
    runs: list[Regime] = []
    for week in sorted(weekly_patterns):
        pattern = weekly_patterns[week]
        if runs and runs[-1].pattern == pattern:
            last = runs[-1]
            runs[-1] = last._replace(end_week=week, weeks=last.weeks + 1)
        else:
            runs.append(Regime(week, week, 1, pattern))
    stable: list[Regime] = []
    for run in runs:
        if run.weeks >= min_stable_weeks:
            stable.append(run)
        else:
            logging.info(
                "Feed '%s': transient weekly variation for %d week(s) starting %s (likely a "
                "holiday or special service) — not counted as a service change.",
                feed_label,
                run.weeks,
                run.start_week,
            )
    return stable


def _pattern_day_map(pattern: frozenset[tuple[str, int]]) -> dict[int, frozenset[str]]:
    """Convert a weekly fingerprint into weekday → active service_id set."""
    by_dow: dict[int, set[str]] = {}
    for sid, dow in pattern:
        by_dow.setdefault(dow, set()).add(sid)
    return {dow: frozenset(sids) for dow, sids in by_dow.items()}


def _weekly_signature(pattern: frozenset[tuple[str, int]]) -> tuple[tuple[int, ...], ...]:
    """Reduce a weekly fingerprint to its service_id-free day-of-week structure.

    Two fingerprints share a signature when their services run on the same
    day-of-week combinations, regardless of what the service_ids are called
    — the test that tells "same pick, renamed ids" apart from a genuine
    restructuring when comparing feeds across a boundary.
    """
    by_sid: dict[str, set[int]] = {}
    for sid, dow in pattern:
        by_sid.setdefault(sid, set()).add(dow)
    return tuple(sorted(tuple(sorted(days)) for days in by_sid.values()))


def _change_date_between(
    services_by_date: Mapping[dt.date, frozenset[str]],
    prev: Regime,
    new: Regime,
) -> dt.date:
    """Pin down the first date the old pattern stops matching and the new one starts.

    Scans day by day from the end of the previous stable regime through the
    first week of the new one, so mid-week markups land on the true day and
    holidays inside the transition window (matching neither pattern) are
    skipped. Falls back to the new regime's first Monday.
    """
    prev_map = _pattern_day_map(prev.pattern)
    new_map = _pattern_day_map(new.pattern)
    day = prev.end_week + dt.timedelta(days=7)
    last = new.start_week + dt.timedelta(days=6)
    while day <= last:
        sids = services_by_date.get(day, frozenset())
        dow = day.weekday()
        if sids != prev_map.get(dow, frozenset()) and sids == new_map.get(dow, frozenset()):
            return day
        day += dt.timedelta(days=1)
    return new.start_week


def analyze_feed_changes(
    active_dates: Mapping[str, set[dt.date]],
    min_stable_weeks: int = MIN_STABLE_WEEKS,
    feed_label: str = "feed",
) -> tuple[list[ServiceChange], list[Regime]]:
    """Detect the lasting service changes within one feed.

    Args:
        active_dates: Output of :func:`expand_service_active_dates`.
        min_stable_weeks: Weeks a new pattern must persist to count.
        feed_label: Feed name used in events and log messages.

    Returns:
        Tuple of (chronological change events, the feed's stable regimes —
        the latter reused for cross-feed boundary comparison).
    """
    services_by_date = _service_ids_by_date(active_dates)
    patterns = weekly_service_patterns(services_by_date)
    regimes = stable_regimes(patterns, min_stable_weeks, feed_label)
    events: list[ServiceChange] = []
    if not regimes:
        if patterns:
            logging.warning(
                "Feed '%s': no weekly service pattern lasts %d week(s) or more — the feed is "
                "too short or too irregular for change detection.",
                feed_label,
                min_stable_weeks,
            )
        return events, regimes
    prev = regimes[0]
    for new in regimes[1:]:
        if new.pattern == prev.pattern:
            logging.info(
                "Feed '%s': the weekly pattern resumes unchanged after a transient variation "
                "before %s — not counted as a service change.",
                feed_label,
                new.start_week,
            )
            prev = new
            continue
        date = _change_date_between(services_by_date, prev, new)
        prev_ids = {sid for sid, _ in prev.pattern}
        new_ids = {sid for sid, _ in new.pattern}
        if not new_ids:
            kind = "Service suspended"
        elif not prev_ids:
            kind = "Service resumes"
        else:
            kind = "Service change"
        added = frozenset(new_ids - prev_ids)
        removed = frozenset(prev_ids - new_ids)
        note = "" if (added or removed) else "same service_ids on a new day-of-week pattern"
        events.append(ServiceChange(date, (feed_label,), kind, added, removed, note))
        prev = new
    return events, regimes


# ---- CROSS-FEED RECONCILIATION ----------------------------------------------


def _regime_in_effect(regimes: Sequence[Regime], week: dt.date) -> Regime:
    """Return the last stable regime starting on or before *week* (else the first)."""
    chosen = regimes[0]
    for regime in regimes:
        if regime.start_week <= week:
            chosen = regime
        else:
            break
    return chosen


def _boundary_event(
    prev_feed: FeedSummary,
    prev_regimes: Sequence[Regime],
    next_feed: FeedSummary,
    next_regimes: Sequence[Regime],
    coverage_gap_days: int = COVERAGE_GAP_DAYS,
) -> Optional[ServiceChange]:
    """Derive the service change (if any) implied by one feed superseding another.

    In a feed archive each markup is often published as a brand-new feed, so
    the succession itself — not any within-feed calendar row — marks the
    change date. Republished snapshots (same weekly pattern on both sides of
    the boundary) produce no event; a boundary where only the service_ids
    differ is reported with an explicit ambiguity note, because renamed ids
    and a genuinely new pick with identical structure are indistinguishable
    from calendars alone.

    Args:
        prev_feed: The earlier feed (by first active date).
        prev_regimes: Its stable regimes.
        next_feed: The later feed.
        next_regimes: Its stable regimes.
        coverage_gap_days: Max uncovered days still treated as contiguous.

    Returns:
        A :class:`ServiceChange` dated at *next_feed*'s first active date,
        or ``None`` when the boundary shows no service change.
    """
    if (
        prev_feed.first_active is None
        or prev_feed.last_active is None
        or next_feed.first_active is None
    ):
        return None
    if next_feed.first_active <= prev_feed.first_active:
        logging.info(
            "Feed '%s' does not start after feed '%s' — treating it as an overlapping "
            "snapshot, not a new service period.",
            next_feed.label,
            prev_feed.label,
        )
        return None
    if not prev_regimes or not next_regimes:
        logging.warning(
            "Cannot compare feeds '%s' and '%s' across their boundary — one of them is too "
            "short or too irregular to establish a stable weekly pattern.",
            prev_feed.label,
            next_feed.label,
        )
        return None

    next_ref = next_regimes[0]
    boundary_week = next_feed.first_active - dt.timedelta(days=next_feed.first_active.weekday())
    prev_ref = _regime_in_effect(prev_regimes, boundary_week)
    prev_ids = frozenset({sid for sid, _ in prev_ref.pattern})
    next_ids = frozenset({sid for sid, _ in next_ref.pattern})
    added = frozenset(next_ids - prev_ids)
    removed = frozenset(prev_ids - next_ids)
    evidence = (prev_feed.label, next_feed.label)

    gap_days = (next_feed.first_active - prev_feed.last_active).days - 1
    if gap_days > coverage_gap_days:
        logging.warning(
            "Coverage gap: feed '%s' has no service after %s and feed '%s' none before %s — "
            "%d day(s) are uncovered by the archive.",
            prev_feed.label,
            prev_feed.last_active,
            next_feed.label,
            next_feed.first_active,
            gap_days,
        )
        return ServiceChange(
            next_feed.first_active,
            evidence,
            "New service period",
            added,
            removed,
            note=f"begins after a {gap_days}-day gap in archive coverage",
        )
    if prev_ref.pattern == next_ref.pattern:
        logging.info(
            "Feed '%s' begins %s with the same weekly pattern feed '%s' already shows for "
            "that date — treating it as a republished snapshot, not a service change.",
            next_feed.label,
            next_feed.first_active,
            prev_feed.label,
        )
        return None
    if _weekly_signature(prev_ref.pattern) == _weekly_signature(next_ref.pattern):
        logging.warning(
            "Feed '%s' takes over from '%s' on %s with an identical day-of-week structure "
            "but different service_ids — probably a new pick (or renamed ids); review the "
            "feeds if %s is not a known change date.",
            next_feed.label,
            prev_feed.label,
            next_feed.first_active,
            next_feed.first_active,
        )
        return ServiceChange(
            next_feed.first_active,
            evidence,
            "New service period",
            added,
            removed,
            note="weekly day-of-week structure is unchanged — date marks where the newer "
            "feed takes over",
        )
    return ServiceChange(
        next_feed.first_active,
        evidence,
        "Service change",
        added,
        removed,
        note=f"feed '{next_feed.label}' supersedes '{prev_feed.label}'",
    )


def _cluster_events(
    events: Sequence[ServiceChange],
    first_active_by_feed: Mapping[str, dt.date],
    tolerance_days: int = _MERGE_TOLERANCE_DAYS,
) -> list[MergedChange]:
    """Merge events that fall within *tolerance_days* of each other.

    When several feeds evidence the same change on slightly different dates,
    the event from the newest feed (largest first active date) supplies the
    representative date and description, on the grounds that newer exports
    supersede older ones.
    """
    ordered = sorted(events, key=lambda e: (e.date, e.feeds))
    clusters: list[list[ServiceChange]] = []
    for event in ordered:
        if clusters and (event.date - clusters[-1][-1].date).days <= tolerance_days:
            clusters[-1].append(event)
        else:
            clusters.append([event])
    merged: list[MergedChange] = []
    for cluster in clusters:
        newest = max(
            cluster,
            key=lambda e: max(first_active_by_feed.get(f, dt.date.min) for f in e.feeds),
        )
        feeds = tuple(sorted({feed for event in cluster for feed in event.feeds}))
        notes = tuple(dict.fromkeys(event.note for event in cluster if event.note))
        merged.append(
            MergedChange(newest.date, newest.kind, newest.added, newest.removed, feeds, notes)
        )
    return merged


def _flag_disputes(
    merged: Sequence[MergedChange],
    summaries: Sequence[FeedSummary],
) -> list[MergedChange]:
    """Mark changes that other covering feeds fail to corroborate.

    A feed disputes a change when its active dates comfortably cover the
    change date yet it contributed no matching event. Overlapping snapshots
    legitimately disagree — an older feed keeps projecting its stale
    calendar past the pick it never knew about — so disputes are flagged and
    logged rather than dropped; the newer feed is usually the one to trust.
    """
    out: list[MergedChange] = []
    buffer_days = dt.timedelta(days=_DISPUTE_COVERAGE_BUFFER_DAYS)
    for change in merged:
        disputed: list[str] = []
        for summary in summaries:
            if summary.label in change.feeds:
                continue
            if summary.first_active is None or summary.last_active is None:
                continue
            if (
                summary.first_active <= change.date - buffer_days
                and summary.last_active >= change.date + buffer_days
            ):
                disputed.append(summary.label)
        if disputed:
            logging.warning(
                "Feeds disagree: a %s on %s is shown by %s, but feed(s) %s cover that date "
                "and show no change within a week of it. The newer feed is usually right; "
                "review the feeds if this date matters.",
                change.kind.lower(),
                change.date,
                ", ".join(change.feeds),
                ", ".join(disputed),
            )
            change = change._replace(disputed_by=tuple(disputed))
        out.append(change)
    return out


def merge_service_changes(
    summaries: Sequence[FeedSummary],
    min_stable_weeks: int = MIN_STABLE_WEEKS,
    coverage_gap_days: int = COVERAGE_GAP_DAYS,
) -> list[MergedChange]:
    """Detect changes in every usable feed and reconcile them across the archive.

    Args:
        summaries: Inspected feeds; unusable ones are skipped.
        min_stable_weeks: Weeks a new pattern must persist to count.
        coverage_gap_days: Max uncovered days between feeds still treated as
            a seamless hand-off.

    Returns:
        Chronological list of merged, dispute-flagged change events.
    """
    ordered = sorted(
        (s for s in summaries if s.usable and s.first_active is not None),
        key=lambda s: (s.first_active, s.label),
    )
    events: list[ServiceChange] = []
    regimes_by_feed: dict[str, list[Regime]] = {}
    for summary in ordered:
        feed_events, regimes = analyze_feed_changes(summary.active, min_stable_weeks, summary.label)
        events.extend(feed_events)
        regimes_by_feed[summary.label] = regimes
    for prev_feed, next_feed in zip(ordered, ordered[1:]):
        boundary = _boundary_event(
            prev_feed,
            regimes_by_feed[prev_feed.label],
            next_feed,
            regimes_by_feed[next_feed.label],
            coverage_gap_days,
        )
        if boundary is not None:
            events.append(boundary)
    first_active_by_feed = {s.label: s.first_active for s in ordered if s.first_active is not None}
    merged = _cluster_events(events, first_active_by_feed)
    return _flag_disputes(merged, ordered)


def apply_cutoffs(
    changes: Sequence[MergedChange],
    max_changes: Optional[int],
    max_years: Optional[float],
    anchor: dt.date,
) -> list[MergedChange]:
    """Trim the change list to the requested recency window.

    Args:
        changes: Merged change events (any order).
        max_changes: Keep only the most recent N changes, or ``None``.
        max_years: Keep only changes within the last X years of *anchor*,
            or ``None``.
        anchor: The newest active date in the archive — the reference point
            for *max_years*, so historical archives filter sensibly.

    Returns:
        Chronological list of the changes that survive both cutoffs.
    """
    kept = sorted(changes, key=lambda c: c.date)
    if max_years is not None:
        cutoff = anchor - dt.timedelta(days=round(365.25 * max_years))
        dropped = sum(1 for c in kept if c.date < cutoff)
        kept = [c for c in kept if c.date >= cutoff]
        if dropped:
            logging.info(
                "MAX_YEARS=%s: dropped %d change(s) before %s (anchored to the newest "
                "active date, %s).",
                max_years,
                dropped,
                cutoff,
                anchor,
            )
    if max_changes is not None and len(kept) > max_changes:
        logging.info(
            "MAX_CHANGES=%d: keeping the most recent %d of %d change(s).",
            max_changes,
            max_changes,
            len(kept),
        )
        kept = kept[len(kept) - max_changes :]
    return kept


# ---- OUTPUT TABLES AND EXPORT -----------------------------------------------


def _join_ids(ids: Iterable[str], limit: Optional[int] = None) -> str:
    """Join service_ids with ``'; '``, optionally truncating long lists for display."""
    ordered = sorted(ids)
    if limit is not None and len(ordered) > limit:
        return "; ".join(ordered[:limit]) + f"; +{len(ordered) - limit} more"
    return "; ".join(ordered)


def _display_id_list(joined: str, limit: int = 6) -> str:
    """Shorten a ``'; '``-joined service_id list for a printable cell."""
    ids = [part for part in joined.split("; ") if part]
    return _join_ids(ids, limit=limit)


def build_changes_table(changes: Sequence[MergedChange]) -> pd.DataFrame:
    """Build the one-row-per-change-date table for the quick reference.

    The table is machine-truthful: service_id and feed lists are complete
    and ``'; '``-joined (the XLSX export shortens long lists at display
    time only), and feeds that dispute a change sit in their own
    ``disputed_by`` column rather than inside free-text notes.

    Args:
        changes: Output of :func:`merge_service_changes` (post-cutoff).

    Returns:
        DataFrame of string columns, oldest change first — written verbatim
        to the changes CSV.
    """
    rows: list[dict[str, Any]] = []
    for change in sorted(changes, key=lambda c: c.date):
        rows.append(
            {
                "change_date": change.date.isoformat(),
                "day_of_week": _DAY_NAMES[change.date.weekday()],
                "change_type": change.kind,
                "services_added": _join_ids(change.added),
                "services_removed": _join_ids(change.removed),
                "source_feeds": "; ".join(change.feeds),
                "disputed_by": "; ".join(change.disputed_by),
                "notes": "; ".join(change.notes),
            }
        )
    columns = [
        "change_date",
        "day_of_week",
        "change_type",
        "services_added",
        "services_removed",
        "source_feeds",
        "disputed_by",
        "notes",
    ]
    return pd.DataFrame(rows, columns=columns)


def build_feeds_table(summaries: Sequence[FeedSummary]) -> pd.DataFrame:
    """Build the one-row-per-feed inventory table.

    Args:
        summaries: All inspected feeds, usable or not.

    Returns:
        DataFrame sorted by first active date (unusable feeds last).
    """
    ordered = sorted(summaries, key=lambda s: (s.first_active or dt.date.max, s.label))
    rows: list[dict[str, Any]] = []
    for summary in ordered:
        active_days = len(set().union(*summary.active.values())) if summary.active else 0
        rows.append(
            {
                "feed": summary.label,
                "agency_names": ", ".join(summary.agency_names),
                "agency_timezone": ", ".join(summary.timezones),
                "feed_publisher": summary.feed_publisher,
                "feed_version": summary.feed_version,
                "declared_start": summary.declared_start.isoformat()
                if summary.declared_start
                else "",
                "declared_end": summary.declared_end.isoformat() if summary.declared_end else "",
                "first_active_date": summary.first_active.isoformat()
                if summary.first_active
                else "",
                "last_active_date": summary.last_active.isoformat() if summary.last_active else "",
                "service_ids": len(summary.active),
                "active_days": active_days,
                "issues": "; ".join(summary.issues),
            }
        )
    columns = [
        "feed",
        "agency_names",
        "agency_timezone",
        "feed_publisher",
        "feed_version",
        "declared_start",
        "declared_end",
        "first_active_date",
        "last_active_date",
        "service_ids",
        "active_days",
        "issues",
    ]
    return pd.DataFrame(rows, columns=columns)


_CHANGE_COLUMN_WIDTHS: dict[str, int] = {
    "change_date": 13,
    "day_of_week": 12,
    "change_type": 18,
    "services_added": 26,
    "services_removed": 26,
    "source_feeds": 26,
    "disputed_by": 22,
    "notes": 42,
}

_FEED_COLUMN_WIDTHS: dict[str, int] = {
    "feed": 24,
    "agency_names": 26,
    "agency_timezone": 18,
    "feed_publisher": 20,
    "feed_version": 14,
    "declared_start": 14,
    "declared_end": 14,
    "first_active_date": 16,
    "last_active_date": 16,
    "service_ids": 12,
    "active_days": 12,
    "issues": 48,
}


def export_quick_reference_xlsx(
    changes_df: pd.DataFrame,
    feeds_df: pd.DataFrame,
    output_file: Path,
    title: str,
    subtitle: str,
) -> None:
    """Write the two-sheet, print-ready XLSX quick reference.

    The *Service Changes* sheet gets a merged title/subtitle banner, styled
    headers, zebra striping, and print settings (landscape, fit-to-width,
    banner and header repeated on every page); the *Feeds* sheet gets the
    same treatment without the banner.

    Args:
        changes_df: Output of :func:`build_changes_table` (may be empty).
        feeds_df: Output of :func:`build_feeds_table`.
        output_file: Destination ``.xlsx`` path (parents are created).
        title: Banner title, e.g. the system name.
        subtitle: Banner second line (generation date, scope).
    """
    output_file = Path(output_file)
    output_file.parent.mkdir(parents=True, exist_ok=True)

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    zebra_fill = PatternFill("solid", fgColor="F2F2F2")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="left", vertical="top", wrap_text=True)

    workbook = Workbook()
    changes_ws = workbook.active
    assert changes_ws is not None
    changes_ws.title = "Service Changes"

    n_cols = len(changes_df.columns)
    changes_ws.append([title])
    changes_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
    changes_ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    changes_ws.append([subtitle])
    changes_ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
    changes_ws.cell(row=2, column=1).font = Font(italic=True, color="808080")

    changes_ws.append([name.replace("_", " ").title() for name in changes_df.columns])
    for col_idx in range(1, n_cols + 1):
        cell = changes_ws.cell(row=3, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    wrap_cols = {"services_added", "services_removed", "source_feeds", "disputed_by", "notes"}
    center_cols = {"change_date", "day_of_week"}
    truncate_cols = {"services_added", "services_removed"}
    if changes_df.empty:
        changes_ws.append(["(no service changes detected in the analyzed window)"])
        changes_ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=n_cols)
        changes_ws.cell(row=4, column=1).font = Font(italic=True)
    else:
        for row in changes_df.itertuples(index=False):
            changes_ws.append(
                [
                    _display_id_list(str(value)) if name in truncate_cols else value
                    for name, value in zip(changes_df.columns, row)
                ]
            )
        for row_idx in range(4, changes_ws.max_row + 1):
            for col_idx, name in enumerate(changes_df.columns, start=1):
                cell = changes_ws.cell(row=row_idx, column=col_idx)
                if name in wrap_cols:
                    cell.alignment = wrap
                elif name in center_cols:
                    cell.alignment = center
                if row_idx % 2 == 0:
                    cell.fill = zebra_fill
        changes_ws.auto_filter.ref = f"A3:{get_column_letter(n_cols)}{changes_ws.max_row}"
    for col_idx, name in enumerate(changes_df.columns, start=1):
        letter = get_column_letter(col_idx)
        changes_ws.column_dimensions[letter].width = _CHANGE_COLUMN_WIDTHS.get(name, 14)
    changes_ws.freeze_panes = "A4"
    changes_ws.page_setup.orientation = changes_ws.ORIENTATION_LANDSCAPE
    changes_ws.page_setup.fitToWidth = 1
    changes_ws.page_setup.fitToHeight = 0
    changes_ws.sheet_properties.pageSetUpPr.fitToPage = True  # ty: ignore[invalid-assignment]
    changes_ws.print_title_rows = "1:3"

    feeds_ws = workbook.create_sheet("Feeds")
    feeds_ws.append([name.replace("_", " ").title() for name in feeds_df.columns])
    for col_idx in range(1, len(feeds_df.columns) + 1):
        cell = feeds_ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
    feed_wrap_cols = {"agency_names", "issues"}
    for row in feeds_df.itertuples(index=False):
        feeds_ws.append(list(row))
    for row_idx in range(2, feeds_ws.max_row + 1):
        for col_idx, name in enumerate(feeds_df.columns, start=1):
            cell = feeds_ws.cell(row=row_idx, column=col_idx)
            if name in feed_wrap_cols:
                cell.alignment = wrap
            if row_idx % 2 == 1:
                cell.fill = zebra_fill
    for col_idx, name in enumerate(feeds_df.columns, start=1):
        letter = get_column_letter(col_idx)
        feeds_ws.column_dimensions[letter].width = _FEED_COLUMN_WIDTHS.get(name, 14)
    feeds_ws.freeze_panes = "A2"
    feeds_ws.page_setup.orientation = feeds_ws.ORIENTATION_LANDSCAPE
    feeds_ws.page_setup.fitToWidth = 1
    feeds_ws.page_setup.fitToHeight = 0
    feeds_ws.sheet_properties.pageSetUpPr.fitToPage = True
    feeds_ws.print_title_rows = "1:1"

    workbook.save(str(output_file))
    logging.info(
        "Wrote %s (%d change(s), %d feed(s)).", output_file, len(changes_df), len(feeds_df)
    )


def write_run_log(output_dir: Path, summary_lines: List[str]) -> bool:
    """Write the verbatim config block plus a run summary into *output_dir*.

    Returns:
        ``True`` if the log was written successfully, ``False`` otherwise.
    """
    log_path = output_dir / "gtfs_service_change_dates_runlog.txt"
    try:
        config_text = extract_config_block(Path(__file__))
    except (OSError, ValueError) as exc:
        logging.error("Could not extract config block for run log: %s", exc)
        return False

    lines: List[str] = [
        "=" * 72,
        "GTFS SERVICE CHANGE DATES RUN LOG",
        "=" * 72,
        f"Run timestamp:    {datetime.now().isoformat(timespec='seconds')}",
        f"Output directory: {output_dir}",
        f"Source script:    {Path(__file__).resolve()}",
        "",
        "-" * 72,
        "RUN SUMMARY",
        "-" * 72,
        *summary_lines,
        "",
        "-" * 72,
        "CONFIGURATION (verbatim)",
        "-" * 72,
        "# === BEGIN CONFIG ===",
        config_text,
        "# === END CONFIG ===",
        "",
    ]
    try:
        log_path.write_text("\n".join(lines), encoding="utf-8")
    except OSError as exc:
        logging.error("Could not write run log '%s': %s", log_path, exc)
        return False
    logging.info("Run log written → %s", log_path)
    return True


# ---- ORCHESTRATION ----------------------------------------------------------


def run(
    feeds_dir: Path | None = None,
    output_dir: Path | None = None,
    max_changes: Optional[int] = None,
    max_years: Optional[float] = None,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """Scan the archive, reconcile service changes, and write the quick reference.

    Writes the printable XLSX, its machine-readable CSV twins, and the
    run-log sidecar. Unset args fall back to the config block at the top of
    this file, so ``m.FEEDS_DIR = ...; m.run()`` works after a plain import.

    Args:
        feeds_dir: Folder of GTFS feeds (zips and/or subfolders).
        output_dir: Folder for the XLSX, CSVs, and run-log sidecar.
        max_changes: Keep only the most recent N changes (``None`` = config).
        max_years: Keep only the last X years of changes (``None`` = config).

    Returns:
        Tuple of (service changes table, feed inventory table).

    Raises:
        OSError: If *feeds_dir* is missing, or the run log is required but
            cannot be written.
        ValueError: If no usable GTFS feed is found in *feeds_dir*.
    """
    feeds_dir = FEEDS_DIR if feeds_dir is None else Path(feeds_dir)
    output_dir = OUTPUT_DIR if output_dir is None else Path(output_dir)
    max_changes = MAX_CHANGES if max_changes is None else max_changes
    max_years = MAX_YEARS if max_years is None else max_years

    feed_paths = discover_feeds(feeds_dir)
    summaries = [inspect_feed(path) for path in feed_paths]
    usable = [s for s in summaries if s.usable]
    if not usable:
        raise ValueError(
            f"None of the {len(summaries)} feed(s) in '{feeds_dir}' contained usable "
            "calendar data — nothing to analyze."
        )
    system_name = check_system_consistency(summaries)

    changes = merge_service_changes(usable, MIN_STABLE_WEEKS, COVERAGE_GAP_DAYS)
    first_dates = [s.first_active for s in usable if s.first_active is not None]
    last_dates = [s.last_active for s in usable if s.last_active is not None]
    anchor = max(last_dates)
    changes = apply_cutoffs(changes, max_changes, max_years, anchor)

    if changes:
        for change in changes:
            logging.info(
                "%s (%s): %s — added: %s; removed: %s [source: %s]",
                change.date,
                _DAY_NAMES[change.date.weekday()],
                change.kind,
                _join_ids(change.added, limit=6) or "-",
                _join_ids(change.removed, limit=6) or "-",
                ", ".join(change.feeds),
            )
    else:
        logging.info("No service changes detected within the analyzed window.")

    changes_df = build_changes_table(changes)
    feeds_df = build_feeds_table(summaries)

    output_dir.mkdir(parents=True, exist_ok=True)
    out_path = output_dir / OUTPUT_FILENAME
    scope_bits: list[str] = []
    if max_years is not None:
        scope_bits.append(f"last {max_years:g} year(s) of service")
    if max_changes is not None:
        scope_bits.append(f"most recent {max_changes} change(s)")
    subtitle = (
        f"Generated {dt.date.today().isoformat()} from {len(usable)} GTFS feed(s) | "
        f"Active service {min(first_dates).isoformat()} → {anchor.isoformat()} | "
        f"Scope: {'; '.join(scope_bits) or 'full archive'}"
    )
    export_quick_reference_xlsx(
        changes_df,
        feeds_df,
        out_path,
        title=f"Service Change Quick Reference — {system_name}",
        subtitle=subtitle,
    )

    changes_csv_path = output_dir / CHANGES_CSV_FILENAME
    feeds_csv_path = output_dir / FEEDS_CSV_FILENAME
    changes_df.to_csv(changes_csv_path, index=False)
    feeds_df.to_csv(feeds_csv_path, index=False)
    logging.info("Wrote machine-readable CSVs → %s, %s", changes_csv_path, feeds_csv_path)

    summary_lines = [
        f"Feeds folder:     {feeds_dir}",
        f"Feeds discovered: {len(summaries)} ({len(usable)} usable)",
        f"Service changes:  {len(changes_df)}",
        f"Cutoffs:          max_changes={max_changes}, max_years={max_years}",
        f"Quick reference:  {out_path}",
        f"Changes CSV:      {changes_csv_path}",
        f"Feeds CSV:        {feeds_csv_path}",
    ]
    if not write_run_log(output_dir, summary_lines) and REQUIRE_RUN_LOG:
        raise OSError(
            f"Run log could not be written to '{output_dir}' and REQUIRE_RUN_LOG is True."
        )

    logging.info(
        "Service change quick reference complete — %d change(s) across %d usable feed(s).",
        len(changes_df),
        len(usable),
    )
    return changes_df, feeds_df


# =============================================================================
# MAIN
# =============================================================================


def build_arg_parser() -> argparse.ArgumentParser:
    """Build the CLI parser; every flag defaults to its CONFIGURATION constant."""
    parser = argparse.ArgumentParser(
        description=(
            "Scan a folder of GTFS feeds (zipped or unzipped), reconcile the dates the "
            "service actually changed, and export a printable XLSX quick reference plus "
            "machine-readable CSVs. Defaults come from the configuration block at the "
            "top of this file."
        ),
        formatter_class=argparse.ArgumentDefaultsHelpFormatter,
    )
    parser.add_argument(
        "--feeds-dir",
        type=Path,
        default=FEEDS_DIR,
        help="Folder to scan for GTFS feeds (zip archives and/or subfolders).",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=OUTPUT_DIR,
        help="Folder for the XLSX quick reference and its run log.",
    )
    parser.add_argument(
        "--max-changes",
        type=int,
        default=MAX_CHANGES,
        help="Keep only the most recent N service changes.",
    )
    parser.add_argument(
        "--max-years",
        type=float,
        default=MAX_YEARS,
        help="Keep only changes within the last X years of the newest active date.",
    )
    parser.add_argument(
        "--log-level",
        default=logging.getLevelName(LOG_LEVEL),
        help="DEBUG / INFO / WARNING / ERROR.",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    """Command-line entry point. Defaults fall back to the config block.

    Returns:
        Process exit code: 0 on success, 1 on failure, 2 if required
        CONFIGURATION values are still placeholders.
    """
    args = build_arg_parser().parse_args(notebook_safe_argv(argv))
    logging.basicConfig(
        level=getattr(logging, str(args.log_level).upper(), LOG_LEVEL),
        format="%(asctime)s | %(levelname)s | %(message)s",
        datefmt="%H:%M:%S",
    )
    sentinels = {Path(r"Path\To\Your\GTFS_Feeds_Folder"), Path(r"Path\To\Your\Output_Folder")}
    if args.feeds_dir in sentinels or args.output_dir in sentinels:
        logging.warning(
            "FEEDS_DIR and/or OUTPUT_DIR are still placeholders. Update the configuration "
            "block or pass --feeds-dir/--output-dir before running."
        )
        return 2
    try:
        run(
            feeds_dir=args.feeds_dir,
            output_dir=args.output_dir,
            max_changes=args.max_changes,
            max_years=args.max_years,
        )
    except (OSError, ValueError) as exc:
        logging.error("%s", exc)
        return 1
    return 0


# Strict parsing; in a notebook, notebook_safe_argv() keeps the kernel's
# injected argv away from argparse so the config block stays in charge.
if __name__ == "__main__":
    raise SystemExit(main())
