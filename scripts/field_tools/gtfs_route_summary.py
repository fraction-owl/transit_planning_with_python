"""Export a one-row-per-route GTFS desk reference spreadsheet for transit planners.

Typical usage
-------------
    python gtfs_route_summary.py --gtfs-dir ./feed --out-dir ./out

Key Features
------------
* Reads a GTFS feed and classifies each service_id as Weekday / Saturday /
  Sunday / Holiday based on the real active-date calendar (calendar.txt +
  calendar_dates.txt exceptions).
* Computes per-route variants, directions, average trip distance, duration,
  and speed.
* Optionally joins service-type, corridor, last-changed, and ridership
  lookups keyed on route_id.
* Exports a formatted, print-ready XLSX wall chart suitable for field use.
"""

from __future__ import annotations

import argparse
import logging
import os
from collections.abc import Mapping, Sequence
from datetime import datetime, timedelta
from typing import Any, Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ==== CONFIGURATION ==========================================================

_DEFAULT_GTFS_FOLDER_PATH = r"Path\To\Your\GTFS_Folder"
_DEFAULT_BASE_OUTPUT_PATH = r"Path\To\Your\Output_Folder"

GTFS_FOLDER_PATH = _DEFAULT_GTFS_FOLDER_PATH  # <<< EDIT HERE
BASE_OUTPUT_PATH = _DEFAULT_BASE_OUTPUT_PATH  # <<< EDIT HERE
OUTPUT_FILENAME = "routes_summary.xlsx"
DISTANCE_UNIT = "meters"  # meters | kilometers | feet | miles
OUTPUT_UNITS = "imperial"  # imperial (mi/mph) or metric (km/kmh)
EXCLUDED_ROUTE_SHORT_NAMES = ["9999A", "9999B", "9999C"]

SERVICE_TYPES_PATH = ""
CORRIDORS_PATH = ""
LAST_CHANGED_PATH = ""
RIDERSHIP_PATH = ""

REQUIRED_GTFS_FILES = ["routes.txt", "trips.txt", "stop_times.txt", "calendar.txt"]

# -----------------------------------------------------------------------------
# Classification thresholds
# -----------------------------------------------------------------------------
# HOLIDAY_MAX_DAYS_PER_YEAR controls when a service_id is treated as a
# "Holiday" pattern rather than a regular weekly pattern. The classifier
# computes the number of real active dates per year (after applying
# calendar_dates.txt exceptions) and, if that rate is at or below this
# threshold, labels the service as Holiday only. The default of 25 comfortably
# covers the ~10-15 federal/observed holidays most North American agencies
# run, with headroom for extras like the day after Thanksgiving, Christmas
# Eve, New Year's Eve, and a handful of agency-specific special days.
#
# Symptoms you should RAISE this value: a known holiday-only service is being
# labeled "Weekday" (or similar) because it happens to be active on, say, 30
# scattered days a year. Symptoms you should LOWER it: a thin seasonal/summer
# weekday service with only ~20 operating days is being lumped in with true
# holidays. When in doubt, inspect the INFO logs this script emits — each
# service_id prints its active-date count and per-year rate.
#
# WEEKDAY_DOW_SHARE controls how dominant a day-of-week bucket must be to
# earn its label. For each service we compute the share of active dates that
# fall on M-F (and independently on Saturday, and on Sunday). If the share
# is at or above this value, the bucket is added. 0.80 is strict enough to
# keep "mostly weekday" services from leaking into Saturday/Sunday labels
# when they run an occasional weekend, while still tolerating a handful of
# holiday cancellations on an otherwise-pure weekday pattern. Raise it
# (e.g. 0.90) if you see weekend labels on services that only run rare
# weekend specials; lower it (e.g. 0.70) if a service that legitimately
# runs most weekends is getting only "Weekday".
HOLIDAY_MAX_DAYS_PER_YEAR: float = 25
WEEKDAY_DOW_SHARE: float = 0.80

# ==== FUNCTIONS ==============================================================

# ---- REUSABLE HELPERS (copied from utils/gtfs_helpers.py) ------------------


def load_gtfs_data(
    gtfs_folder_path: str,
    files: Optional[Sequence[str]] = None,
    dtype: str | type[str] | Mapping[str, Any] = str,
) -> dict[str, pd.DataFrame]:
    """Load one or more GTFS text files into memory.

    Args:
        gtfs_folder_path: Absolute or relative path to the folder
            containing the GTFS feed.
        files: Explicit sequence of file names to load. If ``None``,
            the standard 13 GTFS text files are attempted.
        dtype: Value forwarded to :pyfunc:`pandas.read_csv(dtype=…)` to
            control column dtypes. Supply a mapping for per-column dtypes.

    Returns:
        Mapping of file stem → :class:`pandas.DataFrame`; for example,
        ``data["trips"]`` holds the parsed *trips.txt* table.

    Raises:
        OSError: Folder missing or one of *files* not present.
        ValueError: Empty file or CSV parser failure.
        RuntimeError: Generic OS error while reading a file.

    Notes:
        All columns default to ``str`` to avoid pandas' type-inference
        pitfalls (e.g. leading zeros in IDs).
    """
    if not os.path.exists(gtfs_folder_path):
        raise OSError(f"The directory '{gtfs_folder_path}' does not exist.")

    if files is None:
        files = (
            "agency.txt",
            "stops.txt",
            "routes.txt",
            "trips.txt",
            "stop_times.txt",
            "calendar.txt",
            "calendar_dates.txt",
            "fare_attributes.txt",
            "fare_rules.txt",
            "feed_info.txt",
            "frequencies.txt",
            "shapes.txt",
            "transfers.txt",
        )

    missing = [
        file_name
        for file_name in files
        if not os.path.exists(os.path.join(gtfs_folder_path, file_name))
    ]
    if missing:
        raise OSError(f"Missing GTFS files in '{gtfs_folder_path}': {', '.join(missing)}")

    data: dict[str, pd.DataFrame] = {}
    for file_name in files:
        key = file_name.replace(".txt", "")
        file_path = os.path.join(gtfs_folder_path, file_name)
        try:
            df = pd.read_csv(file_path, dtype=dtype, low_memory=False)
            data[key] = df
            logging.info("Loaded %s (%d records).", file_name, len(df))

        except pd.errors.EmptyDataError as exc:
            raise ValueError(f"File '{file_name}' in '{gtfs_folder_path}' is empty.") from exc

        except pd.errors.ParserError as exc:
            raise ValueError(
                f"Parser error in '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

        except OSError as exc:
            raise RuntimeError(
                f"OS error reading file '{file_name}' in '{gtfs_folder_path}': {exc}"
            ) from exc

    return data


# ---- SCRIPT FUNCTIONS -------------------------------------------------------


def hms_to_seconds(time_str: str) -> Optional[int]:
    """Convert a GTFS ``HH:MM:SS`` string to seconds since service start.

    GTFS allows hours >= 24 to represent trips that span midnight; this
    function preserves that overflow rather than taking a modulo.

    Args:
        time_str: Time string in ``HH:MM:SS`` form.

    Returns:
        Integer seconds, or ``None`` if the input is missing/invalid.
    """
    if time_str is None or (isinstance(time_str, float) and pd.isna(time_str)):
        return None
    try:
        h, m, s = str(time_str).strip().split(":")
        return int(h) * 3600 + int(m) * 60 + int(s)
    except (ValueError, AttributeError):
        return None


def classify_services(
    calendar_df: pd.DataFrame,
    calendar_dates_df: Optional[pd.DataFrame],
    holiday_max_days_per_year: float = HOLIDAY_MAX_DAYS_PER_YEAR,
    weekday_dow_share: float = WEEKDAY_DOW_SHARE,
) -> dict[str, set[str]]:
    """Classify each service_id by its real active-date pattern.

    Args:
        calendar_df: Parsed ``calendar.txt``.
        calendar_dates_df: Parsed ``calendar_dates.txt`` or ``None``.
        holiday_max_days_per_year: Services active at or below this annual
            rate are labelled ``Holiday``.
        weekday_dow_share: Minimum fraction of active dates on a given
            day-of-week bucket to earn that bucket's label.

    Returns:
        Mapping of ``service_id`` to a set of labels drawn from
        ``{"Weekday", "Saturday", "Sunday", "Holiday"}``.
    """
    dow_cols = ["monday", "tuesday", "wednesday", "thursday", "friday", "saturday", "sunday"]
    result: dict[str, set[str]] = {}

    # Build base active-date set from calendar.txt
    base: dict[str, set[datetime]] = {}
    for _, row in calendar_df.iterrows():
        sid = row["service_id"]
        try:
            start = datetime.strptime(str(row["start_date"]), "%Y%m%d")
            end = datetime.strptime(str(row["end_date"]), "%Y%m%d")
        except ValueError:
            base[sid] = set()
            continue
        pattern = [int(str(row.get(c, "0")) or "0") == 1 for c in dow_cols]
        dates: set[datetime] = set()
        d = start
        while d <= end:
            if pattern[d.weekday()]:
                dates.add(d)
            d += timedelta(days=1)
        base[sid] = dates

    # Apply calendar_dates exceptions
    if calendar_dates_df is not None and not calendar_dates_df.empty:
        for _, row in calendar_dates_df.iterrows():
            sid = row["service_id"]
            try:
                d = datetime.strptime(str(row["date"]), "%Y%m%d")
            except ValueError:
                continue
            etype = str(row.get("exception_type", "")).strip()
            dates = base.setdefault(sid, set())
            if etype == "1":
                dates.add(d)
            elif etype == "2":
                dates.discard(d)

    for sid, dates in base.items():
        if not dates:
            result[sid] = set()
            logging.info("Service %s: empty (0 active dates)", sid)
            continue
        span_days = (max(dates) - min(dates)).days + 1
        per_year = len(dates) / max(span_days / 365.25, 0.1)
        labels: set[str] = set()
        if per_year <= holiday_max_days_per_year:
            labels.add("Holiday")
        else:
            n = len(dates)
            wd = sum(1 for d in dates if d.weekday() < 5)
            sat = sum(1 for d in dates if d.weekday() == 5)
            sun = sum(1 for d in dates if d.weekday() == 6)
            if wd / n >= weekday_dow_share:
                labels.add("Weekday")
            if sat / n >= weekday_dow_share:
                labels.add("Saturday")
            if sun / n >= weekday_dow_share:
                labels.add("Sunday")
            if not labels:
                if wd > 0:
                    labels.add("Weekday")
                if sat > 0:
                    labels.add("Saturday")
                if sun > 0:
                    labels.add("Sunday")
        result[sid] = labels
        logging.info(
            "Service %s -> %s (%s dates, %.1f/yr)",
            sid,
            sorted(labels),
            len(dates),
            per_year,
        )
    return result


def trip_distances_meters(
    stop_times_df: pd.DataFrame,
    shapes_df: Optional[pd.DataFrame],
    trips_df: pd.DataFrame,
    distance_unit: str,
) -> pd.Series:
    """Return a ``trip_id`` -> meters Series for all trips that have geometry.

    Args:
        stop_times_df: Parsed ``stop_times.txt``.
        shapes_df: Parsed ``shapes.txt`` or ``None``.
        trips_df: Parsed ``trips.txt``.
        distance_unit: Unit of ``shape_dist_traveled`` in the feed.

    Returns:
        Series indexed by ``trip_id`` with distances in metres.
    """
    factors: dict[str, float] = {
        "meters": 1.0,
        "kilometers": 1000.0,
        "feet": 0.3048,
        "miles": 1609.344,
    }
    factor = factors.get(distance_unit, 1.0)

    if "shape_dist_traveled" in stop_times_df.columns:
        s = pd.to_numeric(stop_times_df["shape_dist_traveled"], errors="coerce")
        if s.notna().any():
            tmp = stop_times_df.assign(_d=s).dropna(subset=["_d"])
            per_trip = tmp.groupby("trip_id")["_d"].max() * factor
            if not per_trip.empty:
                return per_trip

    if shapes_df is not None and not shapes_df.empty and "shape_dist_traveled" in shapes_df.columns:
        s = pd.to_numeric(shapes_df["shape_dist_traveled"], errors="coerce")
        tmp = shapes_df.assign(_d=s).dropna(subset=["_d"])
        per_shape = tmp.groupby("shape_id")["_d"].max() * factor
        if "shape_id" in trips_df.columns:
            return trips_df.set_index("trip_id")["shape_id"].map(per_shape).dropna()

    return pd.Series(dtype=float)


def trip_durations_seconds(stop_times_df: pd.DataFrame) -> pd.Series:
    """Return a ``trip_id`` -> duration-in-seconds Series.

    Args:
        stop_times_df: Parsed ``stop_times.txt``.

    Returns:
        Series indexed by ``trip_id`` with trip durations in seconds.
    """
    df = stop_times_df.copy()
    df["_seq"] = pd.to_numeric(df["stop_sequence"], errors="coerce")
    df = df.sort_values(["trip_id", "_seq"])
    grp = df.groupby("trip_id")
    first_dep = grp["departure_time"].first().map(hms_to_seconds)
    last_arr = grp["arrival_time"].last().map(hms_to_seconds)
    return (last_arr - first_dep).dropna()


def load_optional_lookup(path: str, value_col: str) -> dict[str, str]:
    """Load an optional route_id -> value lookup from CSV or TSV.

    Args:
        path: Filesystem path to a CSV or TSV file.
        value_col: Name of the column to use as the value.

    Returns:
        Mapping of ``route_id`` to the value in *value_col*, or an empty
        dict if the file is absent or cannot be parsed.
    """
    if not path:
        return {}
    if not os.path.isfile(path):
        logging.warning("Optional lookup not found: %s", path)
        return {}
    try:
        with open(path, "r", encoding="utf-8") as fh:
            first = fh.readline()
        sep = "\t" if "\t" in first else ","
        df = pd.read_csv(path, sep=sep, dtype=str)
    except (OSError, ValueError, pd.errors.ParserError) as exc:
        logging.warning("Could not read lookup %s: %s", path, exc)
        return {}
    if "route_id" not in df.columns or value_col not in df.columns:
        logging.warning("Lookup %s missing route_id or %s column", path, value_col)
        return {}
    return dict(zip(df["route_id"], df[value_col]))


def build_summary(
    routes_df: pd.DataFrame,
    trips_df: pd.DataFrame,
    stop_times_df: pd.DataFrame,
    calendar_df: pd.DataFrame,
    calendar_dates_df: Optional[pd.DataFrame],
    shapes_df: Optional[pd.DataFrame],
    distance_unit: str,
    extras: Mapping[str, Mapping[str, str]],
    holiday_max_days_per_year: float = HOLIDAY_MAX_DAYS_PER_YEAR,
    weekday_dow_share: float = WEEKDAY_DOW_SHARE,
) -> pd.DataFrame:
    """Assemble the per-route summary DataFrame.

    Args:
        routes_df: Parsed ``routes.txt``.
        trips_df: Parsed ``trips.txt``.
        stop_times_df: Parsed ``stop_times.txt``.
        calendar_df: Parsed ``calendar.txt``.
        calendar_dates_df: Parsed ``calendar_dates.txt`` or ``None``.
        shapes_df: Parsed ``shapes.txt`` or ``None``.
        distance_unit: Unit of ``shape_dist_traveled`` in the feed.
        extras: Optional lookup dicts keyed by category then ``route_id``.
        holiday_max_days_per_year: Forwarded to :func:`classify_services`.
        weekday_dow_share: Forwarded to :func:`classify_services`.

    Returns:
        One-row-per-route :class:`pandas.DataFrame` ready for export.
    """
    service_labels = classify_services(
        calendar_df,
        calendar_dates_df,
        holiday_max_days_per_year=holiday_max_days_per_year,
        weekday_dow_share=weekday_dow_share,
    )
    dist_m = trip_distances_meters(stop_times_df, shapes_df, trips_df, distance_unit)
    dur_s = trip_durations_seconds(stop_times_df)

    t = trips_df.copy()
    t["_dist_m"] = t["trip_id"].map(dist_m)
    t["_dur_s"] = t["trip_id"].map(dur_s)

    excluded = {str(x) for x in (EXCLUDED_ROUTE_SHORT_NAMES or [])}
    imperial = str(OUTPUT_UNITS).lower() == "imperial"
    dist_col = "avg_distance_mi" if imperial else "avg_distance_km"
    speed_col = "avg_speed_mph" if imperial else "avg_speed_kmh"

    rows = []
    for _, r in routes_df.iterrows():
        rid = r["route_id"]
        if str(r.get("route_short_name", "")) in excluded:
            logging.info("Excluding route %s", r.get("route_short_name", ""))
            continue
        rt = t[t["route_id"] == rid]
        if rt.empty:
            continue

        if "shape_id" in rt.columns and rt["shape_id"].notna().any():
            variants = int(rt["shape_id"].nunique())
        elif "trip_headsign" in rt.columns and rt["trip_headsign"].notna().any():
            variants = int(rt["trip_headsign"].nunique())
        else:
            variants = 1
        variants = max(variants, 1)

        directions = int(rt["direction_id"].nunique()) if "direction_id" in rt.columns else 1
        directions = max(directions, 1)

        day_cats: set[str] = set()
        for sid in rt["service_id"].dropna().unique():
            day_cats |= service_labels.get(sid, set())

        avg_dist_m = rt["_dist_m"].dropna().mean()
        avg_dur_s = rt["_dur_s"].dropna().mean()
        if pd.notna(avg_dist_m):
            avg_distance: Optional[float] = round(
                avg_dist_m / (1609.344 if imperial else 1000.0), 2
            )
        else:
            avg_distance = None
        avg_duration_min: Optional[float] = (
            round(avg_dur_s / 60.0, 1) if pd.notna(avg_dur_s) else None
        )
        if pd.notna(avg_dist_m) and pd.notna(avg_dur_s) and avg_dur_s > 0:
            mps = avg_dist_m / avg_dur_s
            avg_speed: Optional[float] = round(mps * (2.23694 if imperial else 3.6), 1)
        else:
            avg_speed = None

        rows.append(
            {
                "route_short_name": r.get("route_short_name", ""),
                "route_long_name": r.get("route_long_name", ""),
                "variants": variants,
                "directions": directions,
                "weekday": "Y" if "Weekday" in day_cats else "",
                "saturday": "Y" if "Saturday" in day_cats else "",
                "sunday": "Y" if "Sunday" in day_cats else "",
                "holiday": "Y" if "Holiday" in day_cats else "",
                dist_col: avg_distance,
                "avg_duration_min": avg_duration_min,
                speed_col: avg_speed,
                "service_type": extras.get("service_types", {}).get(rid, ""),
                "corridor": extras.get("corridors", {}).get(rid, ""),
                "last_changed": extras.get("last_changed", {}).get(rid, ""),
                "ridership": extras.get("ridership", {}).get(rid, ""),
            }
        )

    cols = [
        "route_short_name",
        "route_long_name",
        "variants",
        "directions",
        "weekday",
        "saturday",
        "sunday",
        "holiday",
        dist_col,
        "avg_duration_min",
        speed_col,
        "service_type",
        "corridor",
        "last_changed",
        "ridership",
    ]
    return pd.DataFrame(rows, columns=cols)


def export_to_xlsx(data_frame: pd.DataFrame, output_file: str) -> None:
    """Write the summary to a formatted XLSX wall chart.

    Freezes the header row, bolds headers, widens columns based on
    content, wraps ``route_long_name``, and enables print-friendly
    settings (landscape, fit-to-width, repeat header on every page).

    Args:
        data_frame: Summary built by :func:`build_summary`.
        output_file: Destination ``.xlsx`` path.
    """
    if data_frame is None or data_frame.empty:
        logging.info("No rows to export; skipping write.")
        return
    if not output_file.lower().endswith(".xlsx"):
        output_file = os.path.splitext(output_file)[0] + ".xlsx"
    os.makedirs(os.path.dirname(output_file) or ".", exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Routes"

    headers = list(data_frame.columns)
    ws.append(headers)
    for row in data_frame.itertuples(index=False):
        ws.append(list(row))

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="305496")
    center = Alignment(horizontal="center", vertical="center")
    wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)

    for col_idx, _name in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    widths: dict[str, Union[int, float]] = {
        "route_short_name": 12,
        "route_long_name": 40,
        "variants": 9,
        "directions": 10,
        "weekday": 9,
        "saturday": 9,
        "sunday": 8,
        "holiday": 9,
        "avg_distance_mi": 14,
        "avg_distance_km": 14,
        "avg_duration_min": 15,
        "avg_speed_mph": 13,
        "avg_speed_kmh": 13,
        "service_type": 14,
        "corridor": 16,
        "last_changed": 14,
        "ridership": 12,
    }
    for col_idx, name in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = widths.get(name, 12)

    long_col = headers.index("route_long_name") + 1 if "route_long_name" in headers else None
    for row_idx in range(2, ws.max_row + 1):
        if long_col is not None:
            ws.cell(row=row_idx, column=long_col).alignment = wrap
        ws.row_dimensions[row_idx].height = 22

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_title_rows = "1:1"

    wb.save(output_file)
    logging.info("Wrote %s (%s rows)", output_file, len(data_frame))


def _parse_cli_args() -> argparse.Namespace:
    """Parse command-line overrides for every CONFIGURATION constant.

    Returns:
        Parsed argument namespace.
    """
    p = argparse.ArgumentParser(description="Export a one-row-per-route GTFS summary XLSX.")
    p.add_argument("--gtfs-dir", default=None)
    p.add_argument("--out-dir", default=None)
    p.add_argument("--out-file", default=None)
    p.add_argument(
        "--distance-unit", default=None, choices=["meters", "kilometers", "feet", "miles"]
    )
    p.add_argument("--service-types", default=None)
    p.add_argument("--corridors", default=None)
    p.add_argument("--last-changed", default=None)
    p.add_argument("--ridership", default=None)
    p.add_argument("--holiday-max-days-per-year", type=float, default=None)
    p.add_argument("--weekday-dow-share", type=float, default=None)
    args, unknown = p.parse_known_args()
    if unknown:
        logging.debug("Ignoring unknown args (e.g. Jupyter kernel): %s", unknown)
    return args


# ==== MAIN ===================================================================


def main() -> None:
    """Entry point: load feed, build summary, write XLSX."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
    )
    args = _parse_cli_args()

    gtfs_dir = args.gtfs_dir or GTFS_FOLDER_PATH
    out_dir = args.out_dir or BASE_OUTPUT_PATH
    out_file = args.out_file or OUTPUT_FILENAME
    dist_unit = args.distance_unit or DISTANCE_UNIT
    svc_types_path = args.service_types if args.service_types is not None else SERVICE_TYPES_PATH
    corridors_path = args.corridors if args.corridors is not None else CORRIDORS_PATH
    last_changed_path = args.last_changed if args.last_changed is not None else LAST_CHANGED_PATH
    ridership_path = args.ridership if args.ridership is not None else RIDERSHIP_PATH
    holiday_max = (
        args.holiday_max_days_per_year
        if args.holiday_max_days_per_year is not None
        else HOLIDAY_MAX_DAYS_PER_YEAR
    )
    weekday_share = (
        args.weekday_dow_share if args.weekday_dow_share is not None else WEEKDAY_DOW_SHARE
    )

    logging.info("==== GTFS Route Summary ====")
    logging.info("GTFS dir      : %s", gtfs_dir)
    logging.info("Output        : %s", os.path.join(out_dir, out_file))
    logging.info("Distance unit : %s", dist_unit)
    logging.info("Holiday max/y : %s", holiday_max)
    logging.info("Weekday share : %s", weekday_share)

    try:
        core = load_gtfs_data(gtfs_dir, files=REQUIRED_GTFS_FILES)

        calendar_dates_df: Optional[pd.DataFrame] = None
        try:
            cd = load_gtfs_data(gtfs_dir, files=("calendar_dates.txt",))
            calendar_dates_df = cd.get("calendar_dates")
        except OSError as exc:
            logging.warning("calendar_dates.txt unavailable: %s", exc)

        shapes_df: Optional[pd.DataFrame] = None
        try:
            sh = load_gtfs_data(gtfs_dir, files=("shapes.txt",))
            shapes_df = sh.get("shapes")
        except OSError as exc:
            logging.warning("shapes.txt unavailable: %s", exc)

        extras: dict[str, dict[str, str]] = {
            "service_types": load_optional_lookup(svc_types_path, "service_type"),
            "corridors": load_optional_lookup(corridors_path, "corridor"),
            "last_changed": load_optional_lookup(last_changed_path, "last_changed"),
            "ridership": load_optional_lookup(ridership_path, "ridership"),
        }

        summary = build_summary(
            routes_df=core["routes"],
            trips_df=core["trips"],
            stop_times_df=core["stop_times"],
            calendar_df=core["calendar"],
            calendar_dates_df=calendar_dates_df,
            shapes_df=shapes_df,
            distance_unit=dist_unit,
            extras=extras,
            holiday_max_days_per_year=holiday_max,
            weekday_dow_share=weekday_share,
        )

        if summary.empty:
            logging.warning("Summary is empty; nothing to write.")
            return

        export_to_xlsx(summary, os.path.join(out_dir, out_file))

    except (OSError, ValueError, RuntimeError) as exc:
        logging.error("Pipeline failed: %s", exc)
    except Exception:
        logging.exception("Unexpected error in pipeline")
    finally:
        logging.info("Exiting script.")


if __name__ == "__main__":
    main()
